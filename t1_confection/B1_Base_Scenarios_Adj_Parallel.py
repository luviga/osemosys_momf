# -*- coding: utf-8 -*-
# Suggested Citation: 
# Victor-Gallardo, L. (2022). Robust Energy System Planning for Decarbonization under Technological Uncertainty. 
# Retrieved from https://www.kerwa.ucr.ac.cr/handle/10669/87273

"""
This code is based on the methodologies and findings discussed in the work of Luis Victor-Gallardo, specifically focusing on robust energy system planning for decarbonization under technological uncertainty. The document provides a comprehensive analysis relevant to the code's development and application. For in-depth understanding and further details, please refer to the original document linked below.

Suggested Citation:
Victor-Gallardo, L. (2022). Robust Energy System Planning for Decarbonization under Technological Uncertainty. 
University of Costa Rica. Retrieved from https://www.kerwa.ucr.ac.cr/handle/10669/87273
"""
import errno
import scipy
import pandas as pd
import numpy as np
import xlrd
import re
import csv
import os, os.path
import sys
import math
import linecache
from copy import deepcopy
import time
import gc
import shutil
import pickle
import multiprocessing as mp
import yaml
import openpyxl
from setup_utils import install_requirements
import subprocess
import platform

'''
We implement OSEMOSYS-CR in a csv based system for semi-automatic manipulation of parameters.
'''
def intersection_2(lst1, lst2): 
    return list(set(lst1) & set(lst2))
#
def interp_max_cap( x ):
    x_change_known = []
    x_unknown = []
    last_big_known = 0
    x_pick = [ x[0] ]
    for n in range( 1, len( x ) ):
        if x[n] > x[n-1] and x[n-1] >= last_big_known:
            last_big_known = x[n-1]
            appender = x[n]
            x_change_known.append( x[n] )
            x_unknown.append( 'none' )
            x_pick.append( x[n] )
        else:
            if x[n] >= last_big_known and x[n] > x_change_known[0]:
                x_change_known.append( x[n] )
                x_unknown.append( 'none' )
                x_pick.append( x[n] )
            else:
                x_change_known.append( 'none' )
                x_unknown.append( x[n] )
                x_pick.append( appender )
    return x_pick
    #
#
def check_enviro_variables(solver_command):
    # Determine the command based on the operating system
    command = 'where' if platform.system() == 'Windows' else 'which'
    
    # Execute the appropriate command
    where_solver = subprocess.run([command, solver_command], capture_output=True, text=True)
    paths = where_solver.stdout.splitlines()
    
    if paths:  # Ensure that at least one path was found
        path_solver = paths[0]
        
        # Check if the path is already in the environment variable PATH
        if path_solver not in os.environ["PATH"]:
            # If not in PATH, add it
            os.environ["PATH"] += os.pathsep + path_solver
            print("Path added:", path_solver)
    else:
        print(f"No '{solver_command}' found on the system.")
    #
#
def main_executer(n1, packaged_useful_elements, scenario_list_print, params):
    set_first_list(scenario_list_print, params)
    # Obtaining the absolute path of the specified script file
    file_absolute_address = os.path.abspath(params['B1_script'])
    # Getting the configuration main path
    file_config_address = get_config_main_path(os.path.abspath(''))
    # Escaping characters that could be interpreted by operating system functions
    file_address = file_absolute_address.replace(params['B1_script'], '')
    #
    # Constructing the address for the case using the parameters provided
    case_address = os.path.join(file_address, params['results'].replace('\\\\', ''), first_list[n1])
    # Listing '.txt' files in the specified directory
    this_case = [e for e in os.listdir(case_address) if '.txt' in e]
    #
    # Preparing the command start sequence
    str_start = params['start'] + file_address
    #
    # Constructing paths for the data file and the output file, adapting for file system differences
    data_file = os.path.join(case_address, this_case[0] )
    output_file = os.path.join(case_address, this_case[0].replace('.txt', '_output') )
    #
    # Determining the solver based on parameters
    solver = params['solver']

    if solver == 'glpk' and params['glpk_option'] == 'old':
        # Using older GLPK options
        # Check and update environment variables if necessary
        check_enviro_variables('glpsol')

        # Composing the command to solve the model
        str_solve = 'glpsol -m ' + params['OSeMOSYS_Model'] + ' -d ' + str(data_file) + ' -o ' + str(output_file) + '.txt'
        os.system(str_start and str_solve)
        #
        # Processing data post-solution
        data_processor(n1, packaged_useful_elements, params)

    elif solver == 'glpk' and params['glpk_option'] == 'new':
        # Using newer GLPK options
        # Check and update environment variables if necessary
        check_enviro_variables('glpsol')
            
        # Composing the command to solve the model with new options
        str_solve = 'glpsol -m ' + params['OSeMOSYS_Model'] + ' -d ' + str(data_file) + ' --wglp ' + output_file + '.glp --write ' + output_file + '.sol'
        os.system(str_start and str_solve)        
    else:      
        # For LP models
        str_solve = 'glpsol -m ' + params['OSeMOSYS_Model'] + ' -d ' + str(data_file) + ' --wlp ' + output_file + '.lp --check'
        os.system(str_start and str_solve)
        if solver == 'cbc':
            # Using CBC solver
            # Check and update environment variables if necessary
            check_enviro_variables('cbc')
            
            # Composing the command for CBC solver
            str_solve = 'cbc ' + output_file + '.lp -seconds ' + str(params['iteration_time']) + ' solve -solu ' + output_file + '.sol'
            os.system(str_start and str_solve)
        elif solver == 'cplex':
            # Using CPLEX solver
            if os.path.exists(output_file + '.sol'):
                shutil.os.remove(output_file + '.sol')
            
            # Check and update environment variables if necessary
            check_enviro_variables('cplex')
                
            # Composing the command for CPLEX solver
            str_solve = 'cplex -c "read ' + output_file + '.lp" "optimize" "write ' + output_file + '.sol"'
            os.system(str_start and str_solve)
    
    # Handling configuration if yaml is not available for otoole use
    if not (solver == 'glpk' and params['glpk_option'] == 'old') and not os.path.exists(file_config_address + 'config'):
        script_otoole_config = os.path.join(file_config_address, params['otoole_config'])
        str_otoole_config = 'python -u ' + script_otoole_config
        os.system(str_start and str_otoole_config)
    
        
    script_config = os.path.join(file_config_address, params['config'])
    file_path_conv_format = os.path.join(script_config, params['conv_format'])
    file_path_template = os.path.join(script_config, params['templates'])
    file_path_outputs = os.path.join(case_address, params['outputs'].replace('/',''))
    
    
    # Converting outputs from .sol to csv format
    if solver == 'glpk' and params['glpk_option'] == 'new':
        str_outputs = 'otoole results ' + solver + ' csv ' + output_file + '.sol ' + file_path_outputs + ' datafile ' + str(data_file) + ' ' + file_path_conv_format + ' --glpk_model ' + output_file + '.glp'
        os.system(str_start and str_outputs)
        
    elif solver == 'cbc' or solver == 'cplex':  # the command line for cbc and cplex is the same, the unique difference is the name of the solver
        # but this attribute comes from the variable 'solver' and that variable comes from yaml parametrization file
        str_outputs = 'otoole results ' + solver + ' csv ' + output_file + '.sol ' + file_path_outputs + ' csv ' + file_path_template + ' ' + file_path_conv_format
        os.system(str_start and str_outputs)
    
    time.sleep(1)
    # Module to concatenate csvs otoole outputs
    if (params['solver'] == 'glpk' and params['glpk_option'] == 'new') or solver == 'cbc' or solver == 'cplex':
        file_conca_csvs = get_config_main_path(os.path.abspath(''),'config_plots')
        script_concate_csv = os.path.join(file_conca_csvs, params['concat_csvs'])
        str_otoole_concate_csv = 'python -u ' + script_concate_csv + ' ' + str(this_case[0]) + ' 1' # last int is the ID tier
        os.system(str_start and str_otoole_concate_csv)


#
def set_first_list(scenario_list_print, params):
    #
    first_list_raw = os.listdir( params['Executables'] )
    #
    global first_list
    scenario_list_print_with_fut = [ e + '_0' for e in scenario_list_print ]
    first_list = [e for e in first_list_raw if ( '.csv' not in e ) and ( 'Table' not in e ) and ( '.py' not in e ) and ( '__pycache__' not in e ) and ( e in scenario_list_print_with_fut ) ]
    #
#
def data_processor( case, unpackaged_useful_elements, params ):
    #
    Reference_driven_distance =     unpackaged_useful_elements[0]
    Reference_occupancy_rate =      unpackaged_useful_elements[1]
    Reference_op_life =             unpackaged_useful_elements[2]
    Fleet_Groups_inv =              unpackaged_useful_elements[3]
    time_range_vector =             unpackaged_useful_elements[4]
    list_param_default_value_params = unpackaged_useful_elements[5]
    list_param_default_value_value = unpackaged_useful_elements[6]
    list_gdp_ref = unpackaged_useful_elements[7]

    # Extract the default (national) discount rate parameter
    dr_default = params['disc_rate']

    # Briefly open up the system coding to use when processing for visualization:
    df_fuel_to_code = pd.read_excel( os.path.join(params['A1_Inputs'], params['Modes_Trans']), sheet_name=params['Fuel_Code'] )
    df_fuel_2_code_fuel_list        = df_fuel_to_code['Code'].tolist()
    df_fuel_2_code_plain_english    = df_fuel_to_code['Plain English'].tolist()
    df_tech_to_code = pd.read_excel( os.path.join(params['A1_Inputs'], params['Modes_Trans']), sheet_name=params['Tech_Code'] )
    df_tech_2_code_fuel_list        = df_tech_to_code['Techs'].tolist()
    df_tech_2_code_plain_english    = df_tech_to_code['Plain English'].tolist()
    #
    # 1 - Always call the structure of the model:
    #-------------------------------------------#
    structure_filename = params['B1_model_Struct']
    structure_file = pd.ExcelFile(structure_filename)
    structure_sheetnames = structure_file.sheet_names  # see all sheet names
    sheet_sets_structure = pd.read_excel(open(structure_filename, 'rb'),
                                         header=None,
                                         sheet_name=structure_sheetnames[0])
    sheet_params_structure = pd.read_excel(open(structure_filename, 'rb'),
                                           header=None,
                                           sheet_name=structure_sheetnames[1])
    sheet_vars_structure = pd.read_excel(open(structure_filename, 'rb'),
                                         header=None,
                                         sheet_name=structure_sheetnames[2])

    S_DICT_sets_structure = params['S_DICT_sets_structure']
    for col in range(1,11+1):  # 11 columns
        S_DICT_sets_structure['set'].append( sheet_sets_structure.iat[0, col] )
        S_DICT_sets_structure['initial'].append( sheet_sets_structure.iat[1, col] )
        S_DICT_sets_structure['number_of_elements'].append( int( sheet_sets_structure.iat[2, col] ) )
        #
        element_number = int( sheet_sets_structure.iat[2, col] )
        this_elements_list = []
        if element_number > 0:
            for n in range( 1, element_number+1 ):
                this_elements_list.append( sheet_sets_structure.iat[2+n, col] )
        S_DICT_sets_structure['elements_list'].append( this_elements_list )
    #
    S_DICT_params_structure = params['S_DICT_params_structure']
    param_category_list = []
    for col in range(1,30+1):  # 30 columns
        if str( sheet_params_structure.iat[0, col] ) != '':
            param_category_list.append( sheet_params_structure.iat[0, col] )
        S_DICT_params_structure['category'].append( param_category_list[-1] )
        S_DICT_params_structure['parameter'].append( sheet_params_structure.iat[1, col] )
        S_DICT_params_structure['number_of_elements'].append( int( sheet_params_structure.iat[2, col] ) )
        #
        index_number = int( sheet_params_structure.iat[2, col] )
        this_index_list = []
        for n in range(1, index_number+1):
            this_index_list.append( sheet_params_structure.iat[2+n, col] )
        S_DICT_params_structure['index_list'].append( this_index_list )
    #
    S_DICT_vars_structure = params['S_DICT_vars_structure']
    var_category_list = []
    for col in range(1,43+1):  # 43 columns
        if str( sheet_vars_structure.iat[0, col] ) != '':
            var_category_list.append( sheet_vars_structure.iat[0, col] )
        S_DICT_vars_structure['category'].append( var_category_list[-1] )
        S_DICT_vars_structure['variable'].append( sheet_vars_structure.iat[1, col] )
        S_DICT_vars_structure['number_of_elements'].append( int( sheet_vars_structure.iat[2, col] ) )
        #
        index_number = int( sheet_vars_structure.iat[2, col] )
        this_index_list = []
        for n in range(1, index_number+1):
            this_index_list.append( sheet_vars_structure.iat[2+n, col] )
        S_DICT_vars_structure['index_list'].append( this_index_list )
    #-------------------------------------------#
    all_vars = params['all_vars']
    #
    more_vars = params['more_vars']
    #
    filter_vars = params['filter_vars']
    #
    all_vars_output_dict = [ {} for e in range( len( first_list ) ) ]
    #
    output_header = params['output_header']
    #-------------------------------------------------------#
    for v in range( len( all_vars ) ):
        output_header.append( all_vars[v] )
    for v in range( len( more_vars ) ):
        output_header.append( more_vars[v] )
    for v in range( len( filter_vars ) ):
        output_header.append( filter_vars[v] )
    #-------------------------------------------------------#
    this_strategy = first_list[case].split('_')[0] 
    this_future   = first_list[case].split('_')[1]
    #-------------------------------------------------------#
    #
    vars_as_appear = []
    # data_name = params['Executables'] + str( '/' + first_list[case] ) + '/' + str(first_list[case]) + '_output.txt'
    data_name = os.path.join(params['Executables'], str( first_list[case] ),  f'{str(first_list[case])}_output.txt')
    #
    n = 0
    break_this_while = False
    while break_this_while == False:
        n += 1
        structure_line_raw = linecache.getline(data_name, n)
        if 'No. Column name  St   Activity     Lower bound   Upper bound    Marginal' in structure_line_raw:
            ini_line = deepcopy( n+2 )
        if params['KKT'] in structure_line_raw:
            end_line = deepcopy( n-1 )
            break_this_while = True
            break
    #
    for n in range(ini_line, end_line, 2 ):
        structure_line_raw = linecache.getline(data_name, n)
        structure_list_raw = structure_line_raw.split(' ')
        #
        structure_list_raw_2 = [ s_line for s_line in structure_list_raw if s_line != '' ]
        structure_line = structure_list_raw_2[1]
        structure_list = structure_line.split('[')
        the_variable = structure_list[0]
        #
        if the_variable in all_vars:
            set_list = structure_list[1].replace(']','').replace('\n','').split(',')
            #--%
            index = S_DICT_vars_structure['variable'].index( the_variable )
            this_variable_indices = S_DICT_vars_structure['index_list'][ index ]
            #
            if 'y' in this_variable_indices:
                data_line = linecache.getline(data_name, n+1)
                data_line_list_raw = data_line.split(' ')
                data_line_list = [ data_cell for data_cell in data_line_list_raw if data_cell != '' ]
                useful_data_cell = data_line_list[1]
                #--%
                if useful_data_cell != '0':
                    #
                    if the_variable not in vars_as_appear:
                        vars_as_appear.append( the_variable )
                        all_vars_output_dict[case].update({ the_variable:{} })
                        all_vars_output_dict[case][the_variable].update({ the_variable:[] })
                        #
                        for n in range( len( this_variable_indices ) ):
                            all_vars_output_dict[case][the_variable].update( { this_variable_indices[n]:[] } )
                    #--%
                    this_variable = vars_as_appear[-1]
                    all_vars_output_dict[case][this_variable][this_variable].append( useful_data_cell )
                    for n in range( len( this_variable_indices ) ):
                        all_vars_output_dict[case][the_variable][ this_variable_indices[n] ].append( set_list[n] )
                #
            #
            elif 'y' not in this_variable_indices:
                data_line = linecache.getline(data_name, n+1)
                data_line_list_raw = data_line.split(' ')
                data_line_list = [ data_cell for data_cell in data_line_list_raw if data_cell != '' ]
                useful_data_cell = data_line_list[1]
                #--%
                if useful_data_cell != '0':
                    #
                    if the_variable not in vars_as_appear:
                        vars_as_appear.append( the_variable )
                        all_vars_output_dict[case].update({ the_variable:{} })
                        all_vars_output_dict[case][the_variable].update({ the_variable:[] })
                        #
                        for n in range( len( this_variable_indices ) ):
                            all_vars_output_dict[case][the_variable].update( { this_variable_indices[n]:[] } )
                    #--%
                    this_variable = vars_as_appear[-1]
                    all_vars_output_dict[case][this_variable][this_variable].append( useful_data_cell )
                    for n in range( len( this_variable_indices ) ):
                        all_vars_output_dict[case][the_variable][ this_variable_indices[n] ].append( set_list[n] )
        #--%
        else:
            pass
    #
    linecache.clearcache()
    #%%
    #-----------------------------------------------------------------------------------------------------------%
    # Extract *AccumulatedNewCapacity* for every vehicle technology:
    this_var_out1 = 'AccumulatedNewCapacity'
    this_var_dict_out1 = all_vars_output_dict[case]['AccumulatedNewCapacity']
    this_var_techs = this_var_dict_out1['t']
    this_var_techs_unique = list( Fleet_Groups_inv.keys() )  # transport techs only
    this_var_years = this_var_dict_out1['y']
    this_var_years_unique = list(set(this_var_dict_out1['y']))
    this_var_years_unique.sort()
    aide_dict_accumnewcap = {}
    for t_aid in this_var_techs_unique:
        aide_dict_accumnewcap.update({t_aid:[]})
        for y_aid in this_var_years_unique:
            tech_indices = \
                [i for i, x in enumerate(this_var_techs) if (t_aid in x )]
            year_indices = \
                [i for i, x in enumerate(this_var_years) if (y_aid in x )]
            position_idx_list = \
                list(set(tech_indices) & set(year_indices))
            if len(position_idx_list) != 0:
                position_idx = position_idx_list[0]
                this_value = float(this_var_dict_out1[this_var_out1][position_idx])
            elif len(position_idx_list) == 0:
                this_value = 0
            else:
                print('*This is an error (out 1)*')
                sys.exit()
            aide_dict_accumnewcap[t_aid].append(this_value)

    # Extract *CapitalInvestment* for every vehicle technology: // gotta divide by *NewCapacity* below:
    this_var_out2 = 'CapitalInvestment'
    this_var_dict_out2 = all_vars_output_dict[case]['CapitalInvestment']
    this_var_techs = this_var_dict_out2['t']
    this_var_techs_unique = list( Fleet_Groups_inv.keys() )  # transport techs only
    this_var_years = this_var_dict_out2['y']
    this_var_years_unique = list(set(this_var_dict_out2['y']))
    this_var_years_unique.sort()
    aide_dict_capitalcost = {}
    for t_aid in this_var_techs_unique:
        aide_dict_capitalcost.update({t_aid:[]})
        for y_aid in this_var_years_unique:
            tech_indices = \
                [i for i, x in enumerate(this_var_techs) if (t_aid in x )]
            year_indices = \
                [i for i, x in enumerate(this_var_years) if (y_aid in x )]
            position_idx_list = \
                list(set(tech_indices) & set(year_indices))
            if len(position_idx_list) != 0:
                position_idx = position_idx_list[0]
                this_value = float(this_var_dict_out2[this_var_out2][position_idx])
            elif len(position_idx_list) == 0:
                this_value = 0
            else:
                print('*This is an error (out 2)*')
                sys.exit()
            aide_dict_capitalcost[t_aid].append(this_value)

    # Extract *NewCapacity* for every vehicle technology:
    this_var_out3 = 'NewCapacity'
    this_var_dict_out3 = all_vars_output_dict[case]['NewCapacity']
    this_var_techs = this_var_dict_out3['t']
    this_var_techs_unique = list( Fleet_Groups_inv.keys() )  # transport techs only
    this_var_years = this_var_dict_out3['y']
    this_var_years_unique = list(set(this_var_dict_out3['y']))
    this_var_years_unique.sort()
    aide_dict_newcap = {}
    for t_aid in this_var_techs_unique:
        aide_dict_newcap.update({t_aid:[]})
        for y_aid in this_var_years_unique:
            tech_indices = \
                [i for i, x in enumerate(this_var_techs) if (t_aid in x )]
            year_indices = \
                [i for i, x in enumerate(this_var_years) if (y_aid in x )]
            position_idx_list = \
                list(set(tech_indices) & set(year_indices))
            if len(position_idx_list) != 0:
                position_idx = position_idx_list[0]
                this_value = float(this_var_dict_out3[this_var_out3][position_idx])
            elif len(position_idx_list) == 0:
                this_value = 0
            else:
                print('*This is an error (out 3)*')
                sys.exit()
            aide_dict_newcap[t_aid].append(this_value)

    # Extract *TotalCapacityAnnual* for every vehicle technology:
    this_var_out4 = 'TotalCapacityAnnual'
    this_var_dict_out4 = all_vars_output_dict[case]['TotalCapacityAnnual']
    this_var_techs = this_var_dict_out4['t']
    this_var_techs_unique = list( Fleet_Groups_inv.keys() )  # transport techs only
    this_var_years = this_var_dict_out4['y']
    this_var_years_unique = list(set(this_var_dict_out4['y']))
    this_var_years_unique.sort()
    aide_dict_totalcap = {}
    for t_aid in this_var_techs_unique:
        aide_dict_totalcap.update({t_aid:[]})
        for y_aid in this_var_years_unique:
            tech_indices = \
                [i for i, x in enumerate(this_var_techs) if (t_aid in x )]
            year_indices = \
                [i for i, x in enumerate(this_var_years) if (y_aid in x )]
            position_idx_list = \
                list(set(tech_indices) & set(year_indices))
            if len(position_idx_list) != 0:
                position_idx = position_idx_list[0]
                this_value = float(this_var_dict_out4[this_var_out4][position_idx])
            elif len(position_idx_list) == 0:
                this_value = 0
            else:
                print('*This is an error (out 4)*')
                sys.exit()
            aide_dict_totalcap[t_aid].append(this_value)

    #-----------------------------------------------------------------------------------------------------------%
    output_adress = os.path.join(params['Executables'], str( first_list[case] ))
    combination_list = [] # [fuel, technology, emission, year]
    data_row_list = []
    for var in range( len( vars_as_appear ) ):
        this_variable = vars_as_appear[var]
        this_var_dict = all_vars_output_dict[case][this_variable]
        #--%
        index = S_DICT_vars_structure['variable'].index( this_variable )
        this_variable_indices = S_DICT_vars_structure['index_list'][ index ]
        #--------------------------------------#
        for k in range( len( this_var_dict[this_variable] ) ):
            this_combination = []
            #
            if 'f' in this_variable_indices:
                this_combination.append( this_var_dict['f'][k] )
            else:
                this_combination.append( '' )
            #
            if 't' in this_variable_indices:
                this_combination.append( this_var_dict['t'][k] )
            else:
                this_combination.append( '' )
            #
            if 'e' in this_variable_indices:
                this_combination.append( this_var_dict['e'][k] )
            else:
                this_combination.append( '' )
            #
            if 'l' in this_variable_indices:
                this_combination.append( '' )
            else:
                this_combination.append( '' )
            #
            if 'y' in this_variable_indices:
                this_combination.append( this_var_dict['y'][k] )
            else:
                this_combination.append( '' )
            #
            if this_combination not in combination_list:
                combination_list.append( this_combination )
                data_row = ['' for n in range( len( output_header ) ) ]
                # print('check', len(data_row), len(run_id) )
                data_row[0] = this_strategy
                data_row[1] = this_future
                data_row[2] = this_combination[0] # Fuel
                data_row[3] = this_combination[1] # Technology
                data_row[4] = this_combination[2] # Emission
                # data_row[7] = this_combination[3]
                data_row[5] = this_combination[4] # Year
                #
                var_position_index = output_header.index( this_variable )
                data_row[ var_position_index ] = this_var_dict[ this_variable ][ k ]
                data_row_list.append( data_row )
            else:
                ref_index = combination_list.index( this_combination )
                this_data_row = deepcopy( data_row_list[ ref_index ] )
                #
                var_position_index = output_header.index( this_variable )
                #
                if 'l' in this_variable_indices: 
                    #
                    if str(this_data_row[ var_position_index ]) != '' and str(this_var_dict[ this_variable ][ k ]) != '' and ( 'Rate' not in this_variable ):
                        this_data_row[ var_position_index ] = str(  float( this_data_row[ var_position_index ] ) + float( this_var_dict[ this_variable ][ k ] ) )
                    elif str(this_data_row[ var_position_index ]) == '' and str(this_var_dict[ this_variable ][ k ]) != '':
                        this_data_row[ var_position_index ] = str( float( this_var_dict[ this_variable ][ k ] ) )
                    elif str(this_data_row[ var_position_index ]) != '' and str(this_var_dict[ this_variable ][ k ]) == '':
                        pass
                else:
                    this_data_row[ var_position_index ] = this_var_dict[ this_variable ][ k ]
                #
                data_row_list[ ref_index ]  = deepcopy( this_data_row )
                #
            #
            if this_variable == 'TotalCapacityAnnual' or this_variable == 'NewCapacity':
                ref_index = combination_list.index( this_combination )
                this_data_row = deepcopy( data_row_list[ ref_index ] ) # this must be updated in a further position of the list
                #
                this_tech = this_combination[1]
                if this_tech in list( Fleet_Groups_inv.keys() ):
                    group_tech = Fleet_Groups_inv[ this_tech ]
                    #
                    this_year = this_combination[4]
                    this_year_index = time_range_vector.index( int( this_year ) )
                    #
                    ref_var_position_index = output_header.index( this_variable )
                    #
                    if params['trains'] not in group_tech and params['group_tech'] not in group_tech:
                        driven_distance = float(Reference_driven_distance[this_strategy][group_tech][this_year_index])
                        driven_distance_change_year = \
                            float(Reference_driven_distance[this_strategy][group_tech][time_range_vector.index(params['change_year_B1'])])
                        driven_distance_final = \
                            float(Reference_driven_distance[this_strategy][group_tech][time_range_vector.index(params['final_year'])])
                        passenger_per_vehicle = float(Reference_occupancy_rate[this_strategy][group_tech][this_year_index])
                        #
                        if this_variable == 'NewCapacity':
                            var_position_index = output_header.index( params['newfleet'] )
                            new_cap_calculated = \
                                float(this_data_row[ref_var_position_index])
                            new_fleet_calculated = \
                                (10**9)*new_cap_calculated/driven_distance

                            accum_new_cap_subtract = 0
                            driven_distance_prev = float(Reference_driven_distance[this_strategy][group_tech][this_year_index-1])
                            if int(this_year) >= params['change_year_B1'] and driven_distance_change_year != driven_distance_final:
                                # We need to select the appropiate accumulated new capacity:
                                accum_new_cap = \
                                    aide_dict_accumnewcap[this_tech][this_year_index-1]

                                if int(this_year) > params['year_first_range']+Reference_op_life[this_strategy][this_tech]:  # works with a timeframe by 2050; longer timeframes may need more lifetime cycle adjustments
                                    subtract_index_accum_old = \
                                        int(Reference_op_life[this_strategy][this_tech])
                                    if int(this_year) - subtract_index_accum_old <= params['change_year_B1']:
                                        driven_distance_17 = \
                                            float(Reference_driven_distance[this_strategy][group_tech][time_range_vector.index(params['year_driven_distance'])])
                                    else:
                                        driven_distance_17 = \
                                            float(Reference_driven_distance[this_strategy][group_tech][this_year_index-subtract_index_accum_old])

                                    adjust_factor_old_cap = \
                                        (1-driven_distance/driven_distance_17)

                                    if int(this_year) - subtract_index_accum_old > params['change_year_B1']:  # this is an odd rule, but works empirically
                                        driven_distance_change_year = \
                                            float(Reference_driven_distance[this_strategy][group_tech][time_range_vector.index(params['change_year_B1'])])
                                        driven_distance_minus3 = \
                                            float(Reference_driven_distance[this_strategy][group_tech][this_year_index-3])
                                        adjust_factor_old_cap -= \
                                            (1-driven_distance_17/driven_distance_change_year)
                                        adjust_factor_old_cap -= \
                                            (1-driven_distance/driven_distance_minus3)

                                    accum_new_cap_subtract += \
                                        aide_dict_newcap[this_tech][this_year_index-subtract_index_accum_old]*adjust_factor_old_cap

                                new_cap_adj = \
                                    accum_new_cap*(1-driven_distance/driven_distance_prev) - accum_new_cap_subtract
                            else:
                                new_cap_adj = 0
                            new_fleet_adj = \
                                (10**9)*new_cap_adj/driven_distance
                            this_data_row[ var_position_index ] = \
                                round(new_fleet_calculated + new_fleet_adj, params['round_#'])
                        #
                        if this_variable == 'TotalCapacityAnnual':
                            #
                            var_position_index = output_header.index( params['fleet'] )
                            this_data_row[ var_position_index ] =  round( (10**9)*float( this_data_row[ ref_var_position_index ] )/driven_distance, params['round_#'])
                            #
                            var_position_index = output_header.index( params['dist_driven'] )
                            this_data_row[ var_position_index ] = round(driven_distance, params['round_#'])
                            #
                            var_position_index = output_header.index( params['produced_mob'] )
                            this_data_row[ var_position_index ] =  round( float( this_data_row[ ref_var_position_index ] )*passenger_per_vehicle, params['round_#'])
                        #
                        data_row_list[ ref_index ] = deepcopy( this_data_row )
                        #
                    #
                    ###################################################################################################
                    #
                #
            #
            # Creating convinience filters for the analysis of model outputs:
            this_tech = this_combination[1]
            if this_tech in list( Fleet_Groups_inv.keys() ):
                ref_index = combination_list.index( this_combination )
                this_data_row = deepcopy( data_row_list[ ref_index ] ) # this must be updated in a further position of the list
                #
                var_position_index = output_header.index( params['filt_fuel_type'] )
                ############################################
                # By Fuel Type
                for r in range( len( df_fuel_2_code_fuel_list ) ):
                    if df_fuel_2_code_fuel_list[r] in this_tech:
                        this_data_row[ var_position_index ] = df_fuel_2_code_plain_english[ r ]
                        break
                ############################################
                var_position_index = output_header.index( params['filt_veh_type'] )
                # By vehicle type
                for r in range( len( df_tech_2_code_fuel_list ) ):
                    if df_tech_2_code_fuel_list[r] in this_tech:
                        this_data_row[ var_position_index ] = df_tech_2_code_plain_english[ r ]
                        break
                data_row_list[ ref_index ] = deepcopy( this_data_row )
            #
            output_csv_r = dr_default*100 # This dr_default is the DiscountRate was defined in default_val_params
            output_csv_year = params['year_apply_discount_rate']
            #
            if this_combination[2] in params['this_combina'] and this_variable == 'AnnualTechnologyEmissionPenaltyByEmission':
                ref_index = combination_list.index( this_combination )
                this_data_row = deepcopy( data_row_list[ ref_index ] ) # this must be updated in a further position of the list
                #
                ref_var_position_index = output_header.index( 'AnnualTechnologyEmissionPenaltyByEmission' )
                new_var_position_index = output_header.index( params['externa_base_year'] )
                new2_var_position_index = output_header.index( params['externa_gdp'] )
                #
                this_year = this_combination[4]
                this_year_index = time_range_vector.index( int( this_year ) )
                #
                resulting_value_raw = float(this_data_row[ ref_var_position_index ]) / ( ( 1 + output_csv_r/100 )**( float(this_year) - output_csv_year ) )
                resulting_value = round( resulting_value_raw, params['round_#'])
                #
                this_data_row[new_var_position_index] = str( resulting_value )
                this_data_row[new2_var_position_index] = str( float(this_data_row[ ref_var_position_index ])/list_gdp_ref[this_year_index] )
                #
                data_row_list[ ref_index ] = deepcopy( this_data_row )
                #
            #
            ''' $ This is new (beginning) $ '''
            #
            if this_variable == 'CapitalInvestment':
                ref_index = combination_list.index( this_combination )
                this_data_row = deepcopy( data_row_list[ ref_index ] ) # this must be updated in a further position of the list
                #
                ref_var_position_index = output_header.index( 'CapitalInvestment' )
                new_var_position_index = output_header.index( params['capex_base_year'] )
                new2_var_position_index = output_header.index( params['capex_gdp'] )
                #
                this_year = this_combination[4]
                this_year_index = time_range_vector.index( int( this_year ) )
                #
                # Here we must add an adjustment to the capital investment to make the fleet constant:
                this_base_cap_inv = \
                    float(this_data_row[ ref_var_position_index ])
                this_plus_cap_inv = 0
                '''
                > We need to add new capital costs if the technology is in transport:
                '''
                if this_tech in list(Fleet_Groups_inv.keys()):
                    group_tech = Fleet_Groups_inv[this_tech]
                    if params['trains'] not in group_tech  and params['group_tech'] not in group_tech:
                        driven_distance = float(Reference_driven_distance[this_strategy][group_tech][this_year_index])
                        driven_distance_change_year = \
                            float(Reference_driven_distance[this_strategy][group_tech][time_range_vector.index(params['change_year_B1'])])
                        driven_distance_final = \
                            float(Reference_driven_distance[this_strategy][group_tech][time_range_vector.index(params['final_year'])])

                        accum_new_cap_subtract = 0
                        driven_distance_prev = float(Reference_driven_distance[this_strategy][group_tech][this_year_index-1])
                        if int(this_year) >= params['change_year_B1'] and driven_distance_change_year != driven_distance_final:  # greater than the base year
                            # We need to select the appropiate accumulated new capacity:
                            accum_new_cap = \
                                aide_dict_accumnewcap[this_tech][this_year_index-1]

                            if int(this_year) > params['year_first_range']+Reference_op_life[this_strategy][this_tech]:
                                subtract_index_accum_old = \
                                    int(Reference_op_life[this_strategy][this_tech])
                                if int(this_year) - subtract_index_accum_old <= params['change_year_B1']:
                                    driven_distance_17 = \
                                        float(Reference_driven_distance[this_strategy][group_tech][time_range_vector.index(params['year_driven_distance'])])
                                else:
                                    driven_distance_17 = \
                                        float(Reference_driven_distance[this_strategy][group_tech][this_year_index-subtract_index_accum_old])

                                adjust_factor_old_cap = \
                                    (1-driven_distance/driven_distance_17)

                                if int(this_year) - subtract_index_accum_old > params['change_year_B1']:  # this is an odd rule, but works empirically
                                    driven_distance_minus3 = \
                                        float(Reference_driven_distance[this_strategy][group_tech][this_year_index-3])
                                    adjust_factor_old_cap -= \
                                        (1-driven_distance_17/driven_distance_change_year)
                                    adjust_factor_old_cap -= \
                                        (1-driven_distance/driven_distance_minus3)

                                accum_new_cap_subtract += \
                                    aide_dict_newcap[this_tech][this_year_index-subtract_index_accum_old]*adjust_factor_old_cap

                            new_cap_adj = \
                                accum_new_cap*(1-driven_distance/driven_distance_prev) - accum_new_cap_subtract
                            this_cap_cost = \
                                aide_dict_capitalcost[this_tech][this_year_index]
                            this_new_cap = \
                                aide_dict_newcap[this_tech][this_year_index]
                            this_plus_cap_inv += \
                                new_cap_adj*this_cap_cost/this_new_cap

                this_cap_inv = this_base_cap_inv + this_plus_cap_inv
                this_data_row[ref_var_position_index] = \
                    str(this_cap_inv)  # Here we re-write the new capacity to adjust the system
                '''
                > Below we continue as usual:
                '''
                resulting_value_raw = this_cap_inv/((1 + output_csv_r/100)**(float(this_year) - output_csv_year))
                resulting_value = round( resulting_value_raw, params['round_#'])
                #
                this_data_row[new_var_position_index] = str( resulting_value )
                this_data_row[new2_var_position_index] = str( float(this_data_row[ ref_var_position_index ])/list_gdp_ref[this_year_index] )
                #
                data_row_list[ ref_index ] = deepcopy( this_data_row )
                #
            #
            if this_variable == 'AnnualFixedOperatingCost':
                ref_index = combination_list.index( this_combination )
                this_data_row = deepcopy( data_row_list[ ref_index ] ) # this must be updated in a further position of the list
                #
                ref_var_position_index = output_header.index( 'AnnualFixedOperatingCost' )
                new_var_position_index = output_header.index( params['fopex_base_year'] )
                new2_var_position_index = output_header.index( params['fopex_gdp'] )
                #
                this_year = this_combination[4]
                this_year_index = time_range_vector.index( int( this_year ) )
                #
                resulting_value_raw = float(this_data_row[ ref_var_position_index ]) / ( ( 1 + output_csv_r/100 )**( float(this_year) - output_csv_year ) )
                resulting_value = round( resulting_value_raw, params['round_#'])
                #
                this_data_row[new_var_position_index] = str( resulting_value )
                this_data_row[new2_var_position_index] = str( float(this_data_row[ ref_var_position_index ])/list_gdp_ref[this_year_index] )
                #
                data_row_list[ ref_index ] = deepcopy( this_data_row )
                #
            #
            if this_variable == 'AnnualVariableOperatingCost':
                ref_index = combination_list.index( this_combination )
                this_data_row = deepcopy( data_row_list[ ref_index ] ) # this must be updated in a further position of the list
                #
                ref_var_position_index = output_header.index( 'AnnualVariableOperatingCost' )
                new_var_position_index = output_header.index( params['vopex_base_year'] )
                new2_var_position_index = output_header.index( params['vopex_gdp'] )
                #
                this_year = this_combination[4]
                this_year_index = time_range_vector.index( int( this_year ) )
                #
                resulting_value_raw = float(this_data_row[ ref_var_position_index ]) / ( ( 1 + output_csv_r/100 )**( float(this_year) - output_csv_year ) )
                resulting_value = round( resulting_value_raw, params['round_#'])
                #
                this_data_row[new_var_position_index] = str( resulting_value )
                this_data_row[new2_var_position_index] = str( float(this_data_row[ ref_var_position_index ])/list_gdp_ref[this_year_index] )
                #
                data_row_list[ ref_index ] = deepcopy( this_data_row )
                #
            #
            if this_variable == 'OperatingCost':
                ref_index = combination_list.index( this_combination )
                this_data_row = deepcopy( data_row_list[ ref_index ] ) # this must be updated in a further position of the list
                #
                ref_var_position_index = output_header.index( 'OperatingCost' )
                new_var_position_index = output_header.index( params['opex_base_year'] )
                new2_var_position_index = output_header.index( params['opex_gdp'] )
                #
                this_year = this_combination[4]
                this_year_index = time_range_vector.index( int( this_year ) )
                #
                resulting_value_raw = float(this_data_row[ ref_var_position_index ]) / ( ( 1 + output_csv_r/100 )**( float(this_year) - output_csv_year ) )
                resulting_value = round( resulting_value_raw, params['round_#'])
                #
                this_data_row[new_var_position_index] = str( resulting_value )
                this_data_row[new2_var_position_index] = str( float(this_data_row[ ref_var_position_index ])/list_gdp_ref[this_year_index] )
                #
                data_row_list[ ref_index ] = deepcopy( this_data_row )
                #
            #
            ''' $ (end) $ '''
            #
        #
    #
    non_year_combination_list = []
    non_year_combination_list_years = []
    for n in range( len( combination_list ) ):
        this_combination = combination_list[ n ]
        this_non_year_combination = [ this_combination[0], this_combination[1], this_combination[2] ]
        if this_combination[4] != '' and this_non_year_combination not in non_year_combination_list:
            non_year_combination_list.append( this_non_year_combination )
            non_year_combination_list_years.append( [ this_combination[4] ] )
        elif this_combination[4] != '' and this_non_year_combination in non_year_combination_list:
            non_year_combination_list_years[ non_year_combination_list.index( this_non_year_combination ) ].append( this_combination[4] )
    #
    for n in range( len( non_year_combination_list ) ):
        if len( non_year_combination_list_years[n] ) != len(time_range_vector):
            #
            this_existing_combination = non_year_combination_list[n]
            # print('flag 1', this_existing_combination )
            this_existing_combination.append( '' )
            # print('flag 2', this_existing_combination )
            this_existing_combination.append( non_year_combination_list_years[n][0] )
            # print('flag 3', this_existing_combination )
            ref_index = combination_list.index( this_existing_combination )
            this_existing_data_row = deepcopy( data_row_list[ ref_index ] )
            #
            for n2 in range( len( time_range_vector ) ):
                #
                if time_range_vector[n2] not in non_year_combination_list_years[n]:
                    #
                    data_row = ['' for n in range( len( output_header ) ) ]
                    data_row[0] = this_strategy
                    data_row[1] = this_future
                    data_row[2] = non_year_combination_list[n][0]
                    data_row[3] = non_year_combination_list[n][1]
                    data_row[4] = non_year_combination_list[n][2]
                    data_row[5] = time_range_vector[n2]
                    #
                    for n3 in range( len( vars_as_appear ) ):
                        this_variable = vars_as_appear[n3]
                        this_var_dict = all_vars_output_dict[case][this_variable]
                        index = S_DICT_vars_structure['variable'].index( this_variable )
                        this_variable_indices = S_DICT_vars_structure['index_list'][ index ]
                        #
                        var_position_index = output_header.index( this_variable )
                        #
                        print_true = False
                        if ( 'f' in this_variable_indices and str(non_year_combination_list[n][0]) != '' ): # or ( 'f' not in this_variable_indices and str(non_year_combination_list[n][0]) == '' ):
                            print_true = True
                        else:
                            pass
                        #
                        if ( 't' in this_variable_indices and str(non_year_combination_list[n][1]) != '' ): # or ( 't' not in this_variable_indices and str(non_year_combination_list[n][1]) == '' ):
                            print_true = True
                        else:
                            pass
                        #
                        if ( 'e' in this_variable_indices and str(non_year_combination_list[n][2]) != '' ): # or ( 'e' not in this_variable_indices and str(non_year_combination_list[n][2]) == '' ):
                            print_true = True
                        else:
                            pass
                        #
                        if 'y' in this_variable_indices and ( str( this_existing_data_row[ var_position_index ] ) != '' ) and print_true == True:
                            data_row[ var_position_index ] = '0'
                            #
                        else:
                            pass
                    #
                    data_row_list.append( data_row )
    #--------------------------------------#
    with open( output_adress + '/' + str( first_list[case] ) + '_Output' + '.csv', 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        csvwriter.writerow( output_header )
        for n in range( len( data_row_list ) ):
            csvwriter.writerow( data_row_list[n] )
    #-----------------------------------------------------------------------------------------------------------%
    if params['del_files']:
        shutil.os.remove(data_name) #-----------------------------------------------------------------------------------------------------------%
    gc.collect(generation=2)
    time.sleep(0.05)
    #-----------------------------------------------------------------------------------------------------------%
    print(  'We finished with printing the outputs: ' + str( first_list[case] ) )
#
def function_C_mathprog( scen, stable_scenarios, unpackaged_useful_elements, params ):
    #
    scenario_list =                     unpackaged_useful_elements[0]
    S_DICT_sets_structure =             unpackaged_useful_elements[1]
    S_DICT_params_structure =           unpackaged_useful_elements[2]
    list_param_default_value_params =   unpackaged_useful_elements[3]
    list_param_default_value_value =    unpackaged_useful_elements[4]
    print_adress =                      unpackaged_useful_elements[5]

    # Extract the default (national) discount rate parameter
    dr_default = params['disc_rate']                                                      

    Reference_driven_distance =         unpackaged_useful_elements[6]
    Fleet_Groups_inv =                  unpackaged_useful_elements[7]
    time_range_vector =                 unpackaged_useful_elements[8]
    if params['Use_Blend_Shares_B1']:
        Blend_Shares =                      unpackaged_useful_elements[9]
    #
    # Briefly open up the system coding to use when processing for visualization:
    df_fuel_to_code = pd.read_excel( os.path.join(params['A1_Inputs'], params['Modes_Trans']), sheet_name=params['Fuel_Code'] )
    df_fuel_2_code_fuel_list        = df_fuel_to_code['Code'].tolist()
    df_fuel_2_code_plain_english    = df_fuel_to_code['Plain English'].tolist()
    df_tech_to_code = pd.read_excel( os.path.join(params['A1_Inputs'], params['Modes_Trans']), sheet_name=params['Tech_Code'] )
    df_tech_2_code_fuel_list        = df_tech_to_code['Techs'].tolist()
    df_tech_2_code_plain_english    = df_tech_to_code['Plain English'].tolist()
    #
    # header = ['Scenario','Parameter','REGION','TECHNOLOGY','FUEL','EMISSION','MODE_OF_OPERATION','TIMESLICE','YEAR','SEASON','DAYTYPE','DAILYTIMEBRACKET','STORAGE','Value']
    header_indices = params['header_indices']
    #
    # for scen in range( len( scenario_list ) ):
    print('# This is scenario ', scenario_list[scen] )
    #
    try:
        scen_file_dir = os.path.join(print_adress, f'{str( scenario_list[scen] )}_0')
        os.mkdir( scen_file_dir )
    except OSError as exc:
        if exc.errno != errno.EEXIST:
            raise
        pass
    this_scenario_data = stable_scenarios[ scenario_list[scen] ]
    #
    g= open( os.path.join(print_adress, f'{str( scenario_list[scen] )}_0', f'{str( scenario_list[scen] )}_0.txt'),"w+")
    g.write( '###############\n#    Sets     #\n###############\n#\n' )
    g.write( 'set DAILYTIMEBRACKET :=  ;\n' )
    g.write( 'set DAYTYPE :=  ;\n' )
    g.write( 'set SEASON :=  ;\n' )
    g.write( 'set STORAGE :=  ;\n' )
    #
    for n1 in range( len( S_DICT_sets_structure['set'] ) ):
        if S_DICT_sets_structure['number_of_elements'][n1] != 0:
            g.write( 'set ' + S_DICT_sets_structure['set'][n1] + ' := ' )
            #
            for n2 in range( S_DICT_sets_structure['number_of_elements'][n1] ):
                if S_DICT_sets_structure['set'][n1] == 'YEAR' or S_DICT_sets_structure['set'][n1] == 'MODE_OF_OPERATION':
                    g.write( str( int( S_DICT_sets_structure['elements_list'][n1][n2] ) ) + ' ' )
                else:
                    g.write( str( S_DICT_sets_structure['elements_list'][n1][n2] ) + ' ' )
            g.write( ';\n' )
    #
    g.write( '\n' )
    g.write( '###############\n#    Parameters     #\n###############\n#\n' )
    #
    for p in range( len( list( this_scenario_data.keys() ) ) ):
        #
        this_param = list( this_scenario_data.keys() )[p]
        #
        default_value_list_params_index = list_param_default_value_params.index( this_param )
        default_value = float( list_param_default_value_value[ default_value_list_params_index ] )
        if default_value >= 0:
            default_value = int( default_value )
        else:
            pass
        #
        this_param_index = S_DICT_params_structure['parameter'].index( this_param )
        this_param_keys = S_DICT_params_structure['index_list'][this_param_index]
        #
        if len( this_scenario_data[ this_param ]['value'] ) != 0:
            #
#            f.write( 'param ' + this_param + ':=\n' )
            if len(this_param_keys) != 2:
                g.write( 'param ' + this_param + ' default ' + str( default_value ) + ' :=\n' )
            else:
                g.write( 'param ' + this_param + ' default ' + str( default_value ) + ' :\n' )
            #
            #-----------------------------------------#
            if len(this_param_keys) == 2: #$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
                # get the last and second last parameters of the list:
                last_set_element = this_scenario_data[ this_param ][ this_param_keys[-1] ] # header_indices.index( this_param_keys[-1] ) ]
                last_set_element_unique = [] # list( set( last_set_element ) )
                for u in range( len( last_set_element ) ):
                    if last_set_element[u] not in last_set_element_unique:
                        last_set_element_unique.append( last_set_element[u] )
                #
                for y in range( len( last_set_element_unique ) ):
                    g.write( str( last_set_element_unique[y] ) + ' ')
                g.write(':=\n')
                #
                second_last_set_element = this_scenario_data[ this_param ][ this_param_keys[-2] ] # header_indices.index( this_param_keys[-2] ) ]
                second_last_set_element_unique = [] # list( set( second_last_set_element ) )
                for u in range( len( second_last_set_element ) ):
                    if second_last_set_element[u] not in second_last_set_element_unique:
                        second_last_set_element_unique.append( second_last_set_element[u] )
                #
                for s in range( len( second_last_set_element_unique ) ):
                    g.write( second_last_set_element_unique[s] + ' ' )
                    value_indices = [ i for i, x in enumerate( this_scenario_data[ this_param ][ this_param_keys[-2] ] ) if x == str( second_last_set_element_unique[s] ) ]
                    these_values = []
                    for val in range( len( value_indices ) ):
                        these_values.append( this_scenario_data[ this_param ]['value'][ value_indices[val] ] )
                    for val in range( len( these_values ) ):
                        g.write( str( these_values[val] ) + ' ' )
                    g.write('\n') #$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
            #%%%
            if len(this_param_keys) == 3:
                this_set_element_unique_all = []
                for pkey in range( len(this_param_keys)-2 ):
                    for i in range( 2, len(header_indices)-1 ):
                        if header_indices[i] == this_param_keys[pkey]:
                            this_set_element = this_scenario_data[ this_param ][ header_indices[i] ]
                    this_set_element_unique_all.append( list( set( this_set_element ) ) )
                #
                this_set_element_unique_1 = deepcopy( this_set_element_unique_all[0] )
                #
                for n1 in range( len( this_set_element_unique_1 ) ): #$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
                    g.write( '[' + str( this_set_element_unique_1[n1] ) + ',*,*]:\n' )
                    # get the last and second last parameters of the list:
                    last_set_element = this_scenario_data[ this_param ][ this_param_keys[-1] ] # header_indices.index( this_param_keys[-1] ) ]
                    last_set_element_unique = [] # list( set( last_set_element ) )
                    for u in range( len( last_set_element ) ):
                        if last_set_element[u] not in last_set_element_unique:
                            last_set_element_unique.append( last_set_element[u] )
                    #
                    for y in range( len( last_set_element_unique ) ):
                        g.write( str( last_set_element_unique[y] ) + ' ')
                    g.write(':=\n')
                    #
                    second_last_set_element = this_scenario_data[ this_param ][ this_param_keys[-2] ] #header_indices.index( this_param_keys[-2] ) ]
                    second_last_set_element_unique = [] # list( set( second_last_set_element ) )
                    for u in range( len( second_last_set_element ) ):
                        if second_last_set_element[u] not in second_last_set_element_unique:
                            second_last_set_element_unique.append( second_last_set_element[u] )
                    #
                    for s in range( len( second_last_set_element_unique ) ):
                        g.write( second_last_set_element_unique[s] + ' ' )
                        #
                        value_indices_s = [ i for i, x in enumerate( this_scenario_data[ this_param ][ this_param_keys[-2] ] ) if x == str( second_last_set_element_unique[s] ) ]
                        value_indices_n1 = [ i for i, x in enumerate( this_scenario_data[ this_param ][ this_param_keys[0] ] ) if x == str( this_set_element_unique_1[n1] ) ]
                        #
                        r_index = set(value_indices_s) & set(value_indices_n1)
                        #
                        value_indices = list( r_index )
                        value_indices.sort()
                        #
                        these_values = []
                        for val in range( len( value_indices ) ):
                            try:
                                these_values.append( this_scenario_data[ this_param ]['value'][ value_indices[val] ] )
                            except:
                                print( this_param, val )
                        for val in range( len( these_values ) ):
                            g.write( str( these_values[val] ) + ' ' )
                        g.write('\n') #$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
            #%%%
            if len(this_param_keys) == 4:
                this_set_element_unique_all = []
                for pkey in range( len(this_param_keys)-2 ):
                    for i in range( 2, len(header_indices)-1 ):
                        if header_indices[i] == this_param_keys[pkey]:
                            this_set_element = this_scenario_data[ this_param ][ header_indices[i] ]
                            this_set_element_unique_all.append( list( set( this_set_element ) ) )
                #
                this_set_element_unique_1 = deepcopy( this_set_element_unique_all[0] )
                this_set_element_unique_2 = deepcopy( this_set_element_unique_all[1] )
                #
                for n1 in range( len( this_set_element_unique_1 ) ):
                    for n2 in range( len( this_set_element_unique_2 ) ): #$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
                        g.write( '[' + str( this_set_element_unique_1[n1] ) + ',' + str( this_set_element_unique_2[n2] ) + ',*,*]:\n' )
                        # get the last and second last parameters of the list:
                        last_set_element = this_scenario_data[ this_param ][ this_param_keys[-1] ] # header_indices.index( this_param_keys[-1] ) ]
                        last_set_element_unique = [] # list( set( last_set_element ) )
                        for u in range( len( last_set_element ) ):
                            if last_set_element[u] not in last_set_element_unique:
                                last_set_element_unique.append( last_set_element[u] )
                        #
                        for y in range( len( last_set_element_unique ) ):
                            g.write( str( last_set_element_unique[y] ) + ' ')
                        g.write(':=\n')
                        #
                        second_last_set_element = this_scenario_data[ this_param ][ this_param_keys[-2] ] # header_indices.index( this_param_keys[-2] ) ]
                        second_last_set_element_unique = [] # list( set( second_last_set_element ) )
                        for u in range( len( second_last_set_element ) ):
                            if second_last_set_element[u] not in second_last_set_element_unique:
                                second_last_set_element_unique.append( second_last_set_element[u] )
                        #
                        for s in range( len( second_last_set_element_unique ) ):
                            g.write( second_last_set_element_unique[s] + ' ' )
                            #
                            value_indices_s = [ i for i, x in enumerate( this_scenario_data[ this_param ][ this_param_keys[-2] ] ) if x == str( second_last_set_element_unique[s] ) ]
                            value_indices_n1 = [ i for i, x in enumerate( this_scenario_data[ this_param ][ this_param_keys[0] ] ) if x == str( this_set_element_unique_1[n1] ) ]
                            value_indices_n2 = [ i for i, x in enumerate( this_scenario_data[ this_param ][ this_param_keys[1] ] ) if x == str( this_set_element_unique_2[n2] ) ]
                            r_index = set(value_indices_s) & set(value_indices_n1) & set(value_indices_n2)
                            value_indices = list( r_index )
                            value_indices.sort()
                            #
                            # these_values = this_scenario_data[ this_param ]['value'][ value_indices[0]:value_indices[-1]+1 ]
                            these_values = []
                            for val in range( len( value_indices ) ):
                                these_values.append( this_scenario_data[ this_param ]['value'][ value_indices[val] ] )
                            for val in range( len( these_values ) ):
                                g.write( str( these_values[val] ) + ' ' )
                            g.write('\n') #$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
            #%%%
            if len(this_param_keys) == 5:
                print(5, p, this_param)
                this_set_element_unique_all = []
                last_set_element_unique = []
                second_last_set_element_unique = []
                for pkey in range(len(this_param_keys)-2):
                    for i in range(2, len(header_indices)-1):
                        if header_indices[i] == this_param_keys[pkey]:
                            this_set_element = this_scenario_data[this_param][header_indices[i]]
                            this_set_element_unique_all.append(list(set(this_set_element)))

                #
                this_set_element_unique_1 = deepcopy(this_set_element_unique_all[0])
                this_set_element_unique_2 = deepcopy(this_set_element_unique_all[1])
                this_set_element_unique_3 = deepcopy(this_set_element_unique_all[2])

                last_set_element = np.array(this_scenario_data[this_param][this_param_keys[-1]])
                last_set_element_unique = np.unique(last_set_element)

                second_last_set_element = np.array(this_scenario_data[this_param][this_param_keys[-2]])

                second_last_set_element_unique = np.unique(second_last_set_element)

                long_list1 = this_scenario_data[this_param][this_param_keys[1]]
                long_list2 = this_scenario_data[this_param][this_param_keys[2]]
                concat_result = list(map(lambda x, y: x + '-' + y, long_list1, long_list2))
                concat_result_set = list(set(concat_result))

                short_list1, short_list2 = \
                    zip(*(s.split('-') for s in concat_result_set))
                short_list1_set = list(set(short_list1))
                short_list2_set = list(set(short_list2))

                for n1 in range(len(this_set_element_unique_1)):
                    # for n2 in range(len(this_set_element_unique_2)):
                    # for n3 in range(len(this_set_element_unique_3)):
                    for nx in range(len(concat_result_set)):
                        n1_faster = concat_result_set[nx].split('-')[0]
                        n2_faster = concat_result_set[nx].split('-')[1]

                        for s in range(len(second_last_set_element_unique)):
                            value_indices_s = [i for i, x in enumerate(this_scenario_data[this_param][this_param_keys[-2]]) if x == str(second_last_set_element_unique[s])]
                            value_indices_n1 = [i for i, x in enumerate(this_scenario_data[this_param][this_param_keys[0]]) if x == str(this_set_element_unique_1[n1])]
                            value_indices_n2 = [i for i, x in enumerate(this_scenario_data[this_param][this_param_keys[1]]) if x == str(n1_faster)]
                            value_indices_n3 = [i for i, x in enumerate(this_scenario_data[this_param][this_param_keys[2]]) if x == str(n2_faster)]

                            r_index = set(value_indices_s) & set(value_indices_n1) & set(value_indices_n2) & set(value_indices_n3)
                            value_indices = list(r_index)
                            value_indices.sort()

                            if len(value_indices) != 0:
                                g.write('[' + str(this_set_element_unique_1[n1]) + ',' + str(n1_faster) + ',' + str(n2_faster) + ',*,*]:\n')

                                for y in range(len(last_set_element_unique)):
                                    g.write(str(last_set_element_unique[y]) + ' ')
                                g.write(':=\n')

                                g.write(second_last_set_element_unique[s] + ' ')

                                these_values = []
                                for val in range(len(value_indices)):
                                    these_values.append(this_scenario_data[this_param]['value'][value_indices[val]])
                                for val in range(len(these_values)):
                                    g.write(str(these_values[val]) + ' ')
                                g.write('\n')

                            else:
                                print('Never happens right?')
                                sys.exit()
            #
            #g.write('\n') 
            #-----------------------------------------#
            g.write( ';\n\n' )
    #
    # Print parameters with it default value
    for parm in params['params_inputs_data']:																																		   
        g.write(f'param {parm} default ' + str(params['default_val_params'][parm]) + ' :=\n;\n')
    for sna, val_parm in params['params_inputs_data_excep_by_scen'].items():																									 
        if scenario_list[scen] == params[sna]:
            if params['params_inputs_data_excep_by_scen'][params[sna]] != []:
                for parm_2 in val_parm:																																			   
                    g.write(f'param {parm_2} default ' + str(params['default_val_params'][parm_2]) + ' :=\n;\n')
    #
    g.write('#\n' + 'end;\n')
    #
    g.close()
    # print(sorted_combined_params)
    # sys.exit()
    #
    ###########################################################################################################################
    # Furthermore, we must print the inputs separately for faste deployment of the input matrix:
    #
    if scenario_list[scen] != params['BAU'] and params['Use_Blend_Shares_B1']:
        this_Blend_Shares_data = Blend_Shares[ scenario_list[scen] ]
    #
    basic_header_elements = params['basic_header_elements']
    #
    parameters_to_print = params['parameters_to_print']
    #
    more_params = params['more_params']
    #
    filter_params = params['filter_params']
    #
    input_params_table_headers = basic_header_elements + parameters_to_print + more_params + filter_params
    all_data_row = []
    all_data_row_partial = []
    #
    combination_list = []
    synthesized_all_data_row = []
    #
    # memory elements:
    f_unique_list, f_counter, f_counter_list, f_unique_counter_list = [], 1, [], []
    t_unique_list, t_counter, t_counter_list, t_unique_counter_list = [], 1, [], []
    e_unique_list, e_counter, e_counter_list, e_unique_counter_list = [], 1, [], []
    l_unique_list, l_counter, l_counter_list, l_unique_counter_list = [], 1, [], []
    y_unique_list, y_counter, y_counter_list, y_unique_counter_list = [], 1, [], []
    #
    for p in range( len( parameters_to_print ) ):
        #
        this_p_index = S_DICT_params_structure[ 'parameter' ].index( parameters_to_print[p] )
        this_p_index_list = S_DICT_params_structure[ 'index_list' ][ this_p_index ]
        #
        for n in range( 0, len( this_scenario_data[ parameters_to_print[p] ][ 'value' ] ) ):
            #
            single_data_row = []
            single_data_row_partial = []
            #
            single_data_row.append( 0 )
            single_data_row.append( scen )
            single_data_row.append( scenario_list[scen] )
            #
            strcode = ''
            #
            if 'f' in this_p_index_list:
                single_data_row.append( this_scenario_data[ parameters_to_print[p] ][ 'f' ][n] ) # Filling FUEL if necessary
                if single_data_row[-1] not in f_unique_list:
                    f_unique_list.append( single_data_row[-1] )
                    f_counter_list.append( f_counter )
                    f_unique_counter_list.append( f_counter )
                    f_counter += 1
                else:
                    f_counter_list.append( f_unique_counter_list[ f_unique_list.index( single_data_row[-1] ) ] )
                strcode += str(f_counter_list[-1])
            else:
                single_data_row.append( '' )
                strcode += '0'
            #
            if 't' in this_p_index_list:
                single_data_row.append( this_scenario_data[ parameters_to_print[p] ][ 't' ][n] ) # Filling TECHNOLOGY if necessary
                if single_data_row[-1] not in t_unique_list:
                    t_unique_list.append( single_data_row[-1] )
                    t_counter_list.append( t_counter )
                    t_unique_counter_list.append( t_counter )
                    t_counter += 1
                else:
                    t_counter_list.append( t_unique_counter_list[ t_unique_list.index( single_data_row[-1] ) ] )
                strcode += str(t_counter_list[-1])
            else:
                single_data_row.append( '' )
                strcode += '0'
            #
            if 'e' in this_p_index_list:
                single_data_row.append( this_scenario_data[ parameters_to_print[p] ][ 'e' ][n] ) # Filling EMISSION if necessary
                if single_data_row[-1] not in e_unique_list:
                    e_unique_list.append( single_data_row[-1] )
                    e_counter_list.append( e_counter )
                    e_unique_counter_list.append( e_counter )
                    e_counter += 1
                else:
                    e_counter_list.append( e_unique_counter_list[ e_unique_list.index( single_data_row[-1] ) ] )
                strcode += str(e_counter_list[-1])
            else:
                single_data_row.append( '' )
                strcode += '0'
            #
            if 'l' in this_p_index_list:
                single_data_row.append( this_scenario_data[ parameters_to_print[p] ][ 'l' ][n] ) # Filling SEASON if necessary
                if single_data_row[-1] not in l_unique_list:
                    l_unique_list.append( single_data_row[-1] )
                    l_counter_list.append( l_counter )
                    l_unique_counter_list.append( l_counter )
                    l_counter += 1
                else:
                    l_counter_list.append( l_unique_counter_list[ l_unique_list.index( single_data_row[-1] ) ] )
                strcode += str(l_counter_list[-1])
            else:
                single_data_row.append( '' )
                strcode += '0'
            #
            if 'y' in this_p_index_list:
                single_data_row.append( this_scenario_data[ parameters_to_print[p] ][ 'y' ][n] ) # Filling YEAR if necessary
                if single_data_row[-1] not in y_unique_list:
                    y_unique_list.append( single_data_row[-1] )
                    y_counter_list.append( y_counter )
                    y_unique_counter_list.append( y_counter )
                    y_counter += 1
                else:
                    y_counter_list.append( y_unique_counter_list[ y_unique_list.index( single_data_row[-1] ) ] )
                strcode += str(y_counter_list[-1])
            else:
                single_data_row.append( '' )
                strcode += '0'
            #
            this_combination_str = str(1) + strcode # deepcopy( single_data_row )
            this_combination = int( this_combination_str )
            #
            for aux_p in range( len(basic_header_elements), len(basic_header_elements) + len( parameters_to_print ) ):
                if aux_p == p + len(basic_header_elements):
                    single_data_row.append( this_scenario_data[ parameters_to_print[p] ][ 'value' ][n] ) # Filling the correct data point
                    single_data_row_partial.append( this_scenario_data[ parameters_to_print[p] ][ 'value' ][n] )
                else:
                    single_data_row.append( '' )
                    single_data_row_partial.append( '' )
            #
            #---------------------------------------------------------------------------------#
            for aide in range( len(more_params)+len(filter_params) ):
                single_data_row.append( '' )
                single_data_row_partial.append( '' )
            #---------------------------------------------------------------------------------#
            #
            all_data_row.append( single_data_row )
            all_data_row_partial.append( single_data_row_partial )
            #
            if this_combination not in combination_list:
                combination_list.append( this_combination )
                synthesized_all_data_row.append( single_data_row )
            else:
                ref_combination_index = combination_list.index( this_combination )
                ref_parameter_index = input_params_table_headers.index( parameters_to_print[p] )
                synthesized_all_data_row[ ref_combination_index ][ ref_parameter_index ] = deepcopy( single_data_row_partial[ ref_parameter_index-len( basic_header_elements ) ] )
                #
            #
            ##################################################################################################################
            #
            if parameters_to_print[p] in ['EmissionActivityRatio']:
                ref_index = combination_list.index( this_combination )
                this_data_row = deepcopy( synthesized_all_data_row[ ref_index ] ) # this must be updated in a further position of the list
                #
                this_tech = this_data_row[4]
                this_emission = this_data_row[5]
                this_year = this_data_row[7]
                this_year_index = time_range_vector.index( int( this_year ) )
                #
                if scenario_list[scen] != params['BAU'] and params['Use_Blend_Shares_B1']:
                    if this_tech in list( this_Blend_Shares_data.keys() ):
                        if this_emission in list( this_Blend_Shares_data[ this_tech ].keys() ):
                            #
                            this_Blend_Shares = this_Blend_Shares_data[ this_tech ][ this_emission ]
                            #
                            var_position_index = input_params_table_headers.index( params['biofuel_shares'] )
                            this_data_row[ var_position_index ] = round(this_Blend_Shares[ this_year_index ], params['round_#'])
                            #
                        #
                    #
                #
                else:
                    var_position_index = input_params_table_headers.index( params['biofuel_shares'] )
                    this_data_row[ var_position_index ] = 0.0
                #
                synthesized_all_data_row[ ref_index ] = deepcopy( this_data_row )
                #
            #
            ##################################################################################################################
            #
            if parameters_to_print[p] in params['param_print']:
                ref_index = combination_list.index( this_combination )
                this_data_row = deepcopy( synthesized_all_data_row[ ref_index ] ) # this must be updated in a further position of the list
                #
                this_tech = this_data_row[4]
                if this_tech in list( Fleet_Groups_inv.keys() ):
                    group_tech = Fleet_Groups_inv[ this_tech ]
                    #
                    this_year = this_data_row[7]
                    this_year_index = time_range_vector.index( int( this_year ) )
                    #
                    ref_var_position_index = input_params_table_headers.index( parameters_to_print[p] )
                    #
                    if params['trains'] not in group_tech  and params['group_tech'] not in group_tech:
                        driven_distance = Reference_driven_distance[ scenario_list[scen] ][ group_tech ][ this_year_index ]
                        #
                        if parameters_to_print[p] == 'TotalAnnualMaxCapacity' or parameters_to_print[p] == 'FixedCost': # this will overwrite for zero-carbon techs, but will not fail.
                            var_position_index = input_params_table_headers.index( params['dist_driven'] )
                            this_data_row[ var_position_index ] = round(driven_distance, params['round_#'])
                        #
                        if parameters_to_print[p] == 'CapitalCost':
                            var_position_index = input_params_table_headers.index( 'UnitCapitalCost (USD)' )
                            this_data_row[ var_position_index ] = round( (10**6)*float( this_data_row[ ref_var_position_index ] )*driven_distance/(10**9), params['round_#'])
                        #
                        if parameters_to_print[p] == 'FixedCost':
                            var_position_index = input_params_table_headers.index( 'UnitFixedCost (USD)' )
                            this_data_row[ var_position_index ] = round( (10**6)*float( this_data_row[ ref_var_position_index ] )*driven_distance/(10**9), params['round_#'])
                        #
                        synthesized_all_data_row[ ref_index ] = deepcopy( this_data_row )
                        #
                    #
                    ###################################################################################################
                    #
                #
                if parameters_to_print[p] == 'FixedCost':
                    # Creating convinience filters for the analysis of model outputs:
                    if params['tr'] in this_tech[:2]:
                        #
                        ref_index = combination_list.index( this_combination )
                        this_data_row = deepcopy( synthesized_all_data_row[ ref_index ] ) # this must be updated in a further position of the list
                        #
                        var_position_index = input_params_table_headers.index( params['filt_fuel_type'] )
                        ############################################
                        # By Fuel Type
                        for r in range( len( df_fuel_2_code_fuel_list ) ):
                            if df_fuel_2_code_fuel_list[r] in this_tech:
                                this_data_row[ var_position_index ] = df_fuel_2_code_plain_english[ r ]
                                break
                        ############################################
                        var_position_index = input_params_table_headers.index( params['filt_veh_type'] )
                        # By vehicle type
                        for r in range( len( df_tech_2_code_fuel_list ) ):
                            if df_tech_2_code_fuel_list[r] in this_tech:
                                this_data_row[ var_position_index ] = df_tech_2_code_plain_english[ r ]
                                break
                        #
                        synthesized_all_data_row[ ref_index ] = deepcopy( this_data_row )
                        #
                    #
                #
            #
        #
    #
    ###########################################################################################################################
    #
    # with open( params['Executables'] + '/' + str( scenario_list[scen] ) + '_0' + '/' + str( scenario_list[scen] ) + '_0' + '_Input.csv', 'w', newline = '') as param_csv:
    with open( os.path.join(params['Executables'], f'{str( scenario_list[scen] )}_0', f'{str( scenario_list[scen] )}_0_Input.csv'), 'w', newline = '') as param_csv:
        csvwriter = csv.writer(param_csv, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        # Print the header:
        csvwriter.writerow( input_params_table_headers )
        # Print all else:
        for n in range( len( synthesized_all_data_row ) ):
            csvwriter.writerow( synthesized_all_data_row[n] )
    #

def generalized_logistic_curve(x, L, Q, k, M):
  return L/( 1 + Q*math.exp( -k*( x-M ) ) )
#
def interpolation_blend( start_blend_point, mid_blend_point, final_blend_point, value_list, time_range_vector ):
    #
    start_blend_year, start_blend_value = start_blend_point[0], start_blend_point[1]/100
    final_blend_year, final_blend_value = final_blend_point[0], final_blend_point[1]/100
    #
    # Now we need to interpolate:
    x_coord_tofill = [] # these are indices that are NOT known - to interpolate
    xp_coord_known = [] # these are known indices - use for interpolation
    fp_coord_known = [] # these are the values known to interpolate the whole series
    #
    if mid_blend_point != 'None': # this means we must use a linear interpolation
        mid_blend_year, mid_blend_value = mid_blend_point[0], mid_blend_point[1]/100
        #
        for t in range( len( time_range_vector ) ):
            something_to_fill = False
            #
            if time_range_vector[t] < start_blend_year:
                fp_coord_known.append( 0.0 )
            #
            if time_range_vector[t] == start_blend_year:
                fp_coord_known.append( start_blend_value )
            #
            if ( time_range_vector[t] > start_blend_year and time_range_vector[t] < mid_blend_year ) or ( time_range_vector[t] > mid_blend_year and time_range_vector[t] < final_blend_year ):
                something_to_fill = True
            #
            if time_range_vector[t] == mid_blend_year:
                fp_coord_known.append( mid_blend_value )
            #
            if time_range_vector[t] == final_blend_year or time_range_vector[t] > final_blend_year:
                fp_coord_known.append( final_blend_value )
            #
            if something_to_fill == True:
                x_coord_tofill.append( t )
            else:
                xp_coord_known.append( t ) # means this value was stored
            #
            y_coord_filled = list( np.interp( x_coord_tofill, xp_coord_known, fp_coord_known ) )
            #
            interpolated_values = []
            for coord in range( len( time_range_vector ) ):
                if coord in xp_coord_known:
                    value_index = xp_coord_known.index(coord)
                    interpolated_values.append( float( fp_coord_known[value_index] ) )
                elif coord in x_coord_tofill:
                    value_index = x_coord_tofill.index(coord)
                    interpolated_values.append( float( y_coord_filled[value_index] ) )
            #
        #
    else:
        #
        for t in range( len( time_range_vector ) ):
            something_to_fill = False
            #
            if time_range_vector[t] < start_blend_year:
                fp_coord_known.append( 0.0 )
            #
            if time_range_vector[t] == start_blend_year:
                fp_coord_known.append( start_blend_value )
            #
            if ( time_range_vector[t] > start_blend_year and time_range_vector[t] < final_blend_year ):
                something_to_fill = True
            #
            if time_range_vector[t] == final_blend_year or time_range_vector[t] > final_blend_year:
                fp_coord_known.append( final_blend_value )
            #
            if something_to_fill == True:
                x_coord_tofill.append( t )
            else:
                xp_coord_known.append( t ) # means this value was stored
            #
            y_coord_filled = list( np.interp( x_coord_tofill, xp_coord_known, fp_coord_known ) )
            #
            interpolated_values = []
            for coord in range( len( time_range_vector ) ):
                if coord in xp_coord_known:
                    value_index = xp_coord_known.index(coord)
                    interpolated_values.append( float( fp_coord_known[value_index] ) )
                elif coord in x_coord_tofill:
                    value_index = x_coord_tofill.index(coord)
                    interpolated_values.append( float( y_coord_filled[value_index] ) )
            #
        #
    #
    new_value_list = []
    for n in range( len( value_list ) ):
        new_value_list.append( value_list[n]*( 1-interpolated_values[n] ) )
    new_value_list_rounded = [ round(elem, params['round_#']) for elem in new_value_list ]
    biofuel_shares = [ round(elem, params['round_#']) for elem in interpolated_values ]
    #
    return new_value_list_rounded, biofuel_shares
    #
#
def interpolation_non_linear_final_flexible( time_list, value_list, new_relative_final_value, Flexible_Initial_Year_of_Uncertainty ):
    # Rememeber that the 'old_relative_final_value' is 1
    old_relative_final_value = 1
    new_value_list = []
    # We select a list that goes from the "Flexible_Initial_Year_of_Uncertainty" to the Final Year of the Time Series
    initial_year_index = time_list.index( Flexible_Initial_Year_of_Uncertainty )
    fraction_time_list = time_list[initial_year_index:]
    fraction_value_list = value_list[initial_year_index:]
    # We now perform the 'non-linear OR linear adjustment':
    xdata = [ fraction_time_list[i] - fraction_time_list[0] for i in range( len( fraction_time_list ) ) ]
    ydata = [ float( fraction_value_list[i] ) for i in range( len( fraction_value_list ) ) ]
    delta_ydata = [ ydata[i]-ydata[i-1] for i in range( 1,len( ydata ) ) ]
    #
    m_original = ( ydata[-1]-ydata[0] ) / ( xdata[-1]-xdata[0] )
    #
    m_new = ( ydata[-1]*(new_relative_final_value/old_relative_final_value) - ydata[0] ) / ( xdata[-1]-xdata[0] )
    #
    if float(m_original) == 0.0:
        delta_ydata_new = [m_new for i in range( 1,len( ydata ) ) ]
    else:
        delta_ydata_new = [ (m_new/m_original)*(ydata[i]-ydata[i-1]) for i in range( 1,len( ydata ) ) ]
    #
    ydata_new = [ 0 for i in range( len( ydata ) ) ]
    ydata_new[0] = ydata[0]
    for i in range( 0, len( delta_ydata ) ):
        ydata_new[i+1] = ydata_new[i] + delta_ydata_new[i]
    #
    # We now recreate the new_value_list considering the fraction before and after the Flexible_Initial_Year_of_Uncertainty
    fraction_list_counter = 0
    for n in range( len( time_list ) ):
        if time_list[n] >= Flexible_Initial_Year_of_Uncertainty:
            new_value_list.append( ydata_new[ fraction_list_counter ] )
            fraction_list_counter += 1
        else:
            new_value_list.append( float( value_list[n] ) )
    #
    # We return the list:
    return new_value_list
#
def interpolation_non_linear_final( time_list, value_list, new_relative_final_value ):
    # Rememeber that the 'old_relative_final_value' is 1
    old_relative_final_value = 1
    new_value_list = []
    # We select a list that goes from the "Initial_Year_of_Uncertainty" to the Final Year of the Time Series
    initial_year_index = time_list.index( Initial_Year_of_Uncertainty )
    fraction_time_list = time_list[initial_year_index:]
    fraction_value_list = value_list[initial_year_index:]
    # We now perform the 'non-linear OR linear adjustment':
    xdata = [ fraction_time_list[i] - fraction_time_list[0] for i in range( len( fraction_time_list ) ) ]
    ydata = [ float( fraction_value_list[i] ) for i in range( len( fraction_value_list ) ) ]
    delta_ydata = [ ydata[i]-ydata[i-1] for i in range( 1,len( ydata ) ) ]
    #
    m_original = ( ydata[-1]-ydata[0] ) / ( xdata[-1]-xdata[0] )
    #
    m_new = ( ydata[-1]*(new_relative_final_value/old_relative_final_value) - ydata[0] ) / ( xdata[-1]-xdata[0] )
    #
    if float(m_original) == 0.0:
        delta_ydata_new = [m_new for i in range( 1,len( ydata ) ) ]
    else:
        delta_ydata_new = [ (m_new/m_original)*(ydata[i]-ydata[i-1]) for i in range( 1,len( ydata ) ) ]
    #
    ydata_new = [ 0 for i in range( len( ydata ) ) ]
    ydata_new[0] = ydata[0]
    for i in range( 0, len( delta_ydata ) ):
        ydata_new[i+1] = ydata_new[i] + delta_ydata_new[i]
    #
    # We now recreate the new_value_list considering the fraction before and after the Initial_Year_of_Uncertainty
    fraction_list_counter = 0
    for n in range( len( time_list ) ):
        if time_list[n] >= Initial_Year_of_Uncertainty:
            new_value_list.append( ydata_new[ fraction_list_counter ] )
            fraction_list_counter += 1
        else:
            new_value_list.append( float( value_list[n] ) )
    #
    # We return the list:
    return new_value_list
#
def linear_interpolation_time_series( time_range, known_years, known_values ):
    # Now we need to interpolate:
    x_coord_tofill = [] # these are indices that are NOT known - to interpolate
    xp_coord_known = [] # these are known indices - use for interpolation
    fp_coord_known = [] # these are the values known to interpolate the whole series
    #
    min_t = min( known_years )
    for t in range( len( time_range ) ):
        if ( time_range[t] in known_years ) or ( time_range[t] < min_t ):
            xp_coord_known.append( t )
            #
            if time_range[t] < min_t:
                fp_coord_known.append( 0.0 )
            else:
                future_year_index = known_years.index( time_range[t] )
                fp_coord_known.append( float(known_values[ future_year_index ]) )
            #
        else:
            x_coord_tofill.append( t )
    #
    y_coord_filled = list( np.interp( x_coord_tofill, xp_coord_known, fp_coord_known ) )
    #
    interpolated_values = []
    for coord in range( len( time_range ) ):
        if coord in xp_coord_known:
            value_index = xp_coord_known.index(coord)
            interpolated_values.append( float( fp_coord_known[value_index] ) )
        elif coord in x_coord_tofill:
            value_index = x_coord_tofill.index(coord)
            interpolated_values.append( float( y_coord_filled[value_index] ) )
    #
    return interpolated_values
#
def csv_printer_parallel(s, scenario_list, stable_scenarios, basic_header_elements, parameters_to_print, S_DICT_params_structure, params):
    #
    input_params_table_headers = basic_header_elements + parameters_to_print
    all_data_row = []
    all_data_row_partial = []
    #
    combination_list = []
    synthesized_all_data_row = []
    #
    # memory elements:
    f_unique_list, f_counter, f_counter_list, f_unique_counter_list = [], 1, [], []
    t_unique_list, t_counter, t_counter_list, t_unique_counter_list = [], 1, [], []
    e_unique_list, e_counter, e_counter_list, e_unique_counter_list = [], 1, [], []
    l_unique_list, l_counter, l_counter_list, l_unique_counter_list = [], 1, [], []
    y_unique_list, y_counter, y_counter_list, y_unique_counter_list = [], 1, [], []
    #
    each_parameter_header = params['sets']
    each_parameter_all_data_row = {}
    #
    print( '    ', s+1, ' // Printing scenarios: ', scenario_list[s] )
    each_parameter_all_data_row.update( { scenario_list[s]:{} } )
    #
    for p in range( len( parameters_to_print ) ):
        each_parameter_all_data_row[ scenario_list[s] ].update( { parameters_to_print[p]:[] } )
        #
        this_p_index = S_DICT_params_structure[ 'parameter' ].index( parameters_to_print[p] )
        this_p_index_list = S_DICT_params_structure[ 'index_list' ][ this_p_index ]
        #
        for n in range( 0, len( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'value' ] ) ):
            #
            single_data_row = []
            single_data_row_partial = []
            #
            single_data_row.append( 0 ) # Filling Future.ID
            single_data_row.append( s ) # Filling Strategy.ID
            single_data_row.append( scenario_list[s] ) # Filling Strategy
            #
            strcode = ''
            #
            if 'f' in this_p_index_list:
                single_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'f' ][n] ) # Filling FUEL if necessary
                if single_data_row[-1] not in f_unique_list:
                    f_unique_list.append( single_data_row[-1] )
                    f_counter_list.append( f_counter )
                    f_unique_counter_list.append( f_counter )
                    f_counter += 1
                else:
                    f_counter_list.append( f_unique_counter_list[ f_unique_list.index( single_data_row[-1] ) ] )
                strcode += str(f_counter_list[-1])
            else:
                single_data_row.append( '' )
                strcode += '0'
            #
            if 't' in this_p_index_list:
                try:
                    single_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 't' ][n] ) # Filling TECHNOLOGY if necessary
                except Exception:
                    print('---------------------------')
                    print(scenario_list[s], parameters_to_print[p], n, len(stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 't' ]))
                    print('---------------------------')
                    sys.exit()

                if single_data_row[-1] not in t_unique_list:
                    t_unique_list.append( single_data_row[-1] )
                    t_counter_list.append( t_counter )
                    t_unique_counter_list.append( t_counter )
                    t_counter += 1
                else:
                    t_counter_list.append( t_unique_counter_list[ t_unique_list.index( single_data_row[-1] ) ] )
                strcode += str(t_counter_list[-1])
            else:
                single_data_row.append( '' )
                strcode += '0'
            #
            if 'e' in this_p_index_list:
                single_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'e' ][n] ) # Filling EMISSION if necessary
                if single_data_row[-1] not in e_unique_list:
                    e_unique_list.append( single_data_row[-1] )
                    e_counter_list.append( e_counter )
                    e_unique_counter_list.append( e_counter )
                    e_counter += 1
                else:
                    e_counter_list.append( e_unique_counter_list[ e_unique_list.index( single_data_row[-1] ) ] )
                strcode += str(e_counter_list[-1])
            else:
                single_data_row.append( '' )
                strcode += '0'
            #
            if 'l' in this_p_index_list:
                single_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'l' ][n] ) # Filling SEASON if necessary
                if single_data_row[-1] not in l_unique_list:
                    l_unique_list.append( single_data_row[-1] )
                    l_counter_list.append( l_counter )
                    l_unique_counter_list.append( l_counter )
                    l_counter += 1
                else:
                    l_counter_list.append( l_unique_counter_list[ l_unique_list.index( single_data_row[-1] ) ] )
                strcode += str(l_counter_list[-1])
            else:
                single_data_row.append( '' )
                strcode += '000' # this is done to avoid repeated characters
            #
            if 'y' in this_p_index_list:
                single_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'y' ][n] ) # Filling YEAR if necessary
                if single_data_row[-1] not in y_unique_list:
                    y_unique_list.append( single_data_row[-1] )
                    y_counter_list.append( y_counter )
                    y_unique_counter_list.append( y_counter )
                    y_counter += 1
                else:
                    y_counter_list.append( y_unique_counter_list[ y_unique_list.index( single_data_row[-1] ) ] )
                strcode += str(y_counter_list[-1])
            else:
                single_data_row.append( '' )
                strcode += '0'
            #
            this_combination_str = str(10) + str(s) + strcode # deepcopy( single_data_row )
            #
            this_combination = int( this_combination_str )
            #
            for aux_p in range( len(basic_header_elements), len(basic_header_elements) + len( parameters_to_print ) ):
                if aux_p == p + len(basic_header_elements):
                    single_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'value' ][n] ) # Filling the correct data point
                    single_data_row_partial.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'value' ][n] )
                else:
                    single_data_row.append( '' )
                    single_data_row_partial.append( '' )
            #
            all_data_row.append( single_data_row )
            all_data_row_partial.append( single_data_row_partial )
            #
            if this_combination not in combination_list:
                combination_list.append( this_combination )
                synthesized_all_data_row.append( single_data_row )
            else:
                ref_combination_index = combination_list.index( this_combination )
                ref_parameter_index = input_params_table_headers.index( parameters_to_print[p] )
                synthesized_all_data_row[ ref_combination_index ][ ref_parameter_index ] = deepcopy( single_data_row_partial[ ref_parameter_index-len( basic_header_elements ) ] )
            #
            # Let us proceed with the list for the each_parameter list:
            this_each_parameter_all_data_row = [] # each_parameter_header = [ 'PARAMETER','Scenario','REGION','TECHNOLOGY','FUEL','EMISSION','MODE_OF_OPERATION','YEAR','TIMESLICE','SEASON','DAYTYPE','DAILYTIMEBRACKET','STORAGE','Value' ]
            #
            this_each_parameter_all_data_row.append( parameters_to_print[p] ) # Alternatively, fill PARAMETER
            this_each_parameter_all_data_row.append( scenario_list[s] ) # Alternatively, fill Scenario
            if 'r' in this_p_index_list:
                this_each_parameter_all_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'r' ][n] ) # Filling REGION if necessary
            else:
                this_each_parameter_all_data_row.append( '' )
            if 't' in this_p_index_list:
                this_each_parameter_all_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 't' ][n] ) # Filling TECHNOLOGY if necessary
            else:
                this_each_parameter_all_data_row.append( '' )
            if 'f' in this_p_index_list:
                this_each_parameter_all_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'f' ][n] ) # Filling FUEL if necessary
            else:
                this_each_parameter_all_data_row.append( '' )
            if 'e' in this_p_index_list:
                this_each_parameter_all_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'e' ][n] ) # Filling EMISSION if necessary
            else:
                this_each_parameter_all_data_row.append( '' )
            if 'm' in this_p_index_list:
                this_each_parameter_all_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'm' ][n] ) # Filling MODE_OF_OPERATION if necessary
            else:
                this_each_parameter_all_data_row.append( '' )
            if 'y' in this_p_index_list:
                this_each_parameter_all_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'y' ][n] ) # Filling YEAR if necessary
            else:
                this_each_parameter_all_data_row.append( '' )
            if 'l' in this_p_index_list:
                this_each_parameter_all_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'l' ][n] ) # Filling TIMESLICE if necessary
            else:
                this_each_parameter_all_data_row.append( '' )
            this_each_parameter_all_data_row.append( '' )
            this_each_parameter_all_data_row.append( '' )
            this_each_parameter_all_data_row.append( '' )
            this_each_parameter_all_data_row.append( '' )
            this_each_parameter_all_data_row.append( stable_scenarios[ scenario_list[s] ][ parameters_to_print[p] ][ 'value' ][n] )
            #
            each_parameter_all_data_row[ scenario_list[s] ][ parameters_to_print[p] ].append( this_each_parameter_all_data_row )

    #########################################################################################
    print('4: We have printed the inputs. We will store the parameters that the experiment uses and be done.')
    # let us be reminded of the 'each_parameter_header'
    for p in range( len( parameters_to_print ) ):
        param_filename = params['B1_Output_Params'] + str( scenario_list[s] ) + '/' + str( parameters_to_print[p] ) + '.csv'
        with open( param_filename, 'w', newline = '') as param_csv:
            csvwriter = csv.writer(param_csv, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            # Print the header:
            csvwriter.writerow( each_parameter_header )
            # Let us locate the data:
            for n in range( len( each_parameter_all_data_row[ scenario_list[s] ][ parameters_to_print[p] ] ) ):
                csvwriter.writerow( each_parameter_all_data_row[ scenario_list[s] ][ parameters_to_print[p] ][n] )

def get_config_main_path(full_path, base_folder='config_main_files'):
    # Split the path into parts
    parts = full_path.split(os.sep)
    
    # Find the index of the target directory 'osemosys_momf'
    target_index = parts.index('osemosys_momf') if 'osemosys_momf' in parts else None
    
    # If the directory is found, reconstruct the path up to that point
    if target_index is not None:
        base_path = os.sep.join(parts[:target_index + 1])
    else:
        base_path = full_path  # If not found, return the original path
    
    # Append the specified directory to the base path
    appended_path = os.path.join(base_path, base_folder)
    
    return appended_path

def load_and_process_yaml(path):
    """
    Load a YAML file and replace the specific placeholder '${year_apply_discount_rate}' with the year specified in the file.
    
    Args:
    path (str): The path to the YAML file.
    
    Returns:
    dict: The updated data from the YAML file where the specific placeholder is replaced.
    """
    with open(path, 'r') as file:
        # Load the YAML content into 'params'
        params = yaml.safe_load(file)

    # Retrieve the reference year from the YAML file and convert it to string for replacement
    reference_year = str(params['year_apply_discount_rate'])

    # Function to recursively update strings containing the placeholder
    def update_strings(obj):
        if isinstance(obj, dict):
            return {k: update_strings(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [update_strings(element) for element in obj]
        elif isinstance(obj, str):
            # Replace the specific placeholder with the reference year
            return obj.replace('${year_apply_discount_rate}', reference_year)
        else:
            return obj

    # Update all string values in the loaded YAML structure
    updated_params = update_strings(params)

    return updated_params

if __name__ == '__main__':
    #
    start1 = time.time()

    # Read yaml file with parameterization
    file_config_address = get_config_main_path(os.path.abspath(''))
    params = load_and_process_yaml(os.path.join(file_config_address, 'MOMF_B1_exp_manager.yaml'))


    all_years = [ y for y in range( params['year_first_range'] , params['final_year']+1 ) ]
    index_change_year = all_years.index( params['change_year_B1'] )
    initial_year = all_years[0]
    final_year = all_years[-1]
    """
    *Abbreviations:*
    """
    
    #
    # We must open useful GDP data for denominator
    df_gdp_ref = pd.read_excel(params['GDP_REF'], params['GDP'])
    list_gdp_growth_ref = df_gdp_ref[params['GDP_Growth']].tolist()
    list_gdp_ref = df_gdp_ref[params['GDP']].tolist()
    #
    '''''
    ################################# PART 1 #################################
    '''''
    print('1: We have defined some data. Now we will read the parameters we have as reference (or previous parameters) into a dictionary.')
    '''
    # 1.A) We extract the strucute setup of the model based on 'Structure.xlsx'
    '''
    structure_filename = params['B1_model_Struct']
    structure_file = pd.ExcelFile(structure_filename)
    structure_sheetnames = structure_file.sheet_names  # see all sheet names
    sheet_sets_structure = pd.read_excel(open(structure_filename, 'rb'),
                                         header=None,
                                         sheet_name=structure_sheetnames[0])
    sheet_params_structure = pd.read_excel(open(structure_filename, 'rb'),
                                           header=None,
                                           sheet_name=structure_sheetnames[1])
    sheet_vars_structure = pd.read_excel(open(structure_filename, 'rb'),
                                         header=None,
                                         sheet_name=structure_sheetnames[2])

    S_DICT_sets_structure = params['S_DICT_sets_structure']
    for col in range(1,11+1):  # 11 columns
        S_DICT_sets_structure['set'].append( sheet_sets_structure.iat[0, col] )
        S_DICT_sets_structure['initial'].append( sheet_sets_structure.iat[1, col] )
        S_DICT_sets_structure['number_of_elements'].append( int( sheet_sets_structure.iat[2, col] ) )
        #
        element_number = int( sheet_sets_structure.iat[2, col] )
        this_elements_list = []
        if element_number > 0:
            for n in range( 1, element_number+1 ):
                this_elements_list.append( sheet_sets_structure.iat[2+n, col] )
        S_DICT_sets_structure['elements_list'].append( this_elements_list )
    #
    S_DICT_params_structure = params['S_DICT_params_structure']
    param_category_list = []
    for col in range(1,30+1):  # 30 columns
        if str( sheet_params_structure.iat[0, col] ) != '':
            param_category_list.append( sheet_params_structure.iat[0, col] )
        S_DICT_params_structure['category'].append( param_category_list[-1] )
        S_DICT_params_structure['parameter'].append( sheet_params_structure.iat[1, col] )
        S_DICT_params_structure['number_of_elements'].append( int( sheet_params_structure.iat[2, col] ) )
        #
        index_number = int( sheet_params_structure.iat[2, col] )
        this_index_list = []
        for n in range(1, index_number+1):
            this_index_list.append( sheet_params_structure.iat[2+n, col] )
        S_DICT_params_structure['index_list'].append( this_index_list )
    #
    S_DICT_vars_structure = params['S_DICT_vars_structure']
    var_category_list = []
    for col in range(1,43+1):  # 43 columns
        if str( sheet_vars_structure.iat[0, col] ) != '':
            var_category_list.append( sheet_vars_structure.iat[0, col] )
        S_DICT_vars_structure['category'].append( var_category_list[-1] )
        S_DICT_vars_structure['variable'].append( sheet_vars_structure.iat[1, col] )
        S_DICT_vars_structure['number_of_elements'].append( int( sheet_vars_structure.iat[2, col] ) )
        #
        index_number = int( sheet_vars_structure.iat[2, col] )
        this_index_list = []
        for n in range(1, index_number+1):
            this_index_list.append( sheet_vars_structure.iat[2+n, col] )
        S_DICT_vars_structure['index_list'].append( this_index_list )
    #-------------------------------------------#
    # LET'S HAVE THE ENTIRE LIST OF TECHNOLOGIES:
    all_techs_list = S_DICT_sets_structure[ 'elements_list' ][ 1 ]
    #-------------------------------------------#
    '''
    Structural dictionary 1: Notes
    a) We use this dictionary relating a specific technology and a group technology and associate the prefix list
    '''
    Fleet_Groups =              pickle.load( open( params['A1_Outputs'] + params['Pickle_Fleet_Groups'],            "rb" ) )
    ### Ind_Groups =                pickle.load( open( './A1_Outputs/A-O_Ind_Groups.pickle',            "rb" ) )
    ### Fleet_Groups['Techs_Microbuses'] += [ 'TRMBUSHYD' ] # this is an add on, kind of a patch
    Fleet_Groups_Distance =     pickle.load( open( params['A1_Outputs'] + params['Pickle_Fleet_Groups_Dist'],   "rb" ) )
    Fleet_Groups_OR =           pickle.load( open( params['A1_Outputs'] + params['Pickle_Fleet_Groups_OR'],         "rb" ) )
    Fleet_Groups_techs_2_dem =  pickle.load( open( params['A1_Outputs'] + params['Pickle_Fleet_Groups_T2D'],        "rb" ) )
    #
    Fleet_Groups_inv = {}
    for k in range( len( list( Fleet_Groups.keys() ) ) ):
        this_fleet_group_key = list( Fleet_Groups.keys() )[k]
        for e in range( len( Fleet_Groups[ this_fleet_group_key ] ) ):
            this_fleet_group_tech = Fleet_Groups[ this_fleet_group_key ][ e ]
            Fleet_Groups_inv.update( { this_fleet_group_tech:this_fleet_group_key } )
    #
    ### Ind_Groups_inv = {}
    ### for k in range( len( list( Ind_Groups.keys() ) ) ):
    ###     this_ind_group_key = list( Ind_Groups.keys() )[k]
    ###     for e in range( len( Ind_Groups[ this_ind_group_key ] ) ):
    ###         this_ind_group_tech = Ind_Groups[ this_ind_group_key ][ e ]
    ###         Ind_Groups_inv.update( { this_ind_group_tech:this_ind_group_key } )
    #
    group_tech_PUBLIC = []
    group_tech_PRIVATE = []
    group_tech_FREIGHT_HEA = []
    group_tech_FREIGHT_LIG = []
    #
    for t in range( len( list( Fleet_Groups_techs_2_dem.keys() ) ) ):
        tech_key = list( Fleet_Groups_techs_2_dem.keys() )[t]
        this_fuel = Fleet_Groups_techs_2_dem[ tech_key ]
        if params['PRI'] in this_fuel:
            group_tech_PRIVATE.append( tech_key )
        # if params['PUB'] in this_fuel:
        if params['PUB'] in this_fuel and params['group_tech'] not in tech_key and params['trains'] not in tech_key:  # delete the trains and telef here, because we did not add them as alternatives                                                                                                                                                               
            group_tech_PUBLIC.append( tech_key )
        if params['FEAHEA'] in this_fuel:
            group_tech_FREIGHT_HEA.append( tech_key )
        if params['FRELIG'] in this_fuel:
            group_tech_FREIGHT_LIG.append( tech_key )
    #
    group_tech_PASSENGER = group_tech_PUBLIC + group_tech_PRIVATE
    group_tech_FREIGHT = group_tech_FREIGHT_HEA + group_tech_FREIGHT_LIG
    #
    global time_range_vector
    time_range_vector = [ int(i) for i in S_DICT_sets_structure[ 'elements_list' ][0] ]
    '''
    ####################################################################################################################################################
    # 1.B) We finish this sub-part, and proceed to read all the base scenarios.
    '''
    header_row = params['header_row']
    #
    scenario_list_sheet = pd.read_excel( params['B1_Scen_Config'], sheet_name=params['Scens'] )
    scenario_list_all = [ scenario_list_sheet[ 'Name' ].tolist()[i] for i in range( len( scenario_list_sheet[ 'Name' ].tolist() ) ) if scenario_list_sheet[ 'Activated' ].tolist()[i] == 'YES' ]
    scenario_list_reference = [ scenario_list_sheet[ 'Reference' ].tolist()[i] for i in range( len( scenario_list_sheet[ 'Name' ].tolist() ) ) if scenario_list_sheet[ 'Activated' ].tolist()[i] == 'YES' ] # the address to the produced dataset
    scenario_list_based = [ scenario_list_sheet[ 'Based_On' ].tolist()[i] for i in range( len( scenario_list_sheet[ 'Name' ].tolist() ) ) if scenario_list_sheet[ 'Activated' ].tolist()[i] == 'YES' ]
    for i in range( len( scenario_list_sheet[ 'Base' ].tolist() ) ):
        if scenario_list_sheet[ 'Base' ].tolist()[i] == 'YES':
            ref_scenario = scenario_list_sheet[ 'Name' ].tolist()[i]
    #
    scenario_list = [ scenario_list_reference[i] for i in range( len( scenario_list_all ) ) if scenario_list_reference[i] != 'based' ]
    #
    base_configuration_overall = pd.read_excel( params['B1_Scen_Config'], sheet_name=params['Over_params'] )
    base_configuration_waste = pd.read_excel(params['B1_Scen_Config'], sheet_name=params['Waste'])
    base_configuration_ippu = pd.read_excel(params['B1_Scen_Config'], sheet_name=params['IPPU'])
    base_configuration_distance = pd.read_excel( params['B1_Scen_Config'], sheet_name=params['Dis_Levers'] )
    base_configuration_modeshift = pd.read_excel( params['B1_Scen_Config'], sheet_name=params['Mode_Shift'] )
    base_configuration_or = pd.read_excel( params['B1_Scen_Config'], sheet_name=params['Occu_Rate'] )
    base_configuration_adoption = pd.read_excel( params['B1_Scen_Config'], sheet_name=params['Tech_Adop'] )
    base_configuration_electrical = pd.read_excel( params['B1_Scen_Config'], sheet_name=params['Elec'] )
    base_configuration_smartGrid = pd.read_excel( params['B1_Scen_Config'], sheet_name=params['Smart_Grid'] )
    base_configuration_E_and_D = pd.read_excel( params['B1_Scen_Config'], sheet_name=params['Effi'] )
    base_configuration_transport_elasticity = pd.read_excel(params['B1_Scen_Config'], sheet_name=params['TElas'])                                                                                                             
    base_configuration_emi_res = pd.read_excel(params['B1_Scen_Config'], sheet_name=params['Emission_Restriction'])                                                                                       
    #
    all_dataset_address = params['A2_Output_Params']
    ''' 
    # Call the default parameters for later use:
    '''
    dict_default_val_params = params['default_val_params']
    list_param_default_value = pd.DataFrame(list(dict_default_val_params.items()), columns=['Parameter', 'Default_Value'])
    list_param_default_value_params = list( list_param_default_value['Parameter'] )
    list_param_default_value_value = list( list_param_default_value['Default_Value'] )
    #
    global Initial_Year_of_Uncertainty
    for n in range( len( base_configuration_overall.index ) ):
        if str( base_configuration_overall.loc[ n, 'Parameter' ] ) == params['ini_year_uncer']:
            Initial_Year_of_Uncertainty = int( base_configuration_overall.loc[ n, 'Value' ] )
    '''
    ####################################################################################################################################################
    '''
    # This section reads a reference data.csv from baseline scenarios and frames Structure-OSEMOSYS_CR.xlsx
    col_position = []
    col_corresponding_initial = []
    for n in range( len( S_DICT_sets_structure['set'] ) ):
        col_position.append( header_row.index( S_DICT_sets_structure['set'][n] ) )
        col_corresponding_initial.append( S_DICT_sets_structure['initial'][n] )
    # Define the dictionary for calibrated database:
    stable_scenarios = {}
    for scen in scenario_list:
        stable_scenarios.update( { scen:{} } )
    #
    for scen in range( len( scenario_list ) ):
        this_paramter_list_dir = os.path.join(all_dataset_address, str( scenario_list[scen] ))
        parameter_list = os.listdir( this_paramter_list_dir )
        #
        for p in range( len( parameter_list ) ):
            this_param = parameter_list[p].replace('.csv','')
            stable_scenarios[ scenario_list[scen] ].update( { this_param:{} } )
            # To extract the parameter input data:
            all_params_list_index = S_DICT_params_structure['parameter'].index(this_param)
            this_number_of_elements = S_DICT_params_structure['number_of_elements'][all_params_list_index]
            this_index_list = S_DICT_params_structure['index_list'][all_params_list_index]
            #
            for k in range(this_number_of_elements):
                stable_scenarios[ scenario_list[scen] ][ this_param ].update({this_index_list[k]:[]})
            stable_scenarios[ scenario_list[scen] ][ this_param ].update({'value':[]})
            # Extract data:
            with open( this_paramter_list_dir + '/' + str(parameter_list[p]), mode='r') as csv_file:
                csv_reader = csv.DictReader(csv_file)
                #
                for row in csv_reader:
                    if row[ header_row[-1] ] != None and row[ header_row[-1] ] != '':
                        #
                        for h in range( 2, len(header_row)-1 ):
                            if row[ header_row[h] ] != None and row[ header_row[h] ] != '':
                                set_index  = S_DICT_sets_structure['set'].index( header_row[h] )
                                set_initial = S_DICT_sets_structure['initial'][ set_index ]
                                stable_scenarios[ scenario_list[scen] ][ this_param ][ set_initial ].append( row[ header_row[h] ] )
                        stable_scenarios[ scenario_list[scen] ][ this_param ][ 'value' ].append( row[ header_row[-1] ] )
                        #
    stable_scenarios_freeze = deepcopy(stable_scenarios)
    ####################################################################################################################################################
    # 1 - DISTANCE DICTIONARY:
    distance_levers = {}
    base_configuration_distance_unique_scenario_list = []
    for n in range( len( base_configuration_distance.index ) ):
        #
        if str( base_configuration_distance.loc[ n , params['scen']] ) not in base_configuration_distance_unique_scenario_list:
            base_configuration_distance_group_set_list = []
            base_configuration_distance_unique_scenario_list.append( str( base_configuration_distance.loc[ n , params['scen']] ) )
            distance_levers.update( { base_configuration_distance_unique_scenario_list[-1]:{} } )
        else:
            pass
        #
        if str( base_configuration_distance.loc[ n ,params['group_set']] ) not in base_configuration_distance_group_set_list:
            base_configuration_distance_group_set_list.append( str( base_configuration_distance.loc[ n ,params['group_set']] ) )
            distance_levers[ base_configuration_distance_unique_scenario_list[-1] ].update( { base_configuration_distance_group_set_list[-1]:0 } )
        else:
            pass
        #
        # Now we fill the data:
        distance_levers[ base_configuration_distance_unique_scenario_list[-1] ][ base_configuration_distance_group_set_list[-1] ] = float( base_configuration_distance.loc[ n ,'Relative reduction to BAU - distance'] )
        #
    #
    ####################################################################################################################################################
    # 2 - MODE SHIFT DICTIONARY:
    modeshift_params = {}
    base_configuration_modeshift_unique_scenario_list = []
    for n in range( len( base_configuration_modeshift.index ) ):
        #
        if str( base_configuration_modeshift.loc[ n ,params['scen']] ) not in base_configuration_modeshift_unique_scenario_list:
            base_configuration_modeshift_group_set_list = []
            base_configuration_modeshift_unique_scenario_list.append( str( base_configuration_modeshift.loc[ n ,params['scen']] ) )
            modeshift_params.update( { base_configuration_modeshift_unique_scenario_list[-1]:{} } )
        else:
            pass
        #
        base_configuration_modeshift_group_set_list.append( str( base_configuration_modeshift.loc[ n ,params['tech_set']] ) )
        modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ].update( { base_configuration_modeshift_group_set_list[-1]:{} } )
        #
        if str( base_configuration_modeshift.loc[ n ,params['logistic']] ) == 'YES' and str( base_configuration_modeshift.loc[ n ,params['linear']] ) == 'NO':
            modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ][ base_configuration_modeshift_group_set_list[-1] ].update( {'Values':[],'Type':params['logistic']} ) # WE MUST REMEMBER THE ORDER OF APPENDING THE PARAMETERS OF LOGISTIC BEHAVIOR: params['R_base_year'], params['R_final_year'], L, C, M
            modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ][ base_configuration_modeshift_group_set_list[-1] ].update( { params['context']:base_configuration_modeshift.loc[ n ,params['context']] } )
            #
            modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ][ base_configuration_modeshift_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_modeshift.loc[ n ,params['R_base_year']] ) )
            modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ][ base_configuration_modeshift_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_modeshift.loc[ n ,params['R_final_year']] ) )
            modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ][ base_configuration_modeshift_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_modeshift.loc[ n ,'L'] ) )
            modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ][ base_configuration_modeshift_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_modeshift.loc[ n ,'C'] ) )
            modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ][ base_configuration_modeshift_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_modeshift.loc[ n ,'M'] ) )
            #
        #
        elif str(base_configuration_modeshift.loc[n ,'Logistic']) == 'NO' and str(base_configuration_modeshift.loc[n ,'Linear']) == 'YES':
            modeshift_params[base_configuration_modeshift_unique_scenario_list[-1]][base_configuration_modeshift_group_set_list[-1]].update({'Values':[], 'Type':'Linear'}) # WE MUST REMEMBER THE ORDER OF APPENDING THE PARAMETERS OF LINEAR BEHAVIOR: y_ini, v_2030, v_2040, v_2050
            modeshift_params[base_configuration_modeshift_unique_scenario_list[-1]][base_configuration_modeshift_group_set_list[-1]].update({ 'Context':base_configuration_modeshift.loc[n ,'Context'] })
            #
            modeshift_params[base_configuration_modeshift_unique_scenario_list[-1]][base_configuration_modeshift_group_set_list[-1]][ 'Values' ].append(float(base_configuration_modeshift.loc[n ,'y_ini']))
            modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ][ base_configuration_modeshift_group_set_list[-1] ][ 'Values' ].append( base_configuration_modeshift.loc[ n ,params['v_first_decade_year']] )
            modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ][ base_configuration_modeshift_group_set_list[-1] ][ 'Values' ].append( base_configuration_modeshift.loc[ n ,params['v_sec_decade_year']] )
            modeshift_params[ base_configuration_modeshift_unique_scenario_list[-1] ][ base_configuration_modeshift_group_set_list[-1] ][ 'Values' ].append( base_configuration_modeshift.loc[ n ,params['v_final_year']] )
            #
        #
        elif str(base_configuration_modeshift.loc[n ,'Built-in']) == 'YES':
            modeshift_params[base_configuration_modeshift_unique_scenario_list[-1]][base_configuration_modeshift_group_set_list[-1]].update({'Values':'Check_Cap', 'Type':'Built-in'}) # WE MUST REMEMBER TO MAKE REFERENCE
            modeshift_params[base_configuration_modeshift_unique_scenario_list[-1]][base_configuration_modeshift_group_set_list[-1]].update({ 'Context':base_configuration_modeshift.loc[n ,'Context'] })
            #
        #
    #
    ####################################################################################################################################################
    # 3 - OCCUPANCY RATE DICTIONARY:
    or_params = {}
    base_configuration_or_unique_scenario_list = []
    for n in range( len( base_configuration_or.index ) ):
        #
        if str( base_configuration_or.loc[ n ,params['scen']] ) not in base_configuration_or_unique_scenario_list:
            base_configuration_or_group_set_list = []
            base_configuration_or_unique_scenario_list.append( str( base_configuration_or.loc[ n ,params['scen']] ) )
            or_params.update( { base_configuration_or_unique_scenario_list[-1]:{} } )
        else:
            pass
        #
        if str( base_configuration_or.loc[ n ,params['group_set']] ) not in base_configuration_or_group_set_list:
            base_configuration_or_group_set_list.append( str( base_configuration_or.loc[ n ,params['group_set']] ) )
            or_params[ base_configuration_or_unique_scenario_list[-1] ].update( { base_configuration_or_group_set_list[-1]:0 } )
        else:
            pass
        #
        # Now we fill the data:
        or_params[ base_configuration_or_unique_scenario_list[-1] ][ base_configuration_or_group_set_list[-1] ] = float( base_configuration_or.loc[ n ,'Relative increase to BAU - occupancy rate'] )
        #
    ####################################################################################################################################################
    # 4 - TECH ADOPTION DICTIONARY:
    adoption_params = {}
    base_configuration_adoption_unique_scenario_list = []
    for n in range( len( base_configuration_adoption.index ) ):
        #
        if str( base_configuration_adoption.loc[ n ,params['scen']] ) not in base_configuration_adoption_unique_scenario_list:
            base_configuration_adoption_group_set_list = []
            base_configuration_adoption_unique_scenario_list.append( str( base_configuration_adoption.loc[ n ,params['scen']] ) )
            adoption_params.update( { base_configuration_adoption_unique_scenario_list[-1]:{} } )
        else:
            pass
        #
        base_configuration_adoption_group_set_list.append( str( base_configuration_adoption.loc[ n ,params['tech_set']] ) )
        adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ].update( { base_configuration_adoption_group_set_list[-1]:{} } )
        store_the_sector = deepcopy(str( base_configuration_adoption.loc[ n ,'Sector'] ))
        #
        if str( base_configuration_adoption.loc[ n ,params['logistic']] ) == 'YES' and str( base_configuration_adoption.loc[ n ,params['linear']] ) == 'NO':
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ].update( {'Values':[], 'Type':params['logistic'], 'Restriction_Type':str( base_configuration_adoption.loc[ n ,'Restriction_Type'] ) } ) # WE MUST REMEMBER THE ORDER OF APPENDING THE PARAMETERS OF LOGISTIC BEHAVIOR: params['R_base_year'], params['R_final_year'], L, C, M
            #
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_adoption.loc[ n ,params['R_base_year']] ) )
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_adoption.loc[ n ,params['R_final_year']] ) )
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_adoption.loc[ n ,'L'] ) )
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_adoption.loc[ n ,'C'] ) )
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_adoption.loc[ n ,'M'] ) )
            #
        #
        elif str( base_configuration_adoption.loc[ n ,params['logistic']] ) == 'NO' and str( base_configuration_adoption.loc[ n ,params['linear']] ) == 'YES':
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ].update( {'Values':[], 'Type':params['linear'], 'Restriction_Type':str( base_configuration_adoption.loc[ n ,'Restriction_Type'] ) } ) # WE MUST REMEMBER THE ORDER OF APPENDING THE PARAMETERS OF LINEAR BEHAVIOR: y_ini, params['v_final_year']
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ][ 'Values' ].append( float( base_configuration_adoption.loc[ n ,'y_ini'] ) )
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ][ 'Values' ].append( base_configuration_adoption.loc[ n ,params['v_first_decade_year']] )
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ][ 'Values' ].append( base_configuration_adoption.loc[ n ,params['v_sec_decade_year']] )
            adoption_params[ base_configuration_adoption_unique_scenario_list[-1] ][ base_configuration_adoption_group_set_list[-1] ][ 'Values' ].append( base_configuration_adoption.loc[ n ,params['v_final_year']] )
            #
        #
        adoption_params[base_configuration_adoption_unique_scenario_list[-1]][base_configuration_adoption_group_set_list[-1]].update({'Sector':store_the_sector})
   #
    '''
    ##########################################################################
    '''
    transport_group_sets = list( Fleet_Groups.keys() )
    transport_group_sets = [ i for i in transport_group_sets if params['train'] not in i and params['group_tech'] not in i ]
    Reference_driven_distance = {}
    Reference_occupancy_rate = {}
    Reference_op_life = {}
    #
    for s in range( len( scenario_list ) ):
        Reference_driven_distance.update( { scenario_list[s]:{} } )
        Reference_occupancy_rate.update( { scenario_list[s]:{} } )
        Reference_op_life.update( { scenario_list[s]:{} } )
        for n in range( len( transport_group_sets ) ):
            this_set = transport_group_sets[n]
            #
            if scenario_list[s] in base_configuration_distance_unique_scenario_list or scenario_list[s] == ref_scenario:
                base_distance = Fleet_Groups_Distance[ Fleet_Groups[this_set][0] ] # the [0] helps call the detail fo distance per vehicle
                if scenario_list[s] != ref_scenario:
                    new_distance = deepcopy( interpolation_non_linear_final_flexible( time_range_vector, base_distance, distance_levers[scenario_list[s]][this_set], Initial_Year_of_Uncertainty ) )
                    new_distance_rounded = [ round(elem, params['round_#']) for elem in new_distance ]
                    Reference_driven_distance[ scenario_list[s] ].update( {this_set:new_distance_rounded} )
                else:
                    Reference_driven_distance[ scenario_list[s] ].update( {this_set:base_distance} )
            #
            if scenario_list[s] in base_configuration_or_unique_scenario_list or scenario_list[s] == ref_scenario:
                these_or_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ][ 't' ] ) if x == str( this_set ) ]
                base_or = deepcopy( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ]['value'][ these_or_indices[0]:these_or_indices[-1]+1 ] )
                base_or = [ float( base_or[j] ) for j in range( len( base_or ) ) ]
                if scenario_list[s] != ref_scenario:
                    new_or = deepcopy( interpolation_non_linear_final_flexible( time_range_vector, base_or, or_params[scenario_list[s]][this_set], Initial_Year_of_Uncertainty ) )
                    new_or_rounded = [ round(elem, params['round_#']) for elem in new_or ]
                    Reference_occupancy_rate[ scenario_list[s] ].update( {this_set:new_or_rounded} )
                else:
                    Reference_occupancy_rate[ scenario_list[s] ].update( {this_set:base_or} )
            #
        #
        for n in range(len(list(Fleet_Groups_inv.keys()))):
            this_set = list(Fleet_Groups_inv.keys())[n]
            if scenario_list[s] in base_configuration_distance_unique_scenario_list or scenario_list[s] == ref_scenario:  # includes distance scenarios
                these_ol_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'OperationalLife' ][ 't' ] ) if x == str( this_set ) ]
                base_ol = deepcopy( stable_scenarios[ scenario_list[s] ][ 'OperationalLife' ]['value'][ these_ol_indices[0]:these_ol_indices[-1]+1 ] )
                base_ol = [ float( base_ol[j] ) for j in range( len( base_ol ) ) ]
                Reference_op_life[ scenario_list[s] ].update( {this_set:base_ol[0]} )
    #
    '''''
    ################################# PART 2 #################################
    '''''
    print('2: We will now manipulate each scenario for later printing of the .csv files in the same adress.')
    '''
    NOTES:
    i) The manipulation of the scenarios is done by taking the demand as reference and the orderly applying the modifications to the transport elements.
    ii) We first perform changes to the _Base_Dataset file to fix the errors found in the system.
    iii) We execute the changes in the scenario restrictions.
    '''
    if params['Use_Blend_Shares_B1']:
        Blend_Shares = {}
        for s in range( len( scenario_list ) ):
            Blend_Shares.update({scenario_list[s]:{}})
        #
    #
    for s in range( len( scenario_list ) ):
        #
        this_set_type_initial = S_DICT_sets_structure['initial'][ S_DICT_sets_structure['set'].index('TECHNOLOGY') ]
        capacity_variables = params['capacity_variables']
        #
        if scenario_list[s] == params['BAU']:
            relative_pkm_to_demand = { 'TotalAnnualMaxCapacity':{}, 'TotalTechnologyAnnualActivityLowerLimit':{} }
            relative_pkm_to_demand_nonrail = { 'TotalAnnualMaxCapacity':{}, 'TotalTechnologyAnnualActivityLowerLimit':{} }
            relative_tkm_to_demand = { 'TotalAnnualMaxCapacity':{}, 'TotalTechnologyAnnualActivityLowerLimit':{} }
            relative_tkm_to_demand_nonrail  = { 'TotalAnnualMaxCapacity':{}, 'TotalTechnologyAnnualActivityLowerLimit':{} }
        #
        ### Call the demand and go from there for restriciton definition ###
        this_set_type_initial = S_DICT_sets_structure['initial'][ S_DICT_sets_structure['set'].index('FUEL') ]
        passpub_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ][ this_set_type_initial ] ) if x == str( params['tra_dem_pub'] ) ]
        passpriv_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ][ this_set_type_initial ] ) if x == str( params['tra_dem_pri'] ) ]
        tourism_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ][ this_set_type_initial ] ) if x == str( params['tra_dem_tur'] ) ]
        nonmotorized_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ][ this_set_type_initial ] ) if x == str( params['tra_non_mot'] ) ]

        hfre_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ][ this_set_type_initial ] ) if x == str( params['tra_dem_hea'] ) ]
        lfre_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ][ this_set_type_initial ] ) if x == str( params['tra_dem_lig'] ) ]
        #
        # We will need to store the BAU data for later use.
        if scenario_list[s] == params['BAU']:
            #
            passpub_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ passpub_range_indices[0]:passpub_range_indices[-1]+1 ] )
            passpriv_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ passpriv_range_indices[0]:passpriv_range_indices[-1]+1 ] )
            nonmotorized_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ nonmotorized_range_indices[0]:nonmotorized_range_indices[-1]+1 ] )
            tourism_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ tourism_range_indices[0]:tourism_range_indices[-1]+1 ] )
            #
            ref_fre_hea = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ hfre_range_indices[0]:hfre_range_indices[-1]+1 ] )
            ref_fre_lig = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ lfre_range_indices[0]:lfre_range_indices[-1]+1 ] )
            #
            Total_Demand = []
            Total_Demand_Fre = []
            for n in range( len( time_range_vector ) ):
                Total_Demand.append(float( passpub_values[n]) + float(passpriv_values[n]) + float(nonmotorized_values[n])  + float(tourism_values[n]))
                Total_Demand_Fre.append( float( ref_fre_hea[n] ) + float( ref_fre_lig[n] ) )
            #
            ref_pass_pub_shares     = [ float(passpub_values[n])/Total_Demand[n] for n in range( len( time_range_vector ) ) ]
            ref_pass_priv_shares    = [ float(passpriv_values[n])/Total_Demand[n] for n in range( len( time_range_vector ) ) ]
            ref_nonmotorized_shares = [ float(nonmotorized_values[n])/Total_Demand[n] for n in range( len( time_range_vector ) ) ]
        #
        else:
            ###################################################################################################
            ### BLOCK 3: modify mode_shift ###
            applicable_sets_demand = []
            applicable_sets_tech = []
            for a_set in range( len( list( modeshift_params[ scenario_list[s] ].keys() ) ) ):
                this_set = list( modeshift_params[ scenario_list[s] ].keys() )[ a_set ]
                if modeshift_params[ scenario_list[s] ][ this_set ][ params['context'] ] == 'Demand':
                    applicable_sets_demand.append( this_set )
                elif modeshift_params[ scenario_list[s] ][ this_set ][ params['context'] ] == 'Technology':
                    applicable_sets_tech.append( this_set )

            # Now we perform the adjustment of the system, after having modified the demand:
            for a_set in range( len( applicable_sets_demand ) ): # FLAG: in the future we must consider non-logistic behaviors
                #
                this_set = applicable_sets_demand[ a_set ]
                if modeshift_params[ scenario_list[s] ][ this_set ][ params['type'] ] == params['logistic']:
                    #
                    R_base_year = modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 0 ] # This is not used / remove in future versions.
                    R_final_year = modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 1 ]
                    L =     modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 2 ]
                    C =     modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 3 ]
                    M =     modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 4 ]
                    #
                    r_final_year = 1/R_final_year
                    Q =     L/C - 1
                    k =     np.log( (r_final_year-1)/Q )/(M-params['final_year'])
                    #
                    shift_years = [ n for n in range( Initial_Year_of_Uncertainty+1,params['final_year']+1 ) ]
                    shift_year_counter = 0
                    adoption_shift = []
                    #
                    for t in range( len( time_range_vector ) ):
                        if time_range_vector[t] > Initial_Year_of_Uncertainty:
                            x = int( shift_years[shift_year_counter] )
                            adoption_shift.append( generalized_logistic_curve(x, L, Q, k, M))
                            shift_year_counter += 1
                        else:
                            adoption_shift.append( 0.0 )
                        #
                    #
                #
                elif modeshift_params[ scenario_list[s] ][ this_set ][ params['type'] ] == params['linear']:
                    y_ini = modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 0 ]
                    v_first_decade_year = modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 1 ]
                    v_sec_decade_year = modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 2 ]                
                    v_final_year = modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 3 ]
                    #
                    x_coord_tofill, xp_coord_known, yp_coord_known = [], [], []
                    for y in range( len( time_range_vector ) ):
                        not_known_e = True
                        if time_range_vector[y] <= y_ini:
                            xp_coord_known.append( float(y) )
                            yp_coord_known.append( float(0) )
                            not_known_e = False
                        if v_first_decade_year != params['interp'] and time_range_vector[y] == params['first_decade_year']:
                            xp_coord_known.append( float(y) )
                            yp_coord_known.append( float(v_first_decade_year) )
                            not_known_e = False
                        if v_sec_decade_year != params['interp'] and time_range_vector[y] == params['sec_decade_year']:
                            xp_coord_known.append( float(y) )
                            yp_coord_known.append( float(v_sec_decade_year) )
                            not_known_e = False
                        if v_final_year != params['interp'] and time_range_vector[y] == params['final_year']:
                            xp_coord_known.append( float(y) )
                            yp_coord_known.append( float(v_final_year) )
                            not_known_e = False
                        if not_known_e == True:
                            x_coord_tofill.append( float(y) )
                    #
                    y_coord_filled = list( np.interp( x_coord_tofill, xp_coord_known, yp_coord_known ) )
                    interpolated_values = []
                    for coord in range( len( time_range_vector ) ):
                        if coord in xp_coord_known:
                            value_index = xp_coord_known.index(coord)
                            interpolated_values.append( float( yp_coord_known[value_index] ) )
                        elif coord in x_coord_tofill:
                            value_index = x_coord_tofill.index(coord)
                            interpolated_values.append( float( y_coord_filled[value_index] ) )
                    #
                    adoption_shift = interpolated_values
                    #
                #
                if params['tra_dem_pub'] in this_set:
                    new_value_list_PUB = []
                elif params['tra_non_mot'] in this_set:
                    new_value_list_NOMOT = []
                #
                for n in range( len( time_range_vector ) ):               
                    if params['tra_dem_pub'] in this_set: # we must subtract the rail activity and then add it again above
                        new_value_list_PUB.append( ( ref_pass_pub_shares[n] + adoption_shift[n] )*( Total_Demand[n] ) ) # - ref_value_range_RAILACTIVITY[n] )
                        new_value_list_PUB_rounded = [ round(elem, params['round_#']) for elem in new_value_list_PUB ]
                    elif params['tra_non_mot'] in this_set:
                        new_value_list_NOMOT.append( ( adoption_shift[n] )*( Total_Demand[n] ) )
                        new_value_list_NOMOT_rounded = [ round(elem, params['round_#']) for elem in new_value_list_NOMOT ]
            #
            adjust_with_rail = False
            for a_set in range( len( applicable_sets_tech ) ): # REMEMBER THIS IS A CHANGE FUNCTION
                this_set = applicable_sets_tech[ a_set ]
                if params['techs_var'] not in this_set:
                    this_set_group = Fleet_Groups_inv[ this_set ]
                else:
                    this_set_group = this_set
                adjust_with_rail = True
                #
                # We need to call the group capacity
                this_set_group_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) if x == str( this_set_group ) ]
                this_set_group_range_indices_lm = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalTechnologyAnnualActivityLowerLimit' ][ 't' ] ) if x == str( this_set_group ) ]
                value_range_set_group = deepcopy( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ]['value'][ this_set_group_range_indices[0]:this_set_group_range_indices[-1]+1 ] )
                value_range_set_group = [ float(value_range_set_group[j]) for j in range( len( time_range_vector ) ) ]
                #
                # We also need the demand correspoding to this set:
                this_demand_set = Fleet_Groups_techs_2_dem[ this_set_group ]
                this_demand_set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'SpecifiedAnnualDemand' ][ 'f' ] ) if x == str( this_demand_set ) ]
                this_demand_set_value = deepcopy( stable_scenarios[ scenario_list[ s ] ][ 'SpecifiedAnnualDemand' ]['value'][ this_demand_set_range_indices[0]:this_demand_set_range_indices[-1]+1 ] )
                this_demand_set_value = [ float(this_demand_set_value[j]) for j in range( len( time_range_vector ) ) ]
                #
                if modeshift_params[ scenario_list[s] ][ this_set ][ params['type'] ] == params['linear']:
                    y_ini = modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 0 ]
                    v_first_decade_year = modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 1 ]
                    v_sec_decade_year = modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 2 ]                
                    v_final_year = modeshift_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 3 ]
                    #
                    x_coord_tofill, xp_coord_known, yp_coord_known = [], [], []
                    for y in range( len( time_range_vector ) ):
                        not_known_e = True
                        if time_range_vector[y] <= y_ini:
                            xp_coord_known.append( y )
                            yp_coord_known.append( 0 )
                            not_known_e = False
                        if v_first_decade_year != params['interp'] and time_range_vector[y] == params['first_decade_year']:
                            xp_coord_known.append( y )
                            yp_coord_known.append( v_first_decade_year )
                            not_known_e = False
                        if v_sec_decade_year != params['interp'] and time_range_vector[y] == params['sec_decade_year']:
                            xp_coord_known.append( y )
                            yp_coord_known.append( v_sec_decade_year )
                            not_known_e = False
                        if v_final_year != params['interp'] and time_range_vector[y] == params['final_year']:
                            xp_coord_known.append( y )
                            yp_coord_known.append( v_final_year )
                            not_known_e = False
                        if not_known_e == True:
                            x_coord_tofill.append( y )
                    #
                    y_coord_filled = list( np.interp( x_coord_tofill, xp_coord_known, yp_coord_known ) )
                    interpolated_values = []
                    for coord in range( len( time_range_vector ) ):
                        if coord in xp_coord_known:
                            value_index = xp_coord_known.index(coord)
                            interpolated_values.append( float( yp_coord_known[value_index] ) )
                        elif coord in x_coord_tofill:
                            value_index = x_coord_tofill.index(coord)
                            interpolated_values.append( float( y_coord_filled[value_index] ) )
                    #
                    new_value_range_set_group = [ float( value_range_set_group[j] ) + interpolated_values[j]*this_demand_set_value[j] for j in range( len( time_range_vector ) ) ]
                    new_value_range_set_group_rounded = [ round(elem, params['round_#']) for elem in new_value_range_set_group ]
                    #
                    # Store the value of train rail (freight) below
                    BOOL_ADJUST_COMPLEMENTS = True
                    if BOOL_ADJUST_COMPLEMENTS:
                        stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ]['value'][ this_set_group_range_indices[0]:this_set_group_range_indices[-1]+1 ] = deepcopy( new_value_range_set_group_rounded )
                        stable_scenarios[ scenario_list[ s ] ][ 'TotalTechnologyAnnualActivityLowerLimit' ]['value'][ this_set_group_range_indices_lm[0]:this_set_group_range_indices_lm[-1]+1 ] = deepcopy( new_value_range_set_group_rounded )

                        for this_parameter in [ 'TotalAnnualMaxCapacity', 'TotalTechnologyAnnualActivityLowerLimit' ]:
                            # We must adjust the capacity of the complementary sets:
                            if this_set == params['tech_train_fre']:
                                this_set_complement_dict = {
                                    params['tech_train_fre']:[params['tech_he_fre']]}
                                this_set_complement_vals_rel_100 = {
                                    params['tech_he_fre']:100}
                            if this_set in [params['tech_telef'], params['tech_train']]:
                                if this_set == params['tech_telef']:
                                    this_set_complement_dict = {
                                        params['tech_telef']:params['techs_trans_pass']}
                                if this_set == params['tech_train']:
                                    this_set_complement_dict = {
                                        params['tech_train']:params['techs_trans_pass']}
                                this_set_complement_vals_rel_100 = params['this_set_complement_vals_rel_100']

                            for acomp in range(len(this_set_complement_dict[this_set])):
                                this_set_complement = this_set_complement_dict[this_set][acomp]
                                this_set_complement_indices = [ i for i, x in enumerate( stable_scenarios[scenario_list[s]][ this_parameter ][ 't' ] ) if x == str( this_set_complement ) ]
                                value_list_complement = deepcopy( stable_scenarios[scenario_list[s]][ this_parameter ]['value'][ this_set_complement_indices[0]:this_set_complement_indices[-1]+1 ] )
                                value_list_complement = [ float(value_list_complement[j]) for j in range( len( time_range_vector ) ) ]

                                this_set_complement_adj_val = this_set_complement_vals_rel_100[this_set_complement]/100

                                # Extract OR parameters:
                                or_group_tech_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'OutputActivityRatio' ][ 't' ] ) if x == str( this_set_complement ) ]
                                or_group_tech_values = deepcopy( stable_scenarios[ scenario_list[ s ] ][ 'OutputActivityRatio' ]['value'][ or_group_tech_indices[0]:or_group_tech_indices[-1]+1 ] )
                                or_group_tech_values = [ float( or_group_tech_values[j] ) for j in range( len( or_group_tech_values ) ) ]

                                # Define new heavy truck capacity
                                if this_parameter == 'TotalAnnualMaxCapacity':
                                    security_multiplier_factor = 1.01
                                elif this_parameter == 'TotalTechnologyAnnualActivityLowerLimit':
                                    security_multiplier_factor = 0.999

                                new_value_list_comp = []
                                for y in range(len(time_range_vector)):
                                    if time_range_vector[y] < Initial_Year_of_Uncertainty:
                                        new_value_list_comp.append(value_list_complement[y])
                                    else:
                                        new_value_list_comp.append(
                                            security_multiplier_factor*this_set_complement_adj_val*(this_demand_set_value[y] - new_value_range_set_group_rounded[y])/or_group_tech_values[y])
                                # Store the value of trucks (after change) below
                                new_value_list_comp_rounded = [ round(elem, 4) for elem in new_value_list_comp ]
                                stable_scenarios[scenario_list[s]][ this_parameter ]['value'][ this_set_complement_indices[0]:this_set_complement_indices[-1]+1 ] = deepcopy( new_value_list_comp_rounded )

                #
                '''
                if modeshift_params[ scenario_list[s] ][ this_set ][ params['type'] ] == params['built_in'] and this_set == params['tech_train_ele']:  # *ALERT* add the elimination of rail as a possible tool
                    this_set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) if x == str( this_set ) ]
                    value_range_set = deepcopy( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ]['value'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ] )
                    value_range_set = [ float(value_range_set[j]) for j in range( len( time_range_vector ) ) ]
                    #
                    new_value_range_set_group = [ float( value_range_set_group[i] ) + float( value_range_set[i] ) for i in range( len( value_range_set ) ) ]
                    new_value_range_set_group_rounded = [ round(elem, params['round_#']) for elem in new_value_range_set_group ]
                    #
                    stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ]['value'][ this_set_group_range_indices[0]:this_set_group_range_indices[-1]+1 ] = deepcopy( new_value_range_set_group_rounded )
                    stable_scenarios[ scenario_list[ s ] ][ 'TotalTechnologyAnnualActivityLowerLimit' ]['value'][ this_set_group_range_indices_lm[0]:this_set_group_range_indices_lm[-1]+1 ] = deepcopy( new_value_range_set_group_rounded )
                #
                if modeshift_params[ scenario_list[s] ][ this_set ][ params['type'] ] == params['linear'] and this_set == params['tech_train_ele']:  # *ALERT* add the elimination of rail as a possible tool
                    #
                    value_range_set_zero = [ 0.0 for j in range( len( time_range_vector ) ) ]
                    #
                    # Set zero for NDP (total capacity)
                    this_set_range_indices_max = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) if x == str( this_set ) ]
                    this_set_range_indices_lower = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalTechnologyAnnualActivityLowerLimit' ][ 't' ] ) if x == str( this_set ) ]
                    #
                    stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ]['value'][ this_set_range_indices_max[0]:this_set_range_indices_max[-1]+1 ] = deepcopy( value_range_set_zero )
                    stable_scenarios[ scenario_list[ s ] ][ 'TotalTechnologyAnnualActivityLowerLimit' ]['value'][ this_set_range_indices_lower[0]:this_set_range_indices_lower[-1]+1 ] = deepcopy( value_range_set_zero )
                    #
                    # Set zero for NDP (costs for passenger rail)
                    cost_sets_eliminate = params['cost_sets_eliminate']
                    for this_cost in cost_sets_eliminate:
                        this_set_range_indices_cost = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'CapitalCost' ][ 't' ] ) if x == str( this_cost ) ]
                        stable_scenarios[ scenario_list[ s ] ][ 'CapitalCost' ]['value'][ this_set_range_indices_cost[0]:this_set_range_indices_cost[-1]+1 ] = deepcopy( value_range_set_zero )
                    #
                    cost_sets_eliminate = ['TRANRAILINF']
                    for this_cost in cost_sets_eliminate:
                        this_set_range_indices_cost = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'FixedCost' ][ 't' ] ) if x == str( this_cost ) ]
                        stable_scenarios[ scenario_list[ s ] ][ 'FixedCost' ]['value'][ this_set_range_indices_cost[0]:this_set_range_indices_cost[-1]+1 ] = deepcopy( value_range_set_zero )
                    #
                    # Re-establish BAU for NDP (group capacity)
                    this_set_range_indices_bau = [ i for i, x in enumerate( stable_scenarios[ params['BAU'] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) if x == str( this_set_group ) ]
                    value_range_set = deepcopy( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ]['value'][ this_set_range_indices_bau[0]:this_set_range_indices_bau[-1]+1 ] )
                    value_range_set = [ float(value_range_set[j]) for j in range( len( time_range_vector ) ) ]
                    stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ]['value'][ this_set_group_range_indices[0]:this_set_group_range_indices[-1]+1 ] = deepcopy( value_range_set )
                    stable_scenarios[ scenario_list[ s ] ][ 'TotalTechnologyAnnualActivityLowerLimit' ]['value'][ this_set_group_range_indices_lm[0]:this_set_group_range_indices_lm[-1]+1 ] = deepcopy( value_range_set )
                '''
            #
            tourism_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ][ 'f' ] ) if x == str( params['tra_dem_tur'] ) ]
            tourism_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ tourism_range_indices[0]:tourism_range_indices[-1]+1 ] )                                                                                                                                                         
            new_value_list_PRIV = [ Total_Demand[n] - new_value_list_PUB_rounded[n] - new_value_list_NOMOT_rounded[n] - float(tourism_values[n]) for n in range( len( time_range_vector ) ) ]
            new_value_list_PRIV_rounded = [ round(elem, params['round_#']) for elem in new_value_list_PRIV ]
            #
            # Assign parameters back: for these subset of uncertainties
            stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ passpub_range_indices[0]:passpub_range_indices[-1]+1 ] = deepcopy( new_value_list_PUB_rounded )
            stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ passpriv_range_indices[0]:passpriv_range_indices[-1]+1 ] = deepcopy( new_value_list_PRIV_rounded )
            stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ nonmotorized_range_indices[0]:nonmotorized_range_indices[-1]+1 ] = deepcopy( new_value_list_NOMOT_rounded )
            #
        #
        ### BLOCK 4: call and modify the occupancy rate ###
        this_set_type_initial = S_DICT_sets_structure['initial'][ S_DICT_sets_structure['set'].index('TECHNOLOGY') ]
        if scenario_list[s] == params['BAU']:
            ref_or_values = {}
        #
        for g in range( len( transport_group_sets ) ):
            or_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ][ this_set_type_initial ] ) if x == str( transport_group_sets[g] ) ]
            #
            if scenario_list[s] == params['BAU']:
                or_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ]['value'][ or_indices[0]:or_indices[-1]+1 ] )
                ref_or_values.update( { transport_group_sets[g]:or_values } )
            #
            else:
                #
                new_value_list = deepcopy( interpolation_non_linear_final_flexible( time_range_vector, ref_or_values[ transport_group_sets[g] ] , or_params[ scenario_list[s] ][ transport_group_sets[g] ], Initial_Year_of_Uncertainty ) )
                new_value_list_rounded = [ round(elem, params['round_#']) for elem in new_value_list ]
                stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ]['value'][ or_indices[0]:or_indices[-1]+1 ] = deepcopy( new_value_list_rounded )
        #
        ### BLOCK 5: call and re-establish the capacity variables of group techs after modal shift change AND occupancy rate change ###
        this_set_type_initial = S_DICT_sets_structure['initial'][ S_DICT_sets_structure['set'].index('TECHNOLOGY') ]
        capacity_variables = params['capacity_variables']
        #
        # We must call the group techs and calculate the passenger kilometers provided by each tech, then store the share relative to the demand
        for capvar in range( len( capacity_variables ) ):
            #
            for g in range( len( group_tech_PRIVATE ) ):
                #
                cap_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ][ this_set_type_initial ] ) if x == str( group_tech_PRIVATE[g] ) ]
                or_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ][ this_set_type_initial ] ) if x == str( group_tech_PRIVATE[g] ) ]
                demand_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ][ 'f' ] ) if x == str( Fleet_Groups_techs_2_dem[ group_tech_PRIVATE[g] ] ) ]
                #
                # here we quickly call the occupancy rates
                or_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ]['value'][ or_indices[0]:or_indices[-1]+1 ] )
                or_values = [ float( or_values[j] ) for j in range( len( or_values ) ) ]
                #
                # here we quickly call the associated demand
                demand_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ demand_indices[0]:demand_indices[-1]+1 ] )
                demand_values = [ float( demand_values[j] ) for j in range( len( demand_values ) ) ]
                #
                cap_values = deepcopy( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices[0]:cap_indices[-1]+1 ] )
                cap_values = [ float( cap_values[j] ) for j in range( len( cap_values ) ) ]
                #
                if scenario_list[s] == params['BAU']: # REMEMBER THE VARIABLES OF BAU DO NOT CHANGE
                    #
                    this_pkm_share = [ cap_values[n]*or_values[n]/demand_values[n] for n in range( len( time_range_vector ) ) ]
                    this_pkm_share_rounded = [ round(elem, params['round_#']) for elem in this_pkm_share ]
                    #
                    relative_pkm_to_demand[ capacity_variables[capvar] ].update( { group_tech_PRIVATE[g]:this_pkm_share_rounded } )
                #
                else: # if scenario_list[s] in [params['NDP'],'NDP_A','OP15C']: # we must apply the same distribution of transport in pkm as in the BAU, considering the new demands. The variable to modify is the capacity.
                    #
                    apply_these_shares = relative_pkm_to_demand[ capacity_variables[capvar] ][ group_tech_PRIVATE[g] ]
                    new_cap_values = [ apply_these_shares[n]*demand_values[n]/or_values[n] for n in range( len( time_range_vector ) ) ]
                    new_cap_values_rounded = [ round(elem, params['round_#']) for elem in new_cap_values ]
                    #
                    stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices[0]:cap_indices[-1]+1 ] = deepcopy( new_cap_values_rounded )
                    #
                #
            #----------------------------------------------------------------------------------------------------------------------------------------------------------
            for g in range( len( group_tech_PUBLIC ) ):
                #
                cap_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ][ this_set_type_initial ] ) if x == str( group_tech_PUBLIC[g] ) ]
                or_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ][ this_set_type_initial ] ) if x == str( group_tech_PUBLIC[g] ) ]
                demand_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ][ 'f' ] ) if x == str( Fleet_Groups_techs_2_dem[ group_tech_PUBLIC[g] ] ) ]
                #
                cap_values = deepcopy( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices[0]:cap_indices[-1]+1 ] )
                cap_values = [ float( cap_values[j] ) for j in range( len( cap_values ) ) ]
                #
                if scenario_list[s] == params['BAU']:
                    #
                    or_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ]['value'][ or_indices[0]:or_indices[-1]+1 ] )
                    or_values = [ float( or_values[j] ) for j in range( len( or_values ) ) ]
                    #
                    # here we quickly call the associated demand
                    demand_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ demand_indices[0]:demand_indices[-1]+1 ] )
                    demand_values = [ float( demand_values[j] ) for j in range( len( demand_values ) ) ]
                    #
                    this_pkm_share = [ cap_values[n]*or_values[n]/demand_values[n] for n in range( len( time_range_vector ) ) ]
                    this_pkm_share_rounded = [ round(elem, params['round_#']) for elem in this_pkm_share ]
                    #
                    relative_pkm_to_demand[ capacity_variables[capvar] ].update( { group_tech_PUBLIC[g]:deepcopy( this_pkm_share_rounded ) } )
                    #
                    if params['train'] not in group_tech_PUBLIC[g]: # we need to find the shares of non-rail modes:
                        try:
                            cap_indices_rail = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ][ this_set_type_initial ] ) if x == str( params['tech_train'] ) ]
                            cap_values_rail = deepcopy( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices_rail[0]:cap_indices_rail[-1]+1 ] )
                        except Exception:
                            cap_values_rail = [0 for i in range(len(time_range_vector))]
                        cap_values_rail = [ float( cap_values_rail[j] ) for j in range( len( cap_values_rail ) ) ]
                        demand_values_norail = [ demand_values[j] - cap_values_rail[j] for j in range( len( demand_values ) ) ]
                        #
                        this_pkm_share = [ cap_values[n]*or_values[n]/demand_values_norail[n] for n in range( len( time_range_vector ) ) ]
                        this_pkm_share_rounded = [ round(elem, params['round_#']) for elem in this_pkm_share ]
                        #
                        relative_pkm_to_demand_nonrail[ capacity_variables[capvar] ].update( { group_tech_PUBLIC[g]:this_pkm_share_rounded } )
                    #
                #
                else: # we must apply the same distribution of transport in pkm as in the BAU, considering the new demands. The variable to modify is the capacity.
                    #
                    if adjust_with_rail == False:
                        or_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ]['value'][ or_indices[0]:or_indices[-1]+1 ] )
                        or_values = [ float( or_values[j] ) for j in range( len( or_values ) ) ]
                        #
                        # here we quickly call the associated demand
                        demand_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ demand_indices[0]:demand_indices[-1]+1 ] )
                        demand_values = [ float( demand_values[j] ) for j in range( len( demand_values ) ) ]
                        #
                        apply_these_shares = relative_pkm_to_demand[ capacity_variables[capvar] ][ group_tech_PUBLIC[g] ]
                        new_cap_values = [ apply_these_shares[n]*demand_values[n]/or_values[n] for n in range( len( time_range_vector ) ) ]
                        new_cap_values_rounded = [ round(elem, params['round_#']) for elem in new_cap_values ]
                        #
                        stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices[0]:cap_indices[-1]+1 ] = deepcopy( new_cap_values_rounded )
                    #
                    elif params['train'] not in group_tech_PUBLIC[g]:
                        or_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ]['value'][ or_indices[0]:or_indices[-1]+1 ] )
                        or_values = [ float( or_values[j] ) for j in range( len( or_values ) ) ]
                        #
                        # here we quickly call the associated demand
                        cap_indices_rail = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ][ this_set_type_initial ] ) if x == str( params['tech_train'] ) ]
                        cap_values_rail = deepcopy( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices_rail[0]:cap_indices_rail[-1]+1 ] )
                        cap_values_rail = [ float( cap_values_rail[j] ) for j in range( len( cap_values_rail ) ) ]
                        demand_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ demand_indices[0]:demand_indices[-1]+1 ] )
                        demand_values_norail = [ demand_values[j] - cap_values_rail[j] for j in range( len( demand_values ) ) ]
                        #
                        apply_these_shares = relative_pkm_to_demand_nonrail[ capacity_variables[capvar] ][ group_tech_PUBLIC[g] ]
                        new_cap_values = [ apply_these_shares[n]*demand_values_norail[n]/or_values[n] for n in range( len( time_range_vector ) ) ]
                        new_cap_values_rounded = [ round(elem, params['round_#']) for elem in new_cap_values ]
                        #
                        stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices[0]:cap_indices[-1]+1 ] = deepcopy( new_cap_values_rounded )
                    #
                #
            #----------------------------------------------------------------------------------------------------------------------------------------------------------
            for g in range( len( group_tech_FREIGHT_HEA ) ):
                #
                cap_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ][ this_set_type_initial ] ) if x == str( group_tech_FREIGHT_HEA[g] ) ]
                or_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ][ this_set_type_initial ] ) if x == str( group_tech_FREIGHT_HEA[g] ) ]
                demand_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ][ 'f' ] ) if x == str( Fleet_Groups_techs_2_dem[ group_tech_FREIGHT_HEA[g] ] ) ]
                # here we quickly call the associated demand
                demand_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ demand_indices[0]:demand_indices[-1]+1 ] )
                demand_values = [ float( demand_values[j] ) for j in range( len( demand_values ) ) ]
                #
                cap_values = deepcopy( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices[0]:cap_indices[-1]+1 ] )
                cap_values = [ float( cap_values[j] ) for j in range( len( cap_values ) ) ]
                #
                if scenario_list[s] == params['BAU']:
                    #
                    or_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ]['value'][ or_indices[0]:or_indices[-1]+1 ] )
                    or_values = [ float( or_values[j] ) for j in range( len( or_values ) ) ]
                    #
                    this_tkm_share = [ cap_values[n]*or_values[n]/demand_values[n] for n in range( len( time_range_vector ) ) ]
                    this_tkm_share_rounded = [ round(elem, params['round_#']) for elem in this_tkm_share ]
                    #
                    relative_tkm_to_demand[ capacity_variables[capvar] ].update( { group_tech_FREIGHT_HEA[g]:deepcopy( this_tkm_share_rounded ) } )
                    #
                    if params['train'] not in group_tech_FREIGHT_HEA[g]: # we need to find the shares of non-rail modes:
                        cap_indices_rail = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ][ this_set_type_initial ] ) if x == str( params['tech_train_fre'] ) ]
                        cap_values_rail = deepcopy( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices_rail[0]:cap_indices_rail[-1]+1 ] )
                        cap_values_rail = [ float( cap_values_rail[j] ) for j in range( len( cap_values_rail ) ) ]
                        demand_values_norail = [ demand_values[j] - cap_values_rail[j] for j in range( len( demand_values ) ) ]
                        #
                        this_tkm_share = [ cap_values[n]*or_values[n]/demand_values_norail[n] for n in range( len( time_range_vector ) ) ]
                        this_tkm_share_rounded = [ round(elem, params['round_#']) for elem in this_tkm_share ]
                        #
                        relative_tkm_to_demand_nonrail[ capacity_variables[capvar] ].update( { group_tech_FREIGHT_HEA[g]:this_tkm_share_rounded } )
                    #
                #
                else: # we must apply the same distribution of transport in pkm as in the BAU, considering the new demands. The variable to modify is the capacity.
                    #
                    if adjust_with_rail == False:
                        or_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ]['value'][ or_indices[0]:or_indices[-1]+1 ] )
                        or_values = [ float( or_values[j] ) for j in range( len( or_values ) ) ]
                        #
                        apply_these_shares = relative_tkm_to_demand[ capacity_variables[capvar] ][ group_tech_FREIGHT_HEA[g] ]
                        new_cap_values = [ apply_these_shares[n]*demand_values[n]/or_values[n] for n in range( len( time_range_vector ) ) ]
                        new_cap_values_rounded = [ round(elem, params['round_#']) for elem in new_cap_values ]
                        #
                        stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices[0]:cap_indices[-1]+1 ] = deepcopy( new_cap_values_rounded )
                    #
                    elif params['train'] not in group_tech_FREIGHT_HEA[g]:
                        or_values = deepcopy( stable_scenarios[ scenario_list[s] ][ 'OutputActivityRatio' ]['value'][ or_indices[0]:or_indices[-1]+1 ] )
                        or_values = [ float( or_values[j] ) for j in range( len( or_values ) ) ]
                        #
                        # here we quickly call the associated demand
                        cap_indices_rail = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ][ this_set_type_initial ] ) if x == str( params['tech_train_fre'] ) ]
                        cap_values_rail = deepcopy( stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices_rail[0]:cap_indices_rail[-1]+1 ] )
                        cap_values_rail = [ float( cap_values_rail[j] ) for j in range( len( cap_values_rail ) ) ]
                        demand_values_norail = [ demand_values[j] - cap_values_rail[j] for j in range( len( demand_values ) ) ]
                        #
                        apply_these_shares = relative_tkm_to_demand_nonrail[ capacity_variables[capvar] ][ group_tech_FREIGHT_HEA[g] ]
                        new_cap_values = [ apply_these_shares[n]*demand_values_norail[n]/or_values[n] for n in range( len( time_range_vector ) ) ]
                        new_cap_values_rounded = [ round(elem, params['round_#']) for elem in new_cap_values ]
                        #
                        stable_scenarios[ scenario_list[s] ][ capacity_variables[capvar] ]['value'][ cap_indices[0]:cap_indices[-1]+1 ] = deepcopy( new_cap_values_rounded )

        #########################################################################################
        ### BLOCK 6: modify demand ###
        
        # Waste
        if params['Use_Waste_B1']:
            if scenario_list[s] in list(set(base_configuration_waste['Scenario'].tolist())):            
                ''' MODIFYING *base_configuration_waste* '''
                this_scenario_df = deepcopy(base_configuration_waste.loc[base_configuration_waste['Scenario'].isin([scenario_list[s]])])

                # Reset index for the filtered DataFrame and drop the old index
                this_scenario_df.reset_index(drop=True, inplace=True)
                
                this_scenario_df_indices_list = this_scenario_df.index.tolist()         
                #
                for i in range(len(this_scenario_df_indices_list)):
                    #
                    set_item = this_scenario_df['Set'].tolist()[i]
                    #
                    waste_Params_built_in = this_scenario_df['Built-in Parameter-Set'].tolist()[i]
                    waste_Params_reference = this_scenario_df['Reference'].tolist()[i]
                    waste_Params_method = this_scenario_df['Method'].tolist()[i]
                    waste_Params_setIndex = this_scenario_df['Set_Index'].tolist()[i]
                    #
                    param_list = this_scenario_df['Parameter'].tolist()[i].split(' ; ')
                    param_str = this_scenario_df['Parameter'].tolist()[i]

                    for p in range(len(param_list)):
                        this_scenario_set_param_df = this_scenario_df.loc[(this_scenario_df['Parameter'] == param_str) & (this_scenario_df['Set'] == set_item)]
                        #                    
                        if waste_Params_method == 'Exact': # only modify demands and capacities if this is turned on
                            value_list_new_vals = []
                            this_index = this_scenario_set_param_df.index.tolist()[0]
                            for y in range(len(time_range_vector)):
                                value_list_new_vals.append(this_scenario_set_param_df.loc[this_index, time_range_vector[y]])
                            #
                            # OBTAINED VALUES, NOW PRINTING
                            if waste_Params_built_in == 'YES':
                                if waste_Params_setIndex not in ['e']:                                  
                                    this_param_indices = [i for i, x in enumerate(stable_scenarios[scenario_list[s]][param_list[p]][waste_Params_setIndex]) if x == set_item]
                                    value_list = deepcopy(stable_scenarios[scenario_list[s]][param_list[p]]['value'][this_param_indices[0]:this_param_indices[-1]+1])
                                    value_list = [float(e) for e in value_list]
                                    #
                                    new_value_list = [round(value_list_new_vals[e], params['round_#']) for e in range(len(value_list))]
                                    stable_scenarios[scenario_list[s]][param_list[p]]['value'][this_param_indices[0]:this_param_indices[-1]+1] = deepcopy(new_value_list)
                                
                                else:
                                
                                    for ite in range(len(value_list_new_vals)):
                                        stable_scenarios[scenario_list[s]][param_list[p]]['r'].append(params['coun_initial'])
                                        stable_scenarios[scenario_list[s]][param_list[p]]['e'].append(set_item)
                                        stable_scenarios[scenario_list[s]][param_list[p]]['y'].append(str(time_range_vector[ite]))
                                        stable_scenarios[scenario_list[s]][param_list[p]]['value'].append(value_list_new_vals[ite])
        
        # IPPU
        if scenario_list[s] in list(set(base_configuration_ippu['Scenario'].tolist())):            
            ''' MODIFYING *base_configuration_waste* '''
            this_scenario_df = deepcopy(base_configuration_ippu.loc[base_configuration_ippu['Scenario'].isin([scenario_list[s]])])

            # Reset index for the filtered DataFrame and drop the old index
            this_scenario_df.reset_index(drop=True, inplace=True)
            
            this_scenario_df_indices_list = this_scenario_df.index.tolist()         
            #
            for i in range(len(this_scenario_df_indices_list)):
                #
                set_item = this_scenario_df['Set'].tolist()[i]
                #
                ippu_Params_built_in = this_scenario_df['Built-in Parameter-Set'].tolist()[i]
                ippu_Params_reference = this_scenario_df['Reference'].tolist()[i]
                ippu_Params_method = this_scenario_df['Method'].tolist()[i]
                ippu_Params_setIndex = this_scenario_df['Set_Index'].tolist()[i]
                #
                param_list = this_scenario_df['Parameter'].tolist()[i].split(' ; ')
                param_str = this_scenario_df['Parameter'].tolist()[i]

                for p in range(len(param_list)):
                    this_scenario_set_param_df = this_scenario_df.loc[(this_scenario_df['Parameter'] == param_str) & (this_scenario_df['Set'] == set_item)]
                    #                    
                    if ippu_Params_method == 'Exact': # only modify demands and capacities if this is turned on
                        value_list_new_vals = []
                        this_index = this_scenario_set_param_df.index.tolist()[0]
                        for y in range(len(time_range_vector)):
                            value_list_new_vals.append(this_scenario_set_param_df.loc[this_index, time_range_vector[y]])
                        #
                        # OBTAINED VALUES, NOW PRINTING
                        if ippu_Params_built_in == 'YES':
                            if ippu_Params_setIndex not in ['e']:
                                this_param_indices = [i for i, x in enumerate(stable_scenarios[scenario_list[s]][param_list[p]][ippu_Params_setIndex]) if x == set_item]
                                value_list = deepcopy(stable_scenarios[scenario_list[s]][param_list[p]]['value'][this_param_indices[0]:this_param_indices[-1]+1])
                                value_list = [float(e) for e in value_list]
                                #
                                new_value_list = [round(value_list_new_vals[e], 4) for e in range(len(value_list))]
                                stable_scenarios[scenario_list[s]][param_list[p]]['value'][this_param_indices[0]:this_param_indices[-1]+1] = deepcopy(new_value_list)
                            else:
                                for ite in range(len(value_list_new_vals)):
                                    stable_scenarios[scenario_list[s]][param_list[p]]['r'].append('RD')
                                    stable_scenarios[scenario_list[s]][param_list[p]]['e'].append(set_item)
                                    stable_scenarios[scenario_list[s]][param_list[p]]['y'].append(str(time_range_vector[ite]))
                                    stable_scenarios[scenario_list[s]][param_list[p]]['value'].append(value_list_new_vals[ite])

        
        
        
        
        # Transport
        if params['Use_Transport_B1']:
            if scenario_list[s] in list(set(base_configuration_transport_elasticity['Scenario'].tolist())):            
                ''' MODIFYING *base_configuration_transport_elasticity* '''
                this_scenario_df = deepcopy( base_configuration_transport_elasticity.loc[ base_configuration_transport_elasticity['Scenario'].isin( [ scenario_list[s] ] ) ] )
                #
                set_list_group_dict = params['set_list_group_dict']
                #
                set_list_group = list( set( this_scenario_df[ 'Set' ].tolist() ) )
                #
                unique_demand_list = []
                unique_demand = {}
                #
                elasticity_Params_built_in_all = list(set(this_scenario_df['Built-in Parameter-Set'].tolist()))[0]
                elasticity_Params_reference_all = list(set(this_scenario_df['Reference'].tolist()))[0]
                elasticity_Params_method_all = list(set(this_scenario_df['Method'].tolist()))[0]
                elasticity_Params_setIndex_all = list(set(this_scenario_df['Set_Index'].tolist()))[0]
                #
                for l0 in range( len( set_list_group ) ):
                    this_scenario_tech_df = this_scenario_df.loc[ this_scenario_df['Set'] == set_list_group[l0] ]
                    set_list = set_list_group_dict[set_list_group[l0]]
                    #
                    for l in range( len( set_list ) ):
                        param_list = list( set( this_scenario_tech_df[ 'Parameter' ].tolist() ) )
                        for p in range( len( param_list ) ):
                            this_scenario_tech_param_df = this_scenario_tech_df.loc[ this_scenario_tech_df['Parameter'] == param_list[p] ]
                            #
                            elasticity_Params_built_in  = this_scenario_tech_param_df['Built-in Parameter-Set'].tolist()[0]
                            elasticity_Params_reference = this_scenario_tech_param_df['Reference'].tolist()[0]
                            elasticity_Params_method    = this_scenario_tech_param_df['Method'].tolist()[0]
                            elasticity_Params_setIndex    = this_scenario_tech_param_df['Set_Index'].tolist()[0]
                            #
                            if elasticity_Params_method == params['exact']: # only modify demands and capacities if this is turned on
                                value_list_num = []
                                this_index = this_scenario_tech_param_df.index.tolist()[0]
                                for y in range( len( time_range_vector ) ):
                                    value_list_num.append( this_scenario_tech_param_df.loc[ this_index, time_range_vector[y] ] )
                                #
                                if set_list_group[l0] == params['pass']:
                                    value_list_mult = [value_list_num[i]/Total_Demand[i] for i in range(len(value_list_num))]
                                elif set_list_group[l0] == params['fre']:
                                    value_list_mult = [value_list_num[i]/Total_Demand_Fre[i] for i in range(len(value_list_num))]
                                #
                                # OBTAINED VALUES, NOW PRINTING
                                if elasticity_Params_built_in == 'YES':
                                    this_param_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ param_list[p] ][ elasticity_Params_setIndex ] ) if x == str( set_list[l] ) ]
                                    value_list = deepcopy( stable_scenarios[ scenario_list[s] ][ param_list[p] ]['value'][ this_param_indices[0]:this_param_indices[-1]+1 ] )
                                    value_list = [float(e) for e in value_list]
                                    #
                                    if set_list[l] not in unique_demand_list:
                                        unique_demand_list.append(set_list[l])
                                        unique_demand.update({set_list[l]:deepcopy(value_list)})
                                    #
                                    new_value_list = [ round( value_list[e]*value_list_mult[e], params['round_#'] ) for e in range(len(value_list)) ]
                                    stable_scenarios[ scenario_list[s] ][ param_list[p] ]['value'][ this_param_indices[0]:this_param_indices[-1]+1 ] = deepcopy( new_value_list )
                

        # This is necessary to lose the restrictions we already defined
        unique_max_cap_techs = list( set( stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) )
        unique_low_lim_techs = list( set( stable_scenarios[ scenario_list[s] ][ 'TotalTechnologyAnnualActivityLowerLimit' ][ 't' ] ) )
        ignore_techs = params['ignore_techs']
        rewrite_techs_maxcap = \
            [e for e in unique_max_cap_techs if
             params['tr'] in e[:2] and params['in'] not in e and e not in ignore_techs]
        rewrite_techs_lowlim = \
            [e for e in unique_low_lim_techs if
             params['tr'] in e[:2] and params['in'] not in e and 'IMP' not in e and e not in ignore_techs]
        '''
        # Here we need to adjust the max capacities based on the adjusted demand
        for tech_gr in Fleet_Groups_techs_2_dem:
            use_dem = Fleet_Groups_techs_2_dem[tech_gr]

            dem_set_range_indices = [ i for i, x in enumerate( stable_scenarios_freeze[ scenario_list[ s ] ][ 'SpecifiedAnnualDemand' ][ 'f' ] ) if x == str( use_dem ) ]
            dem_value_list_fr = stable_scenarios_freeze[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ dem_set_range_indices[0]:dem_set_range_indices[-1]+1 ]
            dem_value_list_fr = [ float(dem_value_list_fr[j]) for j in range( len( dem_value_list_fr ) ) ]

            dem_set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'SpecifiedAnnualDemand' ][ 'f' ] ) if x == str( use_dem ) ]
            dem_value_list = stable_scenarios[ scenario_list[s] ][ 'SpecifiedAnnualDemand' ]['value'][ dem_set_range_indices[0]:dem_set_range_indices[-1]+1 ]
            dem_value_list = [ float(dem_value_list[j]) for j in range( len( dem_value_list ) ) ]

            print(use_dem)
            print(dem_value_list_fr)
            print(dem_value_list)

            cap_vars = [ 'TotalAnnualMaxCapacity', 'TotalTechnologyAnnualActivityLowerLimit' ]
            for cp in cap_vars:
                if cp == 'TotalAnnualMaxCapacity':
                    mult_constant = 1.01
                else:
                    mult_constant = 0.99
                group_set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ cp ][ 't' ] ) if x == str( tech_gr ) ]
                group_value_list = stable_scenarios[ scenario_list[s] ][ cp ]['value'][ group_set_range_indices[0]:group_set_range_indices[-1]+1 ]
                new_group_value_list = [
                    float(gv) * (1/float(d_fr)) * float(d_adj) * mult_constant
                    for gv, d_fr, d_adj in zip(
                    group_value_list, dem_value_list_fr, dem_value_list)]

                if cp == 'TotalAnnualMaxCapacity':
                    print(tech_gr)
                    print([i/j for i, j in zip(dem_value_list, dem_value_list_fr)])
                    print('\n')

                stable_scenarios[ scenario_list[s] ][ cp ]['value'][ group_set_range_indices[0]:group_set_range_indices[-1]+1 ] = deepcopy(new_group_value_list)

        print('Check up until here 1')
        sys.exit()
        '''   
        # Let's adjust existing restrictions to new capacities (transport sector);
        # First, max capacity:
        for t in range(len(rewrite_techs_maxcap)):
            this_set = rewrite_techs_maxcap[t]
            group_tech_set = Fleet_Groups_inv[ this_set ]

            group_set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) if x == str( group_tech_set ) ]
            group_value_list = stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['value'][ group_set_range_indices[0]:group_set_range_indices[-1]+1 ]
            group_value_list = [ float( group_value_list[j] ) for j in range( len( group_value_list ) ) ]

            set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) if x == str( this_set ) ]

            freeze_group_set_range_indices = [ i for i, x in enumerate( stable_scenarios_freeze[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) if x == str( group_tech_set ) ]
            freeze_group_value_list = stable_scenarios_freeze[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['value'][ freeze_group_set_range_indices[0]:freeze_group_set_range_indices[-1]+1 ]
            freeze_group_value_list = [ float( freeze_group_value_list[j] ) for j in range( len( freeze_group_value_list ) ) ]

            freeze_set_range_indices = [ i for i, x in enumerate( stable_scenarios_freeze[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) if x == str( this_set ) ]
            freeze_value_list = stable_scenarios_freeze[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['value'][ freeze_set_range_indices[0]:freeze_set_range_indices[-1]+1 ]
            freeze_value_list = [ float( freeze_value_list[j] ) for j in range( len( freeze_value_list ) ) ]

            value_list = [1.01*group_value_list[i]*freeze_value_list[i]/freeze_group_value_list[i] for i in range(len(group_value_list))]
            value_list_rounded = [ round(elem, params['round_#']) for elem in value_list ]
            stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['value'][ set_range_indices[0]:set_range_indices[-1]+1 ] = deepcopy(value_list_rounded)

        # Second, lower limit:
        for t in range(len(rewrite_techs_lowlim)):
            this_set = rewrite_techs_lowlim[t]
            group_tech_set = Fleet_Groups_inv[ this_set ]

            group_set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalTechnologyAnnualActivityLowerLimit' ][ 't' ] ) if x == str( group_tech_set ) ]
            group_value_list = stable_scenarios[ scenario_list[s] ][ 'TotalTechnologyAnnualActivityLowerLimit' ]['value'][ group_set_range_indices[0]:group_set_range_indices[-1]+1 ]
            group_value_list = [ float( group_value_list[j] ) for j in range( len( group_value_list ) ) ]

            set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalTechnologyAnnualActivityLowerLimit' ][ 't' ] ) if x == str( this_set ) ]

            freeze_group_set_range_indices = [ i for i, x in enumerate( stable_scenarios_freeze[ scenario_list[ s ] ][ 'TotalTechnologyAnnualActivityLowerLimit' ][ 't' ] ) if x == str( group_tech_set ) ]
            freeze_group_value_list = stable_scenarios_freeze[ scenario_list[s] ][ 'TotalTechnologyAnnualActivityLowerLimit' ]['value'][ freeze_group_set_range_indices[0]:freeze_group_set_range_indices[-1]+1 ]
            freeze_group_value_list = [ float( freeze_group_value_list[j] ) for j in range( len( freeze_group_value_list ) ) ]

            freeze_set_range_indices = [ i for i, x in enumerate( stable_scenarios_freeze[ scenario_list[ s ] ][ 'TotalTechnologyAnnualActivityLowerLimit' ][ 't' ] ) if x == str( this_set ) ]
            freeze_value_list = stable_scenarios_freeze[ scenario_list[s] ][ 'TotalTechnologyAnnualActivityLowerLimit' ]['value'][ freeze_set_range_indices[0]:freeze_set_range_indices[-1]+1 ]
            freeze_value_list = [ float( freeze_value_list[j] ) for j in range( len( freeze_value_list ) ) ]

            value_list = [0.999*group_value_list[i]*freeze_value_list[i]/freeze_group_value_list[i] for i in range(len(group_value_list))]
            value_list_rounded = [ round(elem, params['round_#']) for elem in value_list ]
            stable_scenarios[ scenario_list[s] ][ 'TotalTechnologyAnnualActivityLowerLimit' ]['value'][ set_range_indices[0]:set_range_indices[-1]+1 ] = deepcopy(value_list)

        ### BLOCK 7: define the *vehicle/ind_tech fuel tech* composition for each scenario ###
        # NOTE: check whether the scenario is contained in the system dictionary:
        if scenario_list[s] in list( set( list( adoption_params.keys() ) ) ):
            applicable_sets = []
            for a_set in range( len( list( adoption_params[ scenario_list[s] ].keys() ) ) ):
                this_set = list( adoption_params[ scenario_list[s] ].keys() )[ a_set ]
                applicable_sets.append( this_set )

            # Now we perform the adjustment of the system, after having modified the demand:
            for a_set in range( len( applicable_sets ) ):

                this_set = applicable_sets[ a_set ]
                this_sector = adoption_params[ scenario_list[s] ][ this_set ][ 'Sector' ]
                #
                this_param = adoption_params[ scenario_list[s] ][ this_set ][ 'Restriction_Type' ]
                #
                this_type = adoption_params[ scenario_list[s] ][ this_set ][ params['type'] ]
                #
                if this_param == params['min__max'] and this_sector == params['trans']:
                    cap_vars = [ 'TotalAnnualMaxCapacity', 'TotalTechnologyAnnualActivityLowerLimit' ]
                if this_param == params['min__max'] and this_sector == params['indus']:
                    cap_vars = [ 'TotalTechnologyAnnualActivityUpperLimit', 'TotalTechnologyAnnualActivityLowerLimit' ]
                if this_param == params['min']:
                    cap_vars = [ 'TotalTechnologyAnnualActivityLowerLimit' ]
                if this_param == params['max'] and this_sector == params['trans']:
                    cap_vars = [ 'TotalAnnualMaxCapacity']
                if this_param == params['max'] and this_sector == params['indus']:
                    cap_vars = [ 'TotalTechnologyAnnualActivityUpperLimit']
                #########################################################################################
                # Here, let's select the groups:
                if this_sector == params['trans']:
                    group_tech_set = Fleet_Groups_inv[ this_set ]
                #########################################################################################
                if this_type == params['linear']:
                    y_ini  = adoption_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 0 ] # This is not used / remove in future versions.
                    v_first_decade_year = adoption_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 1 ]
                    v_sec_decade_year = adoption_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 2 ]
                    v_final_year = adoption_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 3 ]
                    #
                    known_years_ini = [e for e in range(params['year_first_range'], int(y_ini)+1)]
                    known_years = known_years_ini + [params['first_decade_year'], params['sec_decade_year'], params['final_year'] ]
                    # First decade year
                    try:
                        float_v_first_decade_year = float(v_first_decade_year)
                    except Exception:
                        float_v_first_decade_year = params['interp']
                    # Second decade year
                    try:
                        float_v_sec_decade_year = float(v_sec_decade_year)
                    except Exception:
                        float_v_sec_decade_year = params['interp'] 
                    #Final year
                    try:
                        float_v_final_year = float(v_final_year)
                    except Exception:    
                        float_v_final_year = params['interp']
                    #
                    ### if this_sector == params['indus']:
                    # Here we must know the initial share (extracting the group values):
                    group_set_range_indices = [ i for i, x in enumerate(stable_scenarios[scenario_list[s]]['TotalTechnologyAnnualActivityLowerLimit']['t']) if x == str(group_tech_set)]
                    group_value_list = stable_scenarios[scenario_list[s]]['TotalTechnologyAnnualActivityLowerLimit']['value'][group_set_range_indices[0]:group_set_range_indices[-1]+1]
                    group_value_list = [ float( group_value_list[j] ) for j in range( len( group_value_list ) ) ]

                    this_set_range_indices = [ i for i, x in enumerate(stable_scenarios[scenario_list[s]]['TotalTechnologyAnnualActivityLowerLimit']['t']) if x == str(this_set)]
                    if len( this_set_range_indices ) != 0:
                        value_list = stable_scenarios[scenario_list[s]]['TotalTechnologyAnnualActivityLowerLimit']['value'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ]
                        value_list = [ float( value_list[j] ) for j in range( len( value_list ) ) ]
                        known_value_ini = []
                        for y in range( len( time_range_vector ) ):
                            if time_range_vector[y] <= y_ini:
                                known_value_ini.append(value_list[y]/group_value_list[y])
                                #
                            #
                        #
                    #
                    else:  # If there is no prior restriction, we star from zero.
                        known_value_ini = []
                        for y in range( len( time_range_vector ) ):
                            if time_range_vector[y] <= y_ini:
                                known_value_ini.append(0)
                                #
                            #
                        #
                    #
                    known_values = known_value_ini + [float_v_first_decade_year, float_v_sec_decade_year, float_v_final_year ]
                    #
                    x_coord_tofill, xp_coord_known, yp_coord_known = [], [], []
                    for y in range( len( time_range_vector ) ):
                        not_known_e = True
                        if time_range_vector[y] <= y_ini:
                            xp_coord_known.append( float(y) )
                            yp_coord_known.append( float(known_value_ini[y]) )
                            not_known_e = False
                        if v_first_decade_year != params['interp'] and time_range_vector[y] == params['first_decade_year']:
                            xp_coord_known.append( float(y) )
                            yp_coord_known.append( float(v_first_decade_year) )
                            not_known_e = False
                        if v_sec_decade_year != params['interp'] and time_range_vector[y] == params['sec_decade_year']:
                            xp_coord_known.append( float(y) )
                            yp_coord_known.append( float(v_sec_decade_year) )
                            not_known_e = False
                        if v_final_year != params['interp'] and time_range_vector[y] == params['final_year']:
                            xp_coord_known.append( float(y) )
                            yp_coord_known.append( float(v_final_year) )
                            not_known_e = False
                        if not_known_e == True:
                            x_coord_tofill.append( float(y) )
                    #
                    y_coord_filled = list( np.interp( x_coord_tofill, xp_coord_known, yp_coord_known ) )
                    interpolated_values = []
                    for coord in range( len( time_range_vector ) ):
                        if coord in xp_coord_known:
                            value_index = xp_coord_known.index(coord)
                            interpolated_values.append( float( yp_coord_known[value_index] ) )
                        elif coord in x_coord_tofill:
                            value_index = x_coord_tofill.index(coord)
                            interpolated_values.append( float( y_coord_filled[value_index] ) )
                    #
                    adoption_shift = interpolated_values
                    #
                #
                if this_type == params['logistic']:
                    R_base_year = adoption_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 0 ] # This is not used / remove in future versions.
                    R_final_year = adoption_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 1 ]
                    L =     adoption_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 2 ]
                    C =     adoption_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 3 ]
                    M =     adoption_params[ scenario_list[s] ][ this_set ][ 'Values' ][ 4 ]
                    #
                    r_final_year = 1/R_final_year
                    Q =     L/C - 1
                    k =     np.log( (r_final_year-1)/Q )/(M-params['final_year'])
                    #
                    shift_years = [ n for n in range( Initial_Year_of_Uncertainty+1,params['final_year']+1 ) ]
                    shift_year_counter = 0
                    adoption_shift = []
                    #
                    for t in range( len( time_range_vector ) ):
                        if time_range_vector[t] > Initial_Year_of_Uncertainty:
                            x = int( shift_years[shift_year_counter] )
                            adoption_shift.append( generalized_logistic_curve(x, L, Q, k, M))
                            shift_year_counter += 1
                        else:
                            adoption_shift.append( 0.0 )
                #########################################################################################
                #
                for cp in range( len( cap_vars ) ):
                    #
                    if cap_vars[cp] in params['cap_vars_cond']:
                        security_multiplier_factor = 1.001 # 1.0001
                    elif cap_vars[cp] == 'TotalTechnologyAnnualActivityLowerLimit':
                        security_multiplier_factor = 0.999 # 0.9999
                    #
                    group_set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ cap_vars[cp] ][ 't' ] ) if x == str( group_tech_set ) ]
                    group_value_list = stable_scenarios[ scenario_list[s] ][ cap_vars[cp] ]['value'][ group_set_range_indices[0]:group_set_range_indices[-1]+1 ]
                    group_value_list = [ float( group_value_list[j] ) for j in range( len( group_value_list ) ) ]
                    #**********************************************************************
                    if this_type == params['linear']:
                        if 999 in known_values:
                            indices_999 = [ i for i, x in enumerate( known_values ) if x != float( 999 ) ]
                            max_known_years = known_years[ max( indices_999 ) ]
                            for y in range( len( time_range_vector ) ):
                                if cap_vars[cp] == 'TotalAnnualMaxCapacity' and time_range_vector[y] > max_known_years:
                                    adoption_shift[y] = 1
                                elif cap_vars[cp] == 'TotalTechnologyAnnualActivityLowerLimit' and time_range_vector[y] > max_known_years:
                                    adoption_shift[y] = 0

                    #**********************************************************************
                    new_value_list = []
                    for n in range( len( time_range_vector ) ):
                        new_value_list.append( security_multiplier_factor*adoption_shift[n]*group_value_list[n] )
                        new_value_list_rounded = [ round(elem, params['round_#']) for elem in new_value_list ]
                    #
                    this_set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ cap_vars[cp] ][ 't' ] ) if x == str( this_set ) ]
                    if len( this_set_range_indices ) != 0:
                        stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['value'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ] = deepcopy(new_value_list_rounded)
                        # print('happens??', this_set, this_sector, scenario_list[s], cap_vars[cp])
                    else: # this means we need to add the data to the dictionary
                        # print('happens??', this_set, this_sector, scenario_list[s], cap_vars[cp])
                        for y in range( len( time_range_vector ) ):
                            stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['r'].append( params['coun_initial'] )
                            stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['t'].append( this_set )
                            stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['y'].append( str( time_range_vector[y] ) )
                            stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['value'].append( new_value_list_rounded[y] )
            #
            # This section makes the code have coherent restrictions for transport:
            if scenario_list[s] != params['BAU']:
                cap_vars = params['cap_vars']
                for cp in range( len( cap_vars ) ):
                    all_sets = list( set( stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['t'] ) )
                    all_sets.sort()
                    ref_dict = deepcopy( stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ] )
                    stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['r'] = []
                    stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['t'] = []
                    stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['y'] = []
                    stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['value'] = []
                    for t in range( len( all_sets ) ):
                        ###if all_sets[t] in applicable_sets or all_sets[t][:2] != params['tr]:
                        if ((all_sets[t] in applicable_sets or all_sets[t][:2] != params['tr']) or
                            (all_sets[t] in rewrite_techs_lowlim and all_sets[t] not in applicable_sets and cap_vars[cp] == 'TotalTechnologyAnnualActivityLowerLimit') or
                            (all_sets[t] in rewrite_techs_maxcap and all_sets[t] not in applicable_sets and cap_vars[cp] == 'TotalAnnualMaxCapacity')):
                            this_set_range_indices = [ i for i, x in enumerate( ref_dict[ 't' ] ) if x == str( all_sets[t] ) ]
                            stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['r'] += deepcopy( ref_dict['r'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ] )
                            stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['t'] += deepcopy( ref_dict['t'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ] )
                            stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['y'] += deepcopy( ref_dict['y'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ] )
                            stable_scenarios[ scenario_list[s] ][ cap_vars[ cp ] ]['value'] += deepcopy( ref_dict['value'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ] )

        ### BLOCK 8: Adjust all paramters for distance change ### // we do not apply this to BAU // here the occupancy rate AND the modal shift were already adjusted
        # parameters_involved = ['CapitalCost','FixedCost','ResidualCapacity','TotalAnnualMaxCapacity','TotalTechnologyAnnualActivityLowerLimit']
        parameters_involved = params['parameters_involved']
        #
        if scenario_list[s] != params['BAU']: # we proceed to adjust first the parameters and corresponding sets affected by distance, then the demands.
            for p in range(len(parameters_involved)):
                this_parameter = parameters_involved[p]
                if p == 0:
                    dem_adj_public = True
                    dem_adj_private = True
                    dem_adj_heafre = True
                    # dem_adj_medfre = True
                    dem_adj_ligfre = True
        
                for a_set in range( len( transport_group_sets ) ):
                    # Distance is a parameter that is not in Osemosys directly, but is implicit in multiple parameters.
                    # Now, we proceed to estimate the change in the curve of distance across futures. Note that this modification does not depend on the baseline values.
                    #
                    this_set_type_initial = S_DICT_sets_structure['initial'][ S_DICT_sets_structure['set'].index('TECHNOLOGY') ]
                    #
                    this_group_set = transport_group_sets[a_set] # these are GROUP TECHS, please do not forget
                    #
                    base_distance = Reference_driven_distance[ ref_scenario ][this_group_set]
                    new_distance = Reference_driven_distance[ scenario_list[s] ][this_group_set]

                    # the demand adjustment is done first:
                    this_demand_set = Fleet_Groups_techs_2_dem[ this_group_set ]
                    if ((this_demand_set == params['tra_dem_pri'] and dem_adj_private is True) or
                        (this_demand_set == params['tra_dem_pub'] and dem_adj_public is True) or
                        (this_demand_set == params['tra_dem_hea'] and dem_adj_heafre is True) or
                        # (this_demand_set == params['tra_dem_med'] and dem_adj_medfre is True) or
                        (this_demand_set == params['tra_dem_lig'] and dem_adj_ligfre is True)):
        
                        # let's extract rail capacity to adjust this apropiately
                        try:
                            train_pass_capacity_indices = [ i for i, x in enumerate( stable_scenarios[scenario_list[s]][ 'TotalAnnualMaxCapacity' ][ 't' ] ) if x == str( params['tech_train'] ) ]
                            train_pass_capacity_values = stable_scenarios[scenario_list[s]][ 'TotalAnnualMaxCapacity' ]['value'][ train_pass_capacity_indices[0]:train_pass_capacity_indices[-1]+1 ]
                        except Exception:
                            train_pass_capacity_values = [0 for i in range(len(time_range_vector))]

                        if Fleet_Groups_techs_2_dem[ this_group_set ] == params['tra_dem_pub'] and scenario_list[ s ] == params['NDP']:
                            subtract_list = [float(train_pass_capacity_values[j]) for j in range(len(train_pass_capacity_values))]
                        else:
                            subtract_list = [0 for j in range(len(train_pass_capacity_values))]
        
                        demand_indices = [ i for i, x in enumerate( stable_scenarios[scenario_list[s]][ 'SpecifiedAnnualDemand' ][ 'f' ] ) if x == str( this_demand_set ) ]
                        demand_list = deepcopy( stable_scenarios[scenario_list[s]][ 'SpecifiedAnnualDemand' ]['value'][ demand_indices[0]:demand_indices[-1]+1 ] )
                        demand_list = [float(i) for i in demand_list]
                        #
                        new_value_list = []
                        for n in range( len( demand_list ) ):
                            if n < index_change_year:
                                new_value_list.append( demand_list[n] )
                            else:
                                new_value_list.append( (demand_list[n]-subtract_list[n])*new_distance[n]/base_distance[n] )
                                new_value_list[-1] += subtract_list[n]
                        new_value_list_rounded = [ round(elem, params['round_#']) for elem in new_value_list ]
                        #
                        stable_scenarios[scenario_list[s]][ 'SpecifiedAnnualDemand' ]['value'][ demand_indices[0]:demand_indices[-1]+1 ] = deepcopy(new_value_list_rounded)
                        #                                       
                        if this_demand_set == params['tra_dem_pri']:
                            demand_indices = [ i for i, x in enumerate( stable_scenarios[scenario_list[s]][ 'SpecifiedAnnualDemand' ][ 'f' ] ) if x == str( params['tra_non_mot'] ) ]
                            demand_list = deepcopy( stable_scenarios[scenario_list[s]][ 'SpecifiedAnnualDemand' ]['value'][ demand_indices[0]:demand_indices[-1]+1 ] )
                            #
                            new_value_list = []
                            for n in range( len( demand_list ) ):
                                if n < index_change_year:
                                    new_value_list.append( demand_list[n] )
                                else:
                                    new_value_list.append( demand_list[n]*new_distance[n]/base_distance[n] )
                            new_value_list_rounded = [ round(elem, params['round_#']) for elem in new_value_list ]
                            #
                            stable_scenarios[scenario_list[s]][ 'SpecifiedAnnualDemand' ]['value'][ demand_indices[0]:demand_indices[-1]+1 ] = deepcopy(new_value_list_rounded)
                            dem_adj_private = False
                            #
                        if this_demand_set == params['tra_dem_pub']:
                            dem_adj_public = False
                            #
                        if this_demand_set == params['tra_dem_hea']:
                            dem_adj_heafre = False
                            #
                        # if this_demand_set == params['tra_dem_med']:
                            # dem_adj_medfre = False
                            #
                        if this_demand_set == params['tra_dem_lig']:
                            dem_adj_ligfre = False

                    # group-tech capacity adjustments are done second:
                    if (params['techs_'] in this_group_set and this_parameter in params['cap_vars_cond_2'] ): # THIS CONDITION WILL ALWAYS BE TRUE
                        #
                        this_set_range_indices = [ i for i, x in enumerate( stable_scenarios[scenario_list[s]][ this_parameter ][ this_set_type_initial ] ) if x == str( this_group_set ) ]
                        #
                        if len(this_set_range_indices) != 0:
                            # The obtained distance change is used appropiately from here on (...)
                            # We must proceed with the group set IF the parameter is adequate:
                            value_list = deepcopy( stable_scenarios[scenario_list[s]][ this_parameter ]['value'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ] )
                            value_list = [ float(value_list[j]) for j in range( len(value_list) ) ]
                            #
                            new_value_list = []
                            for n in range( len( value_list ) ):
                                if n < index_change_year:
                                    new_value_list.append( value_list[n] )
                                else:
                                    new_value_list.append( value_list[n]*new_distance[n]/base_distance[n] )
                            #
                            new_value_list_rounded = [ round(elem, params['round_#']) for elem in new_value_list ]
                            stable_scenarios[scenario_list[s]][ this_parameter ]['value'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ] = deepcopy( new_value_list_rounded )
        
                            if this_group_set == params['tech_he_fre'] and scenario_list[s] != params['BAU']:  # here we need to adjust the capacity of freight rail for consistency
                                tsri_frerail = [ i for i, x in enumerate( stable_scenarios[scenario_list[s]][ this_parameter ][ this_set_type_initial ] ) if x == str( params['tech_train_fre'] ) ]
                                value_list_frerail = deepcopy( stable_scenarios[scenario_list[s]][ this_parameter ]['value'][ tsri_frerail[0]:tsri_frerail[-1]+1 ] )
                                value_list_frerail = [ float(value_list_frerail[j]) for j in range( len(value_list_frerail) ) ]
                                #
                                new_value_list_frerail = []
                                for n in range( len( value_list_frerail ) ):
                                    new_value_list_frerail.append( value_list_frerail[n]*new_distance[n]/base_distance[n] )
                                #
                                new_value_list_frerail_rounded = [ round(elem, params['round_#']) for elem in new_value_list_frerail ]
                                stable_scenarios[scenario_list[s]][ this_parameter ]['value'][ tsri_frerail[0]:tsri_frerail[-1]+1 ] = deepcopy( new_value_list_frerail_rounded )

                    '''
                    # Me must act upon the parameters that get affected by the distance change. We must the read the scenario and extract a list of the technologies that require the change.
                    for g in range( len( transport_group_sets ) ):
                        #
                        this_group_set = transport_group_sets[g]
                        #
                        base_distance = Reference_driven_distance[ ref_scenario ][this_group_set]
                        new_distance = Reference_driven_distance[ scenario_list[s] ][this_group_set]
                        #
                    '''
                    # specific-tech capacity adjustments are done third:
                    specific_techs = Fleet_Groups[ this_group_set ]
                    # Let us attack the 02 techs:
                    for n in range( len( specific_techs ) ):
                        #
                        tech_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ this_parameter ][ 't' ] ) if x == str( specific_techs[n] ) ]
                        if len( tech_indices ) != 0:
                            tech_values = deepcopy( stable_scenarios[ scenario_list[ s ] ][ this_parameter ]['value'][ tech_indices[0]:tech_indices[-1]+1 ] )
                            tech_values = [ float( tech_values[j] ) for j in range( len( tech_values ) ) ]
                            #
                            if params['cost'] not in this_parameter:
                                adjustment_factor = \
                                    [new_distance[z]/base_distance[z] for z in range(len(tech_values))]
                            else:
                                adjustment_factor = \
                                    [(new_distance[z]/base_distance[z])**(-1) for z in range(len(tech_values))]
                            tech_values_new = [ tech_values[z]*(adjustment_factor[z]) for z in range(len(tech_values))]
                            #
                            tech_values_new_rounded = [ round(elem, params['round_#']) for elem in tech_values_new ]
                            #
                            stable_scenarios[ scenario_list[ s ] ][ this_parameter ]['value'][ tech_indices[0]:tech_indices[-1]+1 ] = deepcopy( tech_values_new_rounded )
        #***********************************************************************************************
        #if scenario_list[s] != params['BAU']:
        for g in range( len( transport_group_sets ) ): # this actually avoids the maxcap issue
            this_group_set = transport_group_sets[g]
            this_set_range_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ][ this_set_type_initial ] ) if x == str( this_group_set ) ]
            value_list = deepcopy( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ]['value'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ] )
            value_list = [ 1.01*float(value_list[j]) for j in range( len(value_list) ) ]
            #
            new_value_list_rounded = [ round(elem, params['round_#']) for elem in value_list ]
            new_value_list_rounded = deepcopy( interp_max_cap( new_value_list_rounded ) )
            #
            stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMaxCapacity' ]['value'][ this_set_range_indices[0]:this_set_range_indices[-1]+1 ] = deepcopy( new_value_list_rounded )
        #***********************************************************************************************
        #'''
        # Let's define rules to limit the adoption of technology in the early years:
        unique_residual_cap_techs = list( set( stable_scenarios[ scenario_list[s] ][ 'ResidualCapacity' ][ 't' ] ) )
        unique_max_cap_techs = list( set( stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) )
        unique_low_lim_techs = list( set( stable_scenarios[ scenario_list[s] ][ 'TotalTechnologyAnnualActivityLowerLimit' ][ 't' ] ) )
        #
        for g in range( len( transport_group_sets ) ): # this actually avoids the maxcap issue
            this_group_set = transport_group_sets[g]
            this_tech_set_list = Fleet_Groups[ this_group_set ]
            #
            this_g_tech_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ][ 't' ] ) if x == str( this_group_set ) ]
            this_g_tech_values = stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['value'][ this_g_tech_indices[0]:this_g_tech_indices[-1]+1 ]
            #
            for t in range( len( this_tech_set_list ) ):
                this_tech = this_tech_set_list[t]
                #
                if this_tech in unique_residual_cap_techs and this_tech not in unique_max_cap_techs:
                    this_res_tech_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[s] ][ 'ResidualCapacity' ][ 't' ] ) if x == str( this_tech ) ]
                    this_res_tech_values = stable_scenarios[ scenario_list[s] ][ 'ResidualCapacity' ]['value'][ this_res_tech_indices[0]:this_res_tech_indices[-1]+1 ]
                    this_res_tech_values = [ float( this_res_tech_values[j] ) for j in range( len( this_res_tech_values ) ) ]
                    
                    ### print( scenario_list[s], this_tech, 'here1', len(scenario_list) )
                    
                    #
                    for y in range( len( time_range_vector ) ):
                        stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['r'].append( params['coun_initial'] )
                        stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['t'].append( this_tech )                            
                        stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['y'].append( str( time_range_vector[y] ) )
                        if time_range_vector[y] <= Initial_Year_of_Uncertainty:
                            stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['value'].append( this_g_tech_values[y] )
                        else:
                            stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['value'].append( params['top_maxcap'] ) # this 9999 is a top that is hardly ever reached

                elif this_tech not in unique_residual_cap_techs and this_tech not in unique_max_cap_techs and this_tech not in unique_low_lim_techs: # or ('LPG' in this_tech and 'He' in this_group_set):
                    #
                    ### print( '*', scenario_list[s], this_tech, 'here2', len(scenario_list) )
                    
                    for y in range( len( time_range_vector ) ):
                        stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['r'].append( params['coun_initial'] )
                        stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['t'].append( this_tech )                            
                        stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['y'].append( str( time_range_vector[y] ) )
                        if time_range_vector[y] <= Initial_Year_of_Uncertainty:
                            stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['value'].append( 0 )
                        else:
                            stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['value'].append( params['top_maxcap'] ) # this 9999 is a top that is hardly ever reached
        #

        unique_min_cap_techs = list(set(stable_scenarios[scenario_list[s]]['TotalAnnualMinCapacity']['t']))
        for t in range( len( unique_min_cap_techs ) ):
            this_tech = unique_min_cap_techs[t]
            
            this_min_tech_indices = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMinCapacity' ][ 't' ] ) if x == str( this_tech ) ]
            this_min_tech_values = deepcopy( stable_scenarios[ scenario_list[ s ] ][ 'TotalAnnualMinCapacity' ]['value'][ this_min_tech_indices[0]:this_min_tech_indices[-1]+1 ] )
            this_min_tech_values = [ float( this_min_tech_values[j] ) for j in range( len( this_min_tech_values ) ) ]
            
            for y in range( len( time_range_vector ) ):
                stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['r'].append( params['coun_initial'] )
                stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['t'].append( this_tech )                            
                stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['y'].append( str( time_range_vector[y] ) )
                stable_scenarios[ scenario_list[s] ][ 'TotalAnnualMaxCapacity' ]['value'].append( str( this_min_tech_values[y] ) )

        '''
        ### BLOCK 9: makes the code have coherent restrictions for industry // applies for any scenario:
        '''
        '''
        #***********************************************************************************************
        ### BLOCK 10: Integrate biofuels according to RECOPE / *we must use a new interpolation function*.
        #  Note: the biodiesel blend will affect all the sectors, not only transport.
        if scenario_list[s] != params['BAU']: # GENERALIZE THIS LATER.
            #
            Diesel_Techs = params['Diesel_Techs']
            #
            Diesel_Techs_Emissions = params['Diesel_Techs_Emissions']
            #
            Gasoline_Techs = params['Gasoline_Techs']
            #
            Gasoline_Techs_Emissions = params['Gasoline_Techs_Emissions']
            #
            # 10.1. Let us adjust for diesel and biodiesel: (remember to use  *time_range_vector*)
            for n in range( len( Diesel_Techs ) ):
                #
                if params['Use_Blend_Shares_B1']:
                    Blend_Shares[ scenario_list[s] ].update( { Diesel_Techs[n]:{} } )
                #
                for n2 in range( len( Diesel_Techs_Emissions[ Diesel_Techs[n] ] ) ):
                    this_tech_emission_indices_diesel = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'EmissionActivityRatio' ][ 't' ] ) if x == Diesel_Techs[n] ]
                    this_tech_emission_indices_emission = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'EmissionActivityRatio' ][ 'e' ] ) if x == Diesel_Techs_Emissions[ Diesel_Techs[n] ][n2] ]
                    #
                    r_index = set(this_tech_emission_indices_diesel) & set(this_tech_emission_indices_emission)
                    this_tech_emission_indices = list( r_index )
                    this_tech_emission_indices.sort()
                    #
                    value_list = deepcopy( stable_scenarios[ scenario_list[ s ] ][ 'EmissionActivityRatio' ]['value'][ this_tech_emission_indices[0]:this_tech_emission_indices[-1]+1 ] )
                    value_list = [ float(value_list[j]) for j in range( len(value_list) ) ]
                    #
                    start_blend_point = [params['year_start_blend_point_2'], 1]
                    final_blend_point = [params['year_final_blend_point_2'], 5]
                    new_value_list_rounded, biofuel_shares = interpolation_blend( start_blend_point, 'None', final_blend_point, value_list, time_range_vector )
                    #
                    stable_scenarios[ scenario_list[ s ] ][ 'EmissionActivityRatio' ]['value'][ this_tech_emission_indices[0]:this_tech_emission_indices[-1]+1 ] = deepcopy( new_value_list_rounded )
                    #
                    if params['Use_Blend_Shares_B1']:
                        Blend_Shares[ scenario_list[s] ][  Diesel_Techs[n] ].update( { Diesel_Techs_Emissions[ Diesel_Techs[n] ][n2]: deepcopy( biofuel_shares ) } )
                    #
                #
            #
            # 10.2. Let us adjust for gasoline and ethanol: (remember to use  *time_range_vector*)
            for n in range( len( Gasoline_Techs ) ):
                #
                if params['Use_Blend_Shares_B1']:
                    Blend_Shares[ scenario_list[s] ].update( { Gasoline_Techs[n]:{} } )
                #
                for n2 in range( len( Gasoline_Techs_Emissions[ Gasoline_Techs[n] ] ) ):
                    this_tech_emission_indices_gasoline = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'EmissionActivityRatio' ][ 't' ] ) if x == Gasoline_Techs[n] ]
                    this_tech_emission_indices_emission = [ i for i, x in enumerate( stable_scenarios[ scenario_list[ s ] ][ 'EmissionActivityRatio' ][ 'e' ] ) if x == Gasoline_Techs_Emissions[ Gasoline_Techs[n] ][n2] ]
                    #
                    r_index = set(this_tech_emission_indices_gasoline) & set(this_tech_emission_indices_emission)
                    this_tech_emission_indices = list( r_index )
                    this_tech_emission_indices.sort()
                    #
                    value_list = deepcopy( stable_scenarios[ scenario_list[ s ] ][ 'EmissionActivityRatio' ]['value'][ this_tech_emission_indices[0]:this_tech_emission_indices[-1]+1 ] )
                    value_list = [ float(value_list[j]) for j in range( len(value_list) ) ]
                    #
                    start_blend_point = [params['year_start_blend_point'], 2]
                    mid_blend_point = [params['year_mid_blend_point'], 4]
                    final_blend_point = [params['year_final_blend_point'], 8]
                    new_value_list_rounded, biofuel_shares = interpolation_blend( start_blend_point, mid_blend_point, final_blend_point, value_list, time_range_vector )
                    #
                    stable_scenarios[ scenario_list[ s ] ][ 'EmissionActivityRatio' ]['value'][ this_tech_emission_indices[0]:this_tech_emission_indices[-1]+1 ] = deepcopy( new_value_list_rounded )
                    #
                    if params['Use_Blend_Shares_B1']:
                        Blend_Shares[ scenario_list[s] ][  Gasoline_Techs[n] ].update( { Gasoline_Techs_Emissions[ Gasoline_Techs[n] ][n2]:biofuel_shares } )
        '''
        # BLOCK 11: Integrate emission restriction
        if scenario_list[s] in list(set(base_configuration_emi_res['Scenario'].tolist())):
            this_scenario_df = deepcopy(base_configuration_emi_res.loc[base_configuration_emi_res['Scenario'].isin([scenario_list[s]])])
            set_list = list(set(this_scenario_df['Emission'].tolist()))
            for l in range(len(set_list)):
                this_scenario_tech_df = this_scenario_df.loc[this_scenario_df['Emission'] == set_list[l]]
                # Assuming you are only interested in the first row based on some condition, as your code suggests.
                this_index = this_scenario_tech_df.index.tolist()[0]

                # Retrieve the row corresponding to `this_index` and only keep columns that are in `time_range_vector`.
                # Then, convert the resulting Series to a list.
                value_list = this_scenario_tech_df.loc[this_index, time_range_vector].tolist()

                for n in range(len(time_range_vector)):
                    stable_scenarios[scenario_list[s]]['AnnualEmissionLimit']['r'].append(params['coun_initial'])
                    stable_scenarios[scenario_list[s]]['AnnualEmissionLimit']['e'].append(set_list[l])
                    stable_scenarios[scenario_list[s]]['AnnualEmissionLimit']['y'].append(time_range_vector[n])
                    stable_scenarios[scenario_list[s]]['AnnualEmissionLimit']['value'].append(value_list[n])
                                            
    print('  finished the standard adjustment, now to processing')

    '''
    # We must create the *Based_On* parameters:
        * Affects:
            - base_configuration_electrical
            - base_configuration_E_and_D
    '''
    for sb in range( len( scenario_list_based ) ):
        if scenario_list_based[sb] != 'ref':
            this_scenario_name = scenario_list_all[sb]
            stable_scenarios.update( { this_scenario_name:{} } )
            stable_scenarios[ this_scenario_name ] = deepcopy( stable_scenarios[ scenario_list_based[sb] ] )
            #
            if params['Use_Blend_Shares_B1']:
                if scenario_list_based[sb] != params['BAU']:
                    Blend_Shares.update( { this_scenario_name:Blend_Shares[ scenario_list_based[sb] ] } )
                else:
                    Blend_Shares.update( { this_scenario_name:Blend_Shares[ params['NDP'] ] } )
            Reference_driven_distance.update( { this_scenario_name:Reference_driven_distance[ scenario_list_based[sb] ] } )
            Reference_occupancy_rate.update( { this_scenario_name:Reference_occupancy_rate[ scenario_list_based[sb] ] } )
            Reference_op_life.update( { this_scenario_name:Reference_op_life[ scenario_list_based[sb] ] } )
    #
    if params['Use_Blend_Shares_B1']:
        with open(params['Blend_Shares_0'], 'wb') as handle:
            pickle.dump(Blend_Shares, handle, protocol=pickle.HIGHEST_PROTOCOL)
        handle.close()

    #print('check')
    #sys.exit()

    scenario_list_all.sort()
    for sb in range( len( scenario_list_all ) ):
        this_scenario_name = scenario_list_all[sb]
        #
        #########################################################################################
        ''' MODIFYING *base_configuration_electrical* '''
        local_application_iter = ['All', this_scenario_name]
        for iter_scen in local_application_iter:
            this_scenario_df = deepcopy( base_configuration_electrical.loc[ base_configuration_electrical['Scenario'].isin( [ iter_scen ] ) ] )
            #
            set_list = list( set( this_scenario_df[ 'Tech_Set' ].tolist() ) )
            for l in range( len( set_list ) ):
                this_scenario_tech_df = this_scenario_df.loc[ this_scenario_df['Tech_Set'] == set_list[l] ]
                #
                param_list = list( set( this_scenario_tech_df[ 'Parameter' ].tolist() ) )
                for p in range( len( param_list ) ):
                    this_scenario_tech_param_df = this_scenario_tech_df.loc[ this_scenario_tech_df['Parameter'] == param_list[p] ]
                    #
                    electrical_Params_built_in      = this_scenario_tech_param_df['Built-in Parameter-Set'].tolist()[0].replace(' ','')
                    electrical_Params_linear        = this_scenario_tech_param_df['Linear'].tolist()[0].replace(' ','')
                    electrical_Params_exactYears    = this_scenario_tech_param_df['Exact_Years'].astype(str).tolist()[0].replace(' ','')
                    electrical_Params_exactValues   = this_scenario_tech_param_df['Exact_Values'].astype(str).tolist()[0].replace(' ','')
                    electrical_Params_initialYear   = this_scenario_tech_param_df['y_ini'].tolist()[0]
                    electrical_Params_mYears        = this_scenario_tech_param_df['Milestone_Years'].astype(str).tolist()[0].replace(' ','')
                    electrical_Params_mValues       = this_scenario_tech_param_df['Milestone_Value'].astype(str).tolist()[0].replace(' ','')
                    electrical_Params_Unit          = this_scenario_tech_param_df['Unit'].tolist()[0].replace(' ','')
                    electrical_Params_methods       = this_scenario_tech_param_df['Method'].tolist()[0].replace(' ','')
                    #
                    electrical_Params_secMultiplier = this_scenario_tech_param_df['Security_Multiplier'].tolist()[0]
                    #
                    electrical_Params_exactYears    = electrical_Params_exactYears.split( ';' )
                    electrical_Params_exactValues   = electrical_Params_exactValues.split( ';' )
                    electrical_Params_mYears        = electrical_Params_mYears.split( ';' )
                    electrical_Params_mValues       = electrical_Params_mValues.split( ';' )
                    electrical_Params_methods       = electrical_Params_methods.split( ';' )
                    #
                    # Let us create the value vector when the parameter is not built in:
                    if electrical_Params_built_in == 'NO':
                        if params['inter_escal'] in electrical_Params_methods and params['fix_last'] in electrical_Params_methods and params['intact'] not in electrical_Params_exactYears:
                            last_electrical_Params_exactValues = deepcopy( electrical_Params_exactValues[-1] )
    
                            for i in range( int( electrical_Params_exactYears[-1] )+1, int( electrical_Params_mYears[0] ) ):
                                electrical_Params_exactYears += [ i ]
                                electrical_Params_exactValues += [ last_electrical_Params_exactValues ]
        
                            known_years     = [ float(e) for e in electrical_Params_exactYears ] + [ float(e) for e in electrical_Params_mYears ]
                            known_values     = [ float(e) for e in electrical_Params_exactValues ] + [ float(e) for e in electrical_Params_mValues ]
                            value_list = linear_interpolation_time_series( time_range_vector, known_years, known_values )
    
                        if params['interpo'] in electrical_Params_methods and params['fix_last'] in electrical_Params_methods and params['intact'] not in electrical_Params_exactYears:
                            known_years     = [ float(e) for e in electrical_Params_exactYears ] + [ float(e) for e in electrical_Params_mYears ]
                            known_values     = [ float(e) for e in electrical_Params_exactValues ] + [ float(e) for e in electrical_Params_mValues ]
                            value_list = linear_interpolation_time_series( time_range_vector, known_years, known_values )
    
                        if params['fix_ind'] in electrical_Params_methods and params['intact'] not in electrical_Params_exactYears:
                            known_years     = [ float(e) for e in electrical_Params_exactYears ] +  [ electrical_Params_initialYear ]
                            known_values     = [ float(e) for e in electrical_Params_exactValues ] + [ float( electrical_Params_mValues[0] ) ]
                            value_list = linear_interpolation_time_series( time_range_vector, known_years, known_values )
    
                        if params['write'] in electrical_Params_methods:
                            value_list = [ round( e*electrical_Params_secMultiplier, params['round_#'] ) for e in value_list ]
                            for y in range( len( time_range_vector ) ):
                                stable_scenarios[ this_scenario_name ][ param_list[p] ]['r'].append( params['coun_initial'] )
                                stable_scenarios[ this_scenario_name ][ param_list[p] ]['t'].append( set_list[l] )                            
                                stable_scenarios[ this_scenario_name ][ param_list[p] ]['y'].append( str( time_range_vector[y] ) )
                                stable_scenarios[ this_scenario_name ][ param_list[p] ]['value'].append( float( value_list[y] ) )
    
                    # Let us modify the value vector if it already exists:
                    if electrical_Params_built_in == 'YES':
                        this_param_indices = [ i for i, x in enumerate( stable_scenarios[ this_scenario_name ][ param_list[p] ][ 't' ] ) if x == str( set_list[l] ) ]
                        this_param_values = deepcopy( stable_scenarios[ this_scenario_name ][ param_list[p] ]['value'][ this_param_indices[0]:this_param_indices[-1]+1 ] )
                        #
                        if params['interpo'] in electrical_Params_methods and params['intact'] in electrical_Params_exactYears:
                            known_years = [ y for y in time_range_vector if y <= electrical_Params_initialYear ]
                            known_values = [ float( this_param_values[y] ) for y in range( len(time_range_vector) ) if time_range_vector[y] <= electrical_Params_initialYear ]
                            known_years += [ float(e) for e in electrical_Params_mYears ]
                            known_values += [ float(e) for e in electrical_Params_mValues ]
                            value_list = linear_interpolation_time_series( time_range_vector, known_years, known_values )
                        #
                        if params['interpo'] in electrical_Params_methods and params['fix_last'] in electrical_Params_methods and params['intact'] not in electrical_Params_exactYears:
                            known_years     = [ float(e) for e in electrical_Params_exactYears ] + [ float(e) for e in electrical_Params_mYears ]
                            known_values     = [ float(e) for e in electrical_Params_exactValues ] + [ float(e) for e in electrical_Params_mValues ]
                            value_list = linear_interpolation_time_series( time_range_vector, known_years, known_values )
                        #
                        if params['overwrite'] in electrical_Params_methods:
                            value_list = [ round( e, params['round_#'] ) for e in value_list ]
                            stable_scenarios[ this_scenario_name ][ param_list[p] ]['value'][ this_param_indices[0]:this_param_indices[-1]+1 ] = deepcopy( value_list )

        #########################################################################################
        ''' MODIFYING *base_configuration_smartGrid* '''
        this_scenario_df = deepcopy( base_configuration_smartGrid.loc[ base_configuration_smartGrid['Scenario'].isin( [ this_scenario_name ] ) ] )
        #
        set_list = list( set( this_scenario_df[ 'Tech_Set' ].tolist() ) )
        for l in range( len( set_list ) ):
            this_scenario_tech_df = this_scenario_df.loc[ this_scenario_df['Tech_Set'] == set_list[l] ]
            #
            if 'Contains' in set_list[l]:
                tech_indicator_list_raw = set_list[l].replace( ' ','' ).split( ':' )[1] # This is a string with a list after a colon
                tech_indicator_list = tech_indicator_list_raw.split( ';' ) # This could be a list of strings separated by semicolon
                for tech_indicator in tech_indicator_list:
                    tech_list = [ e for e in all_techs_list if tech_indicator in e and 'IMP' not in e]
            else:
                tech_list = [ set_list[l] ]
            #
            for t in range( len( tech_list ) ):
                this_set = tech_list[t]
                param_list = list( set( this_scenario_tech_df[ 'Parameter' ].tolist() ) )
                for p in range( len( param_list ) ):
                    this_scenario_tech_param_df = this_scenario_tech_df.loc[ this_scenario_tech_df['Parameter'] == param_list[p] ]
                    #
                    smartGrid_Params_built_in  = this_scenario_tech_param_df['Built-in Parameter-Set'].tolist()[0]
                    smartGrid_Params_reference = this_scenario_tech_param_df['Reference'].tolist()[0]
                    smartGrid_Params_method    = this_scenario_tech_param_df['Method'].tolist()[0]
                    #------------------------------
                    # TAKE REFERENCE FOR INTERPOLATION, IF AVAILABLE
                    if smartGrid_Params_built_in == 'YES':
                        this_param_indices = [ i for i, x in enumerate( stable_scenarios[ this_scenario_name ][ param_list[p] ][ 't' ] ) if x == str( this_set ) ]
                        this_param_values = deepcopy( stable_scenarios[ this_scenario_name ][ param_list[p] ]['value'][ this_param_indices[0]:this_param_indices[-1]+1 ] )
                    #------------------------------
                    # CALCULATE THE VALUE LIST
                    if smartGrid_Params_method == params['exat_mult']: # use *time_range_horizon*
                        value_list = []
                        this_index = this_scenario_tech_param_df.index.tolist()[0]
                        for y in range( len( time_range_vector ) ):
                            value_list.append( float( this_param_values[y] )*float( this_scenario_tech_param_df.loc[ this_index, time_range_vector[y] ] ) )
                    #
                    if smartGrid_Params_method == params['exact']: # use *time_range_horizon*
                        value_list = []
                        this_index = this_scenario_tech_param_df.index.tolist()[0]
                        for y in range( len( time_range_vector ) ):
                            value_list.append( this_scenario_tech_param_df.loc[ this_index, time_range_vector[y] ] )
                    #
                    if smartGrid_Params_method == params['linear']: # use *time_range_horizon*
                        smartGrid_Params_initialYear  = this_scenario_tech_param_df['y_ini'].tolist()[0]
                        smartGrid_Params_mYears = this_scenario_tech_param_df['Milestone_Years'].astype(str).tolist()[0].replace(' ','')
                        smartGrid_Params_mValue = this_scenario_tech_param_df['Milestone_Value'].astype(str).tolist()[0].replace(' ','')
                        #
                        smartGrid_Params_mYears = smartGrid_Params_mYears.split( ';' )
                        smartGrid_Params_mValue = smartGrid_Params_mValue.split( ';' )
                        #
                            
                        known_years = [ y for y in time_range_vector if y <= smartGrid_Params_initialYear ]
                        known_values = [ float( this_param_values[y] ) for y in range( len(time_range_vector) ) if time_range_vector[y] <= smartGrid_Params_initialYear ]
                        known_years += [ float(e) for e in smartGrid_Params_mYears ]
                        known_values += [ float(e) for e in smartGrid_Params_mValue ]
                        value_list = linear_interpolation_time_series( time_range_vector, known_years, known_values )
                    #
                    if smartGrid_Params_method == 'Copy':
                                                     
                        this_param_indices = [ i for i, x in enumerate( stable_scenarios[ smartGrid_Params_reference ][ param_list[p] ][ 't' ] ) if x == str( this_set ) ]                             
                        value_list = deepcopy( stable_scenarios[ smartGrid_Params_reference ][ param_list[p] ]['value'][ this_param_indices[0]:this_param_indices[-1]+1 ] )
                    #------------------------------
                    # OBTAINED VALUES, NOW PRINTING
                    if smartGrid_Params_built_in == 'YES':
                        value_list = [round(e, params['round_#']) for e in value_list]
                        this_param_indices = [i for i, x in enumerate(stable_scenarios[this_scenario_name][param_list[p]]['t']) if x == str(this_set)]
                        stable_scenarios[this_scenario_name][param_list[p]]['value'][this_param_indices[0]:this_param_indices[-1]+1] = deepcopy(value_list)

        #########################################################################################
        ''' MODIFYING *base_configuration_E_and_D* '''
        this_scenario_df = deepcopy( base_configuration_E_and_D.loc[ base_configuration_E_and_D['Scenario'].isin( [ this_scenario_name ] ) ] )
        #
        set_list = list( set( this_scenario_df[ 'Set' ].tolist() ) )
        for l in range( len( set_list ) ):
            this_scenario_tech_df = this_scenario_df.loc[ this_scenario_df['Set'] == set_list[l] ]
            #
            param_list = list( set( this_scenario_tech_df[ 'Parameter' ].tolist() ) )
            for p in range( len( param_list ) ):
                this_scenario_tech_param_df = this_scenario_tech_df.loc[ this_scenario_tech_df['Parameter'] == param_list[p] ]
                #
                efficiency_Params_built_in  = this_scenario_tech_param_df['Built-in Parameter-Set'].tolist()[0]
                efficiency_Params_reference = this_scenario_tech_param_df['Reference'].tolist()[0]
                efficiency_Params_method    = this_scenario_tech_param_df['Method'].tolist()[0]
                efficiency_Params_setIndex    = this_scenario_tech_param_df['Set_Index'].tolist()[0]
                #
                if efficiency_Params_method == params['exact']: # use *time_range_horizon*
                    value_list = []
                    this_index = this_scenario_tech_param_df.index.tolist()[0]
                    for y in range( len( time_range_vector ) ):
                        value_list.append( this_scenario_tech_param_df.loc[ this_index, time_range_vector[y] ] )
                #
                if efficiency_Params_method == params['copy']:
                    this_param_indices = [ i for i, x in enumerate( stable_scenarios[ efficiency_Params_reference ][ param_list[p] ][ efficiency_Params_setIndex ] ) if x == str( set_list[l] ) ]
                    value_list = deepcopy( stable_scenarios[ efficiency_Params_reference ][ param_list[p] ]['value'][ this_param_indices[0]:this_param_indices[-1]+1 ] )
                #
                # OBTAINED VALUES, NOW PRINTING
                if efficiency_Params_built_in == 'YES':
                    value_list = [ round( e, params['round_#'] ) for e in value_list ]
                    this_param_indices = [ i for i, x in enumerate( stable_scenarios[ this_scenario_name ][ param_list[p] ][ efficiency_Params_setIndex ] ) if x == str( set_list[l] ) ]
                    stable_scenarios[ this_scenario_name ][ param_list[p] ]['value'][ this_param_indices[0]:this_param_indices[-1]+1 ] = deepcopy( value_list )

        #########################################################################################

    print('  finished, now to processing')
    # sys.exit()

    scenario_list = list( stable_scenarios.keys() ) # This applies for all other scenarios

    '''
    Control inputs:
    '''
    is_this_last_update = params['is_this_last_update']
    # is_this_last_update = False
    # is_this_last_update = False
    # generator_or_executor = 'None'
    generator_or_executor = params['generator_or_executor']
    # generator_or_executor = 'Generator'
    # generator_or_executor = 'Executor'
    #
    #
    #########################################################################################
    if is_this_last_update == True:
        #
        print('3: Let us store the inputs for later analysis.')
        basic_header_elements = params['basic_header_elements']
        #
        parameters_to_print = params['parameters_to_print_last']

        if params['parallel']:
            print('Entered Parallelization of control inputs')
            x = len(scenario_list)
            max_x_per_iter = params['max_x_per_iter'] # FLAG: This is an input
            y = x / max_x_per_iter
            y_ceil = math.ceil( y )
            #
            for n in range(0,y_ceil):
                n_ini = n*max_x_per_iter
                processes = []
                #
                if n_ini + max_x_per_iter <= x:
                    max_iter = n_ini + max_x_per_iter
                else:
                    max_iter = x
                #    
                for n2 in range( n_ini , max_iter ):
                    p = mp.Process(target=csv_printer_parallel, args=(n2, scenario_list, stable_scenarios, basic_header_elements, parameters_to_print, S_DICT_params_structure, params) )
                    processes.append(p)
                    p.start()
                #
                for process in processes:
                    process.join()

    #########################################################################################
    global scenario_list_print
    scenario_list_print = deepcopy( scenario_list ) # [params['BAU'],'OP15C',params['NDP']]
    # scenario_list_print = [params['BAU'],params['NDP'],'NDP_A','OP15C'] # FLAG: This is an input
    # scenario_list_print = [params['NDP']]

    # print('stop before printing')
    # sys.exit()

    if generator_or_executor == params['gen_or_exe_1'] or generator_or_executor == params['gen_or_exe_2']:
        #
        print('5: We have finished all manipulations of base scenarios. We will now print.')
        #
        print_adress = params['Executables']
        #
        if params['Use_Blend_Shares_B1']:
            packaged_useful_elements = [scenario_list_print, S_DICT_sets_structure, S_DICT_params_structure, list_param_default_value_params, list_param_default_value_value,
                                        print_adress, Reference_driven_distance,
                                        Fleet_Groups_inv, time_range_vector, Blend_Shares ]
        else:
            packaged_useful_elements = [scenario_list_print, S_DICT_sets_structure, S_DICT_params_structure, list_param_default_value_params, list_param_default_value_value,
                                        print_adress, Reference_driven_distance,
                                        Fleet_Groups_inv, time_range_vector ]
        #
        if params['parallel']:
            print('Entered Parallelization of .txt printing')
            x = len(scenario_list)
            max_x_per_iter = params['max_x_per_iter'] # FLAG: This is an input
            y = x / max_x_per_iter
            y_ceil = math.ceil( y )
            #
            for n in range(0,y_ceil):
                n_ini = n*max_x_per_iter
                processes = []
                #
                if n_ini + max_x_per_iter <= x:
                    max_iter = n_ini + max_x_per_iter
                else:
                    max_iter = x
                #    
                for n2 in range( n_ini , max_iter ):
                    p = mp.Process(target=function_C_mathprog, args=(n2, stable_scenarios, packaged_useful_elements, params) )
                    processes.append(p)
                    p.start()
                #
                for process in processes:
                    process.join()

    #########################################################################################
    if generator_or_executor == params['gen_or_exe_3'] or generator_or_executor == params['gen_or_exe_4']:
        #
        print('6: We have finished printing base scenarios. We must now execute.')
        #
        packaged_useful_elements = [Reference_driven_distance, Reference_occupancy_rate, Reference_op_life,
                                    Fleet_Groups_inv, time_range_vector,
                                    list_param_default_value_params, list_param_default_value_value, list_gdp_ref]
        #
        set_first_list(scenario_list_print, params)
        
        if params['parallel']:
            print('Entered Parallelization of model execution')
            x = len(first_list)
            max_x_per_iter = params['max_x_per_iter'] # FLAG: This is an input
            y = x / max_x_per_iter
            y_ceil = math.ceil( y )
            #
            for n in range(0,y_ceil):
                n_ini = n*max_x_per_iter
                processes = []
                #
                if n_ini + max_x_per_iter <= x:
                    max_iter = n_ini + max_x_per_iter
                else:
                    max_iter = x
                #    
                for n2 in range( n_ini , max_iter ):
                    p = mp.Process(target=main_executer, args=(n2, packaged_useful_elements, scenario_list_print, params) )
                    processes.append(p)
                    p.start()
                #
                for process in processes:
                    process.join()
            
        # This is for the linear version
        else:
            print('Started Linear Runs')
            for n in range( len( first_list ) ):
                main_executer(n, packaged_useful_elements, scenario_list_print, params)
    
    else:
        # Module to run test when mode is 'Generator'
        file_aboslute_address = os.path.abspath(params['B1_script'])
        file_config_plots_csvs = get_config_main_path(os.path.abspath(''),'config_plots')
        file_address = re.escape( file_aboslute_address.replace( params['B1_script'], '' ) )
        #
        str_start = params['start'] + file_address
        path_test_config_plots = os.path.join(file_config_plots_csvs, params['test_path'])
        str_tests = 'python -u ' + path_test_config_plots + ' ' + str(file_address) + ' 1' # last int is the ID tier
        os.system( str_start and str_tests )            
    
    # Delete log files when solver='cplex'
    if params['solver'] == 'cplex' and params['del_files']:
        shutil.os.remove('cplex.log')
        shutil.os.remove('clone1.log')
        shutil.os.remove('clone2.log')

    end_1 = time.time()   
    time_elapsed_1 = -start1 + end_1
    print( str( time_elapsed_1 ) + ' seconds /', str( time_elapsed_1/60 ) + ' minutes' )
    print('*: For all effects, we have finished the work of this script.')
    #########################################################################################